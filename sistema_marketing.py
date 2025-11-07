# -*- coding: utf-8 -*-
# SISTEMA MARKETING EXTREMEMAX
# Sistema profesional de gestión de marketing y seguimiento de clientes

import os, time
from pathlib import Path
from io import BytesIO
import base64
import re
import json
import streamlit as st
from openpyxl import Workbook, load_workbook
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Configuración de página
st.set_page_config(
    page_title="Marketing Extrememax",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado para diseño futurista (neón / glassmorphism)
st.markdown("""
<style>
    /* Importar fuentes modernas */
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700;800;900&family=Inter:wght@300;400;500;600;700;800&display=swap');
    
    * {
        font-family: 'Poppins', 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    }
    
    /* ===== VARIABLES DE COLOR ===== */
    :root {
        --primary-color: #00E5FF;   /* cian neón */
        --primary-dark: #7C3AED;    /* morado neón */
        --primary-light: #67E8F9;
        --accent-color: #22D3EE;
        --secondary-color: #0b0f17; /* fondo oscuro */
        --panel-color: rgba(255, 255, 255, 0.06);
        --glass-border: 1px solid rgba(255, 255, 255, 0.15);
        --text-dark: #e5e7eb;
        --text-light: #a3a3a3;
        --bg-white: #ffffff;
        --bg-light: #0f172a;
        --border-color: rgba(103, 232, 249, 0.25);
        --shadow-sm: 0 10px 20px rgba(0, 229, 255, 0.08);
        --shadow-md: 0 20px 40px rgba(124, 58, 237, 0.15);
        --shadow-lg: 0 30px 60px rgba(0, 0, 0, 0.45);
        --shadow-xl: 0 40px 80px rgba(0, 0, 0, 0.6);
    }
    
    /* ===== APLICACIÓN PRINCIPAL ===== */
    .stApp {
        background: #ffffff;
        background-attachment: fixed;
    }
    
    /* ===== CONTENEDOR PRINCIPAL ===== */
    .main .block-container {
        padding: 2rem;
        max-width: 1400px;
        background: #ffffff;
        border-radius: 20px;
        margin: 2rem 0;
        box-shadow: 0 10px 40px rgba(0,0,0,0.08);
        border: 1px solid #f0f0f0;
    }
    
    /* ===== HEADER PROFESIONAL ===== */
    h1 {
        color: #000000;
        font-family: 'Poppins', sans-serif;
        font-weight: 800;
        font-size: 2.75rem;
        margin-bottom: 1rem;
        text-shadow: 0 0 20px rgba(34,211,238,0.35);
        position: relative;
        padding-bottom: 1rem;
    }
    
    h1::after {
        content: '';
        position: absolute;
        bottom: 0;
        left: 0;
        width: 80px;
        height: 4px;
        background: linear-gradient(90deg, var(--primary-color), var(--primary-dark));
        border-radius: 2px;
    }
    
    h2 {
        color: #000000;
        font-family: 'Poppins', sans-serif;
        font-weight: 700;
        font-size: 2rem;
        margin-top: 2rem;
        margin-bottom: 1rem;
        padding-left: 1rem;
        border-left: 4px solid var(--primary-color);
    }
    
    h3 {
        color: var(--text-dark);
        font-family: 'Poppins', sans-serif;
        font-weight: 600;
        font-size: 1.5rem;
        margin-top: 1.5rem;
        margin-bottom: 0.75rem;
    }
    
    h4 {
        color: var(--text-dark);
        font-weight: 600;
        font-size: 1.25rem;
    }
    
    /* ===== MÉTRICAS PROFESIONALES ===== */
    [data-testid="stMetric"] {
        background: linear-gradient(135deg, #ffffff 0%, #f8f8f8 100%);
        border: 2px solid rgba(255,154,0,0.2);
        padding: 1.5rem;
        border-radius: 16px;
        box-shadow: var(--shadow-sm);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        position: relative;
        overflow: hidden;
    }
    
    [data-testid="stMetric"]::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        width: 4px;
        height: 100%;
        background: linear-gradient(180deg, var(--primary-color), var(--primary-dark));
    }
    
    [data-testid="stMetric"]:hover {
        transform: translateY(-5px);
        box-shadow: var(--shadow-md);
        border-color: rgba(103, 232, 249, 0.45);
    }
    
    [data-testid="stMetricValue"] {
        font-size: 2rem !important;
        font-weight: 700 !important;
        color: #000000 !important;
    }
    
    [data-testid="stMetricLabel"] {
        font-weight: 600 !important;
        color: var(--text-light) !important;
        font-size: 0.9rem !important;
    }
    
    /* ===== BOTONES PROFESIONALES ===== */
    .stButton > button {
        background: linear-gradient(135deg, #FF9A00 0%, #FF6D00 100%);
        border: none;
        border-radius: 12px;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        font-size: 1rem;
        color: #000000;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        box-shadow: 0 4px 15px rgba(255,154,0,0.4);
        cursor: pointer;
        text-transform: none;
    }
    
    /* Botones de tipo de evento - negros con texto blanco y pequeños */
    button[data-testid*="baseButton-secondary"][data-baseweb="button"]:nth-of-type(n+1),
    .stButton button:contains("ACTIVACION"),
    .stButton button:contains("CARRERA"),
    .stButton button:contains("PRESTAR") {
        background-color: #000000 !important;
        background: #000000 !important;
        color: #FFFFFF !important;
        border: 1px solid #333333 !important;
        border-radius: 4px !important;
        padding: 0.15rem 0.25rem !important;
        font-size: 0.55rem !important;
        font-weight: 600 !important;
        min-height: 32px !important;
        height: auto !important;
        line-height: 1.1 !important;
        box-shadow: none !important;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 12px 28px rgba(124,58,237,0.35);
        background: linear-gradient(135deg, var(--primary-color) 0%, var(--primary-dark) 100%);
    }
    
    .stButton > button:active {
        transform: translateY(0);
    }
    
    /* Descargar (botón): forzar texto negro y fondo blanco */
    .stDownloadButton > button,
    [data-testid="stDownloadButton"] button,
    [data-testid="stDownloadButton"] > div > button {
        color: #1a1a1a !important; /* negro para legibilidad */
        background: #ffffff !important;
        border: 2px solid var(--border-color) !important;
    }
    /* Asegurar que el texto interno del botón sea negro */
    [data-testid="stDownloadButton"] button span,
    [data-testid="stDownloadButton"] > div > button span,
    .stDownloadButton > button span,
    [data-testid="stDownloadButton"] a,
    [data-testid="stDownloadButton"] a span {
        color: #1a1a1a !important;
        text-shadow: none !important;
    }
    .stDownloadButton > button:hover,
    [data-testid="stDownloadButton"] button:hover,
    [data-testid="stDownloadButton"] > div > button:hover {
        background: #f8f9fa !important;
        box-shadow: var(--shadow-lg) !important;
    }
    
    /* ===== INPUTS ELEGANTES ===== */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div > select {
        background: var(--bg-white);
        border: 2px solid rgba(0, 0, 0, 0.1);
        border-radius: 12px;
        padding: 0.75rem 1rem;
        color: #000000 !important;
        transition: all 0.3s ease;
        font-size: 1rem;
    }
    
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus,
    .stSelectbox > div > div > select:focus {
        border-color: var(--primary-color);
        box-shadow: 0 0 0 4px rgba(255, 154, 0, 0.1);
        outline: none;
    }
    
    .stTextArea > div > div > textarea {
        background: var(--bg-white);
        border: 2px solid rgba(0, 0, 0, 0.1);
        border-radius: 12px;
        padding: 0.75rem 1rem;
        color: #000000 !important;
        transition: all 0.3s ease;
    }
    
    .stTextArea > div > div > textarea:focus {
        border-color: var(--primary-color);
        box-shadow: 0 0 0 4px rgba(255, 154, 0, 0.1);
        outline: none;
    }
    
    /* ===== SIDEBAR PROFESIONAL ===== */
    [data-testid="stSidebar"] {
        background: #ffffff;
        border-right: 1px solid #e5e7eb;
    }
    
    /* Texto del sidebar en blanco (sin afectar inputs y componentes blancos) */
    [data-testid="stSidebar"] .stMarkdown,
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] h4,
    [data-testid="stSidebar"] h5,
    [data-testid="stSidebar"] h6,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span,
    [data-testid="stSidebar"] div,
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] .streamlit-expanderHeader {
        color: #000000 !important;
    }

    /* Mantener textos de campos editables legibles en el sidebar */
    [data-testid="stSidebar"] input,
    [data-testid="stSidebar"] textarea,
    [data-testid="stSidebar"] select {
        color: #000000 !important;
        background: #ffffff !important;
    }

    /* EXCEPCIONES: componentes con fondo blanco dentro del sidebar deben tener texto negro */
    [data-testid="stSidebar"] .stFileUploader *,
    [data-testid="stSidebar"] .stFileUploader p,
    [data-testid="stSidebar"] .stFileUploader span,
    [data-testid="stSidebar"] .stFileUploader label,
    [data-testid="stSidebar"] .stFileUploader div {
        color: #1a1a1a !important;
    }

    /* Botón de descarga dentro del sidebar: texto negro forzado */
    [data-testid="stSidebar"] [data-testid="stDownloadButton"] button,
    [data-testid="stSidebar"] [data-testid="stDownloadButton"] button span,
    [data-testid="stSidebar"] [data-testid="stDownloadButton"] a,
    [data-testid="stSidebar"] [data-testid="stDownloadButton"] a span {
        color: #1a1a1a !important;
    }

    /* Placeholders de inputs en sidebar en gris oscuro para contraste */
    [data-testid="stSidebar"] input::placeholder,
    [data-testid="stSidebar"] textarea::placeholder {
        color: #4b5563 !important;
        opacity: 1;
    }
    
    [data-testid="stSidebar"] .stMarkdown h3 {
        color: #000000 !important;
        -webkit-text-fill-color: initial;
        background: none;
        border-bottom: 1px solid #e5e7eb;
    }
    
    /* ===== TABS PROFESIONALES ===== */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 12px;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, rgba(34,211,238,0.25), rgba(124,58,237,0.25));
        color: #e5e7eb;
        border: var(--glass-border);
    }
    
    /* ===== EXPANDERS ELEGANTES ===== */
    .streamlit-expanderHeader {
        background: linear-gradient(135deg, rgba(255,255,255,0.06), rgba(255,255,255,0.03));
        border: var(--glass-border);
        border-radius: 12px;
        padding: 1rem;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .streamlit-expanderHeader:hover {
        background: linear-gradient(135deg, rgba(103,232,249,0.08), rgba(124,58,237,0.08));
        border-color: rgba(103,232,249,0.45);
    }
    
    /* ===== MENSAJES PROFESIONALES ===== */
    .stSuccess {
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.1), rgba(16, 185, 129, 0.05));
        border-left: 4px solid #10b981;
        border-radius: 12px;
        padding: 1rem;
        box-shadow: var(--shadow-sm);
    }
    
    .stError {
        background: linear-gradient(135deg, rgba(239, 68, 68, 0.1), rgba(239, 68, 68, 0.05));
        border-left: 4px solid #ef4444;
        border-radius: 12px;
        padding: 1rem;
        box-shadow: var(--shadow-sm);
    }
    
    .stWarning {
        background: linear-gradient(135deg, rgba(255, 154, 0, 0.15), rgba(255, 154, 0, 0.05));
        border-left: 4px solid var(--primary-color);
        border-radius: 12px;
        padding: 1rem;
        box-shadow: var(--shadow-sm);
    }
    
    .stInfo {
        background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), rgba(59, 130, 246, 0.05));
        border-left: 4px solid #3b82f6;
        border-radius: 12px;
        padding: 1rem;
        box-shadow: var(--shadow-sm);
    }
    
    /* ===== PROGRESS BARS ===== */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, var(--primary-color), var(--primary-dark));
        border-radius: 10px;
    }
    
    /* ===== DATA FRAMES ===== */
    .stDataFrame {
        border-radius: 12px;
        overflow: hidden;
        box-shadow: var(--shadow-md);
    }
    
    /* ===== LABELS ===== */
    label {
        color: #000000 !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
        margin-bottom: 0.5rem;
    }
    
    /* ===== DIVIDERS ===== */
    hr {
        border: none;
        height: 2px;
        background: linear-gradient(90deg, transparent, var(--border-color), transparent);
        margin: 2rem 0;
    }
    
    /* ===== CHECKBOX ===== */
    .stCheckbox > label {
        color: var(--text-dark) !important;
        font-weight: 500 !important;
    }
    
    /* ===== RADIO BUTTONS ===== */
    .stRadio > div {
        /* Fondo del grupo de radios: más oscuro para sidebar */
        background: rgba(255, 255, 255, 0.06);
        border-radius: 12px;
        padding: 0.5rem;
    }
    
    .stRadio > div > label {
        border-radius: 10px;
        padding: 0.75rem 1rem;
        transition: all 0.3s ease;
    }
    
    .stRadio > div > label:hover {
        background: rgba(255, 154, 0, 0.15);
    }

    /* Forzar color de texto negro en opciones del menú del sidebar */
    [data-testid="stSidebar"] .stRadio * {
        color: #000000 !important;
    }
    [data-testid="stSidebar"] [data-baseweb="radio"] label,
    [data-testid="stSidebar"] [data-baseweb="radio"] div,
    [data-testid="stSidebar"] [data-baseweb="radio"] span {
        color: #000000 !important;
    }
    [data-testid="stSidebar"] [data-baseweb="radio"] {
        background: transparent !important;
    }
    
    /* ===== FOOTER ===== */
    footer {
        display: none;
    }
    
    .stApp > header {
        display: none;
    }
    
    /* ===== ANIMACIONES ===== */
    @keyframes fadeInUp {
        from {
            opacity: 0;
            transform: translateY(30px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    @keyframes pulse {
        0%, 100% {
            opacity: 1;
        }
        50% {
            opacity: 0.7;
        }
    }
    
    /* ===== SCROLLBAR PERSONALIZADA ===== */
    ::-webkit-scrollbar {
        width: 10px;
        height: 10px;
    }
    
    ::-webkit-scrollbar-track {
        background: var(--bg-light);
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb {
        background: linear-gradient(180deg, var(--primary-color), var(--primary-dark));
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(180deg, var(--primary-dark), var(--primary-color));
    }
    
    /* ===== TABLAS ===== */
    table {
        border-radius: 12px;
        overflow: hidden;
    }

    /* Texto global en negro para alto contraste */
    body, .stApp, .main .block-container, p, span, div {
        color: #000000 !important;
    }
    
    /* ===== SELECTBOX ===== */
    [data-baseweb="select"] {
        border-radius: 12px;
    }
    
    /* ===== DATE INPUT ===== */
    .stDateInput > div > div > input {
        border-radius: 12px;
        border: 2px solid rgba(0, 0, 0, 0.1);
        color: #000000 !important;
    }
    
    /* ===== Asegurar que todos los inputs tengan texto negro ===== */
    input[type="text"],
    input[type="number"],
    input[type="email"],
    input[type="tel"],
    input[type="date"],
    textarea,
    select {
        color: #000000 !important;
    }
    
    input::placeholder,
    textarea::placeholder {
        color: #9ca3af !important;
        opacity: 1 !important;
    }
    
    /* ===== FILE UPLOADER ===== */
    .stFileUploader > div {
        border-radius: 12px;
        border: 2px dashed var(--border-color);
        padding: 2rem;
        transition: all 0.3s ease;
    }
    
    .stFileUploader > div:hover {
        border-color: var(--primary-color);
        background: rgba(103, 232, 249, 0.08);
    }
</style>
""", unsafe_allow_html=True)

# ===== RUTA DEL EXCEL =====
EXCEL_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = "REGISTRO_MARKETING.xlsx"
EXCEL_PATH = Path(EXCEL_DIR) / EXCEL_FILE
EMPRESAS_FILE = "EMPRESAS.xlsx"
EMPRESAS_PATH = Path(EXCEL_DIR) / EMPRESAS_FILE
VENTAS_FILE = "VENTAS POR PRODUCTOS.xlsx"
VENTAS_PATH = Path(EXCEL_DIR) / VENTAS_FILE
WEIGHTS_FILE = Path(EXCEL_DIR) / "WEIGHTS_EVENTOS.json"
OLD_WEIGHTS_EXTREME = Path(EXCEL_DIR) / "WEIGHTS_ACTIVACION_EXTREME.json"
PRODUCTOS_FILE = Path(EXCEL_DIR) / "PRODUCTOS_EVENTOS.json"

def load_saved_products_for_type(event_type: str):
    """Cargar productos guardados para un tipo de evento"""
    try:
        if PRODUCTOS_FILE.exists():
            with open(PRODUCTOS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
            productos = data.get(event_type, [])
            return productos if isinstance(productos, list) else []
    except Exception:
        pass
    return []

def save_saved_products_for_type(event_type: str, productos_list: list):
    """Guardar productos para un tipo de evento"""
    try:
        data = {}
        if PRODUCTOS_FILE.exists():
            with open(PRODUCTOS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
        # Convertir productos a diccionarios serializables
        productos_serializables = []
        for prod in productos_list:
            productos_serializables.append({
                'codigo': str(prod.get('codigo', '')),
                'nombre': str(prod.get('nombre', '')),
                'categoria': str(prod.get('categoria', '')),
                'unidad': str(prod.get('unidad', 'pc')),
                'cantidad': float(prod.get('cantidad', 1.0)),
                'precio_unit': float(prod.get('precio_unit', 0.0)),
                'descuento': float(prod.get('descuento', 0.0)),
                'nota': str(prod.get('nota', ''))
            })
        data[event_type] = productos_serializables
        with open(PRODUCTOS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"Error guardando productos: {e}")
        return False

def load_saved_weights_for_type(event_type: str):
    try:
        # Estructura esperada: { "ACTIVACION EXTREME": {item: peso, ...}, "ACTIVACION PANTRO": {...}, ... }
        if WEIGHTS_FILE.exists():
            with open(WEIGHTS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
            pesos = data.get(event_type, {})
            return {str(k): float(v) for k, v in pesos.items()}
        # Backward compatibility: archivo antiguo solo para EXTREME
        if event_type == "ACTIVACION EXTREME" and OLD_WEIGHTS_EXTREME.exists():
            with open(OLD_WEIGHTS_EXTREME, "r", encoding="utf-8") as f:
                old = json.load(f)
            return {str(k): float(v) for k, v in (old or {}).items()}
    except Exception:
        pass
    return {}

def save_saved_weights_for_type(event_type: str, weights_dict: dict):
    try:
        data = {}
        if WEIGHTS_FILE.exists():
            with open(WEIGHTS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
        data[event_type] = {str(k): float(v) for k, v in weights_dict.items()}
        with open(WEIGHTS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False

# ===== FUNCIONES PARA CARGAR CLIENTES =====
def buscar_cliente_inteligente(cliente, search_term):
    """Búsqueda inteligente en múltiples campos y sin importar el orden"""
    if not search_term:
        return True
    
    search_lower = search_term.lower().strip()
    if not search_lower:
        return True
    
    # Buscar en ID
    if search_lower in cliente.get('id', '').lower():
        return True
    
    # Buscar en nombre fiscal
    if search_lower in cliente.get('nombre', '').lower():
        return True
    
    # Buscar en nombre comercial
    nombre_comercial = cliente.get('nombre_comercial', '')
    if nombre_comercial and search_lower in nombre_comercial.lower():
        return True
    
    # Buscar en identificación (RUC/Cédula)
    if cliente.get('identificacion') and search_lower in str(cliente.get('identificacion', '')).lower():
        return True
    
    # Buscar en teléfono
    if cliente.get('telefono') and search_lower in str(cliente.get('telefono', '')).lower():
        return True
    
    # Búsqueda en desorden (palabras separadas)
    # Busca si TODAS las palabras del término están en algún campo
    palabras = search_lower.split()
    if len(palabras) > 1:
        texto_completo = f"{cliente.get('nombre', '')} {cliente.get('nombre_comercial', '')} {cliente.get('id', '')} {cliente.get('identificacion', '')} {cliente.get('telefono', '')}".lower()
        if all(palabra in texto_completo for palabra in palabras):
            return True
    
    return False

def cargar_clientes():
    """Cargar lista de clientes desde EMPRESAS.xlsx"""
    clientes = []
    if EMPRESAS_PATH.exists():
        try:
            df = pd.read_excel(EMPRESAS_PATH)
            print(f"📊 Archivo leído: {len(df)} filas encontradas")
            
            for _, row in df.iterrows():
                cliente_id = str(row.get('id_empresa', '')).strip()
                nombre_comercial = str(row.get('nombre_comercial', '')).strip()
                nombre_fiscal = str(row.get('nombre_fiscal', '')).strip()
                # Usar nombre_fiscal como nombre principal
                nombre = nombre_fiscal if nombre_fiscal else nombre_comercial
                
                if cliente_id and cliente_id != 'nan' and nombre and nombre != 'nan':
                    clientes.append({
                        'id': cliente_id,
                        'nombre': nombre,
                        'nombre_comercial': nombre_comercial,
                        'identificacion': str(row.get('identificacion', '')).strip(),
                        'telefono': str(row.get('telefono1', '')).strip(),
                        'ciudad': str(row.get('ciudad', '')).strip(),
                        'provincia': str(row.get('provincia', '')).strip(),
                        'direccion': str(row.get('direccion1', '')).strip(),
                        'agente': str(row.get('agente_comercial', '')).strip()
                    })
            
            print(f"✅ {len(clientes)} clientes cargados exitosamente")
        except Exception as e:
            print(f"❌ Error al cargar clientes: {e}")
    else:
        print(f"⚠️ Archivo no encontrado: {EMPRESAS_PATH}")
    return clientes

def cargar_agentes():
    """Cargar lista de agentes únicos desde EMPRESAS.xlsx"""
    agentes = [""]  # Opción vacía para que sea opcional
    if EMPRESAS_PATH.exists():
        try:
            df = pd.read_excel(EMPRESAS_PATH)
            agentes_lista = df['agente_comercial'].dropna().unique().tolist()
            agentes.extend([str(a).strip() for a in agentes_lista if str(a).strip() and str(a).strip() != 'nan'])
            agentes = sorted(set([a for a in agentes if a]))  # Eliminar duplicados y ordenar
            print(f"✅ {len(agentes)-1} agentes cargados exitosamente")
        except Exception as e:
            print(f"❌ Error al cargar agentes: {e}")
    else:
        print(f"⚠️ Archivo no encontrado: {EMPRESAS_PATH}")
    return agentes

# ===== ENCABEZADOS =====
HEADERS = {
    "LETREROS": ["ID", "Cliente", "Tipo Solicitud", "Tipo de Letrero", "Medidas (cm)", "Costo", "Comercial/Agente", "Fecha", "Estado", "Fecha Envío Fabricación", "Fecha Entrega Cliente", "Observaciones", "Número de Guía"],
    "EVENTOS": ["ID", "Cliente", "Tipo", "Fecha", "Estado", "Descripcion", "Productos", "Comercial/Agente", "Número de Guía", "Observaciones Envío", "Productos Retornados", "Fecha Retorno", "Observaciones Retorno"],
    "PUBLICIDAD": ["ID", "Cliente", "Comercial/Vendedor", "Fecha", "Productos", "Observaciones"],
    "PERCHAS": ["ID", "Cliente", "Comercial/Vendedor", "Fecha", "Estado", "Cantidad de Compra", "Productos", "Observaciones", "Número de Guía", "Medio de Envío", "Fotos"],
    "COMERCIALES": ["ID", "Comercial/Vendedor", "Tipo Entrega", "Fecha", "Productos", "Cliente Destino", "Estado", "Observaciones", "Productos Entregados", "Fotos", "ID_Entrega_Original"],
    "INVENTARIO": ["ID", "Tipo", "Cantidad", "Fecha", "Estado", "Observaciones"],
    "INVENTARIO_VENDEDORES": ["Vendedor", "Codigo_Producto", "Cantidad"],
    "EXPOFERIAS": ["ID", "Cliente", "Tipo", "Fecha", "Estado", "Descripcion"]
}

# ===== FUNCIONES ROBUSTAS DE EXCEL =====
def safe_load_workbook(path, tries=5, wait=0.2):
    """Cargar workbook con reintentos optimizado"""
    last = None
    for i in range(tries):
        try:
            return load_workbook(path)
        except PermissionError as e:
            last = e
            if i < tries - 1:  # No esperar en el último intento
                time.sleep(wait)
    raise last if last else PermissionError("No se pudo abrir el Excel (bloqueado).")

def safe_save_workbook(wb, path: Path, tries=30, wait=0.3):
    """Guardar workbook con reintentos optimizado"""
    last = None
    for i in range(tries):
        try:
            wb.save(path)
            if i > 0:
                print(f"✅ Archivo guardado exitosamente en intento {i+1}")
            try:
                wb.close()
            except:
                pass
            return True
        except PermissionError as e:
            last = e
            if i % 5 == 0:  # Solo mostrar cada 5 intentos
                print(f"⚠️ Intento {i+1}/{tries} falló - OneDrive está bloqueando, reintentando...")
            if i < tries - 1:  # No esperar en el último intento
                time.sleep(wait)
        except Exception as e:
            print(f"❌ Error inesperado en intento {i+1}: {e}")
            last = e
            break

    # Si llegamos aquí, no se pudo guardar
    print(f"❌ No se pudo guardar después de {tries} intentos - OneDrive bloqueando el archivo")
    
    # Fallback: copia con timestamp
    stamp = time.strftime("%Y%m%d-%H%M%S")
    alt = path.with_name(f"{path.stem}_copia_{stamp}{path.suffix}")
    try:
        wb.save(alt)
        print(f"✅ Guardé una copia de seguridad: {alt.name}")
        try:
            wb.close()
        except:
            pass
        st.warning(f"⚠️ **El archivo principal está bloqueado.** Guardé una COPIA: `{alt.name}`")
        return True
    except Exception as e2:
        try:
            wb.close()
        except:
            pass
        print(f"❌ Error al guardar copia: {e2}")
        st.error(f"❌ **Error crítico**: No se pudo guardar ni el original ni la copia.\n\n{last}\n\n{e2}")
        return False

def ensure_workbook(path: Path):
    """Crear Excel si no existe"""
    if not path.exists():
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        for name, headers in HEADERS.items():
            ws = wb.create_sheet(name)
            ws.append(headers)
        
        # Crear hoja de BASE DE DATOS CLIENTES
        ws_clientes = wb.create_sheet("BASE DE DATOS CLIENTES")
        headers_clientes = ['ID', 'Nombre Fiscal', 'Nombre Comercial / Local', 'Identificación', 'Teléfono', 'Ciudad', 'Provincia', 'Dirección', 'Agente']
        ws_clientes.append(headers_clientes)
        
        safe_save_workbook(wb, path)
        return

def copiar_base_datos_clientes():
    """Copiar datos de EMPRESAS.xlsx a REGISTRO_MARKETING.xlsx en hoja BASE DE DATOS CLIENTES"""
    try:
        # Leer EMPRESAS.xlsx
        if not EMPRESAS_PATH.exists():
            return False
        
        df = pd.read_excel(EMPRESAS_PATH)
        
        # Abrir REGISTRO_MARKETING.xlsx
        ensure_workbook(EXCEL_PATH)
        wb = safe_load_workbook(EXCEL_PATH)
        
        # Crear o limpiar hoja BASE DE DATOS CLIENTES
        if "BASE DE DATOS CLIENTES" not in wb.sheetnames:
            ws = wb.create_sheet("BASE DE DATOS CLIENTES")
        else:
            ws = wb["BASE DE DATOS CLIENTES"]
        
        # Limpiar datos existentes (excepto headers)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        
        # Headers
        headers = ['ID', 'Nombre Fiscal', 'Nombre Comercial / Local', 'Identificación', 'Teléfono', 'Ciudad', 'Provincia', 'Dirección', 'Agente']
        
        # Escribir headers si no existen
        if ws.max_row == 0:
            ws.append(headers)
        
        # Copiar datos
        for _, row in df.iterrows():
            cliente_id = str(row.get('id_empresa', '')).strip()
            nombre_comercial = str(row.get('nombre_comercial', '')).strip()
            nombre_fiscal = str(row.get('nombre_fiscal', '')).strip()
            # Usar nombre_fiscal como nombre principal
            nombre = nombre_fiscal if nombre_fiscal else nombre_comercial
            
            if cliente_id and cliente_id != 'nan' and nombre and nombre != 'nan':
                ws.append([
                    cliente_id,
                    nombre_fiscal,
                    nombre_comercial,
                    str(row.get('identificacion', '')).strip(),
                    str(row.get('telefono1', '')).strip(),
                    str(row.get('ciudad', '')).strip(),
                    str(row.get('provincia', '')).strip(),
                    str(row.get('direccion1', '')).strip(),
                    str(row.get('agente_comercial', '')).strip()
                ])
        
        safe_save_workbook(wb, EXCEL_PATH)
        return True
    except Exception as e:
        print(f"Error al copiar base de datos: {e}")
        return False

@st.cache_data(ttl=30)  # Cache por 30 segundos
def next_id(sheet_name):
    """Generar próximo ID con caché"""
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_id = ws.max_row
            wb.close()
            return max_id
    except:
        pass
    return 1

def es_fecha(valor):
    """Verificar si un valor es una fecha"""
    if valor is None:
        return False
    if isinstance(valor, datetime):
        return True
    if isinstance(valor, str):
        # Verificar si tiene formato de fecha/hora
        if ":" in valor or "-" in valor:
            try:
                # Intentar parsear como fecha
                datetime.strptime(valor.split()[0], "%Y-%m-%d")
                return True
            except:
                pass
    return False


def es_url(valor):
    """Determinar si una cadena apunta a una URL remota."""
    if not valor:
        return False
    valor_str = str(valor).strip().lower()
    return valor_str.startswith("http://") or valor_str.startswith("https://")

@st.cache_data(ttl=60)  # Cache por 60 segundos
def leer_letreros():
    """Leer todos los letreros del Excel con caché"""
    letreros = []
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "LETREROS" in wb.sheetnames:
            ws = wb["LETREROS"]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:  # Si hay ID
                    # Leer estado solo si no es una fecha
                    estado_val = ""
                    if len(row) > 8 and row[8]:
                        estado_str = str(row[8]).strip()
                        if estado_str and estado_str not in ["", "None"] and not es_fecha(row[8]):
                            estado_val = estado_str
                    
                    letrero = {
                        "ID": row[0],
                        "Cliente": row[1] if len(row) > 1 else "",
                        "Tipo Solicitud": row[2] if len(row) > 2 else "NUEVO",
                        "Tipo": row[3] if len(row) > 3 else "",
                        "Medidas": row[4] if len(row) > 4 else "",
                        "Costo": row[5] if len(row) > 5 else "",
                        "Comercial": row[6] if len(row) > 6 else "",
                        "Fecha": row[7] if len(row) > 7 else "",
                        "Estado": estado_val,
                        "Fecha_Envio_Fab": row[9] if len(row) > 9 else "",
                        "Fecha_Entrega_Cliente": row[10] if len(row) > 10 else "",
                        "Observaciones": row[11] if len(row) > 11 else "",
                        "Numero_Guia": row[12] if len(row) > 12 else ""
                    }
                    letreros.append(letrero)
        wb.close()
    except Exception as e:
        print(f"Error al leer letreros: {e}")
    return letreros

def guardar_fotos_percha(percha_id, fotos_uploaded):
    """Guardar fotos de una percha utilizando Google Drive o disco local como respaldo."""
    try:
        rutas_fotos = []
        try:
            from gdrive_helper import save_photo_percha_drive
            for idx, foto in enumerate(fotos_uploaded):
                result = save_photo_percha_drive(f"{percha_id}_{idx}", foto)
                if result and result.get('url'):
                    rutas_fotos.append(result['url'])
            if rutas_fotos:
                return rutas_fotos
        except Exception as e:
            print(f"Google Drive no disponible para perchas: {e}")
        
        fotos_dir = EXCEL_PATH.parent / "fotos_perchas"
        fotos_dir.mkdir(exist_ok=True)
        
        for idx, foto in enumerate(fotos_uploaded):
            extension = foto.name.split('.')[-1] if '.' in foto.name else 'jpg'
            nombre_foto = f"percha_{percha_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{idx}.{extension}"
            ruta_foto = fotos_dir / nombre_foto
            with open(ruta_foto, "wb") as f:
                f.write(foto.getbuffer())
            rutas_fotos.append(str(ruta_foto))
        return rutas_fotos
    except Exception as e:
        print(f"Error al guardar fotos de percha: {e}")
        return []

def guardar_fotos_comercial(entrega_id, fotos_uploaded):
    """Guardar fotos de una entrega comercial en Google Drive o local como respaldo."""
    try:
        rutas_fotos = []
        try:
            from gdrive_helper import save_photo_comercial_drive
            for idx, foto in enumerate(fotos_uploaded):
                result = save_photo_comercial_drive(f"{entrega_id}_{idx}", foto)
                if result and result.get('url'):
                    rutas_fotos.append(result['url'])
            if rutas_fotos:
                return rutas_fotos
        except Exception as e:
            print(f"Google Drive no disponible para comerciales: {e}")
        
        fotos_dir = EXCEL_PATH.parent / "fotos_comerciales"
        fotos_dir.mkdir(exist_ok=True)
        
        for idx, foto in enumerate(fotos_uploaded):
            extension = foto.name.split('.')[-1] if '.' in foto.name else 'jpg'
            nombre_foto = f"comercial_{entrega_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{idx}.{extension}"
            ruta_foto = fotos_dir / nombre_foto
            with open(ruta_foto, "wb") as f:
                f.write(foto.getbuffer())
            rutas_fotos.append(str(ruta_foto))
        return rutas_fotos
    except Exception as e:
        print(f"Error al guardar fotos comerciales: {e}")
        return []

def actualizar_fotos_comercial(entrega_id, nuevas_rutas_fotos):
    """Actualizar las rutas de fotos en el Excel"""
    wb = None
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "COMERCIALES" not in wb.sheetnames:
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        ws = wb["COMERCIALES"]
        entrega_encontrada = False
        
        # Buscar la entrega por ID
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            current_id = str(row[0].value).strip() if row[0].value is not None else ""
            buscado_id = str(entrega_id).strip()
            
            if current_id == buscado_id:
                entrega_encontrada = True
                
                # Obtener fotos existentes si hay
                fotos_existentes = []
                if len(row) > 9 and row[9].value:
                    fotos_existentes = str(row[9].value).strip().split('|') if '|' in str(row[9].value) else [str(row[9].value).strip()]
                    fotos_existentes = [f for f in fotos_existentes if f]
                
                # Combinar fotos existentes con nuevas
                todas_las_fotos = fotos_existentes + nuevas_rutas_fotos
                
                # Asegurarse de que la columna existe
                headers = [cell.value for cell in ws[1]]
                if "Fotos" not in headers:
                    col_fotos = ws.max_column + 1
                    ws.cell(row=1, column=col_fotos).value = "Fotos"
                else:
                    col_fotos = headers.index("Fotos") + 1
                
                # Guardar todas las rutas separadas por |
                ws.cell(row=idx, column=col_fotos).value = '|'.join(todas_las_fotos)
                
                # Guardar el archivo
                if safe_save_workbook(wb, EXCEL_PATH):
                    print(f"✅ Fotos actualizadas para entrega #{entrega_id}")
                    return True
                else:
                    print(f"❌ Error al guardar fotos")
                    if wb:
                        try:
                            wb.close()
                        except:
                            pass
                    return False
        
        if not entrega_encontrada:
            print(f"❌ No se encontró la entrega con ID: {entrega_id}")
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        if wb:
            try:
                wb.close()
            except:
                pass
        return True
        
    except Exception as e:
        print(f"❌ Error al actualizar fotos de entrega comercial: {e}")
        if wb:
            try:
                wb.close()
            except:
                pass
        return False

def actualizar_fotos_percha(percha_id, nuevas_rutas_fotos):
    """Actualizar las rutas de fotos en el Excel"""
    wb = None
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "PERCHAS" not in wb.sheetnames:
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        ws = wb["PERCHAS"]
        percha_encontrada = False
        
        # Buscar la percha por ID
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            current_id = str(row[0].value).strip() if row[0].value is not None else ""
            buscado_id = str(percha_id).strip()
            
            if current_id == buscado_id:
                percha_encontrada = True
                
                # Obtener fotos existentes si hay
                fotos_existentes = []
                if len(row) > 10 and row[10].value:
                    fotos_existentes = str(row[10].value).strip().split('|') if '|' in str(row[10].value) else [str(row[10].value).strip()]
                    fotos_existentes = [f for f in fotos_existentes if f]
                
                # Combinar fotos existentes con nuevas
                todas_las_fotos = fotos_existentes + nuevas_rutas_fotos
                
                # Asegurarse de que la columna existe
                headers = [cell.value for cell in ws[1]]
                if "Fotos" not in headers:
                    col_fotos = ws.max_column + 1
                    ws.cell(row=1, column=col_fotos).value = "Fotos"
                else:
                    col_fotos = headers.index("Fotos") + 1
                
                # Guardar todas las rutas separadas por |
                ws.cell(row=idx, column=col_fotos).value = '|'.join(todas_las_fotos)
                
                # Guardar el archivo
                if safe_save_workbook(wb, EXCEL_PATH):
                    print(f"✅ Fotos actualizadas para percha #{percha_id}")
                    leer_perchas.clear()
                    return True
                else:
                    print(f"❌ Error al guardar fotos")
                    if wb:
                        try:
                            wb.close()
                        except:
                            pass
                    return False
        
        if not percha_encontrada:
            print(f"❌ No se encontró la percha con ID: {percha_id}")
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        if wb:
            try:
                wb.close()
            except:
                pass
        return True
        
    except Exception as e:
        print(f"❌ Error al actualizar fotos de percha: {e}")
        if wb:
            try:
                wb.close()
            except:
                pass
        return False

@st.cache_data(ttl=30)
def leer_perchas():
    """Leer todas las perchas desde Excel"""
    perchas = []
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "PERCHAS" in wb.sheetnames:
            ws = wb["PERCHAS"]
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] and row[1]:  # ID y Cliente deben existir
                    percha = {
                        'ID': row[0],
                        'Cliente': str(row[1]).strip() if row[1] else '',
                        'Comercial/Vendedor': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                        'Fecha': str(row[3]).strip() if len(row) > 3 and row[3] else '',
                        'Estado': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                        'Cantidad de Compra': str(row[5]).strip() if len(row) > 5 and row[5] else '',
                        'Productos': str(row[6]).strip() if len(row) > 6 and row[6] else '',
                        'Observaciones': str(row[7]).strip() if len(row) > 7 and row[7] else '',
                        'Número de Guía': str(row[8]).strip() if len(row) > 8 and row[8] else '',
                        'Medio de Envío': str(row[9]).strip() if len(row) > 9 and row[9] else '',
                        'Fotos': str(row[10]).strip() if len(row) > 10 and row[10] else ''
                    }
                    perchas.append(percha)
        wb.close()
    except Exception as e:
        print(f"Error al cargar perchas: {e}")
    return perchas

@st.cache_data(ttl=60)  # Cache por 60 segundos
def leer_eventos():
    """Leer todos los eventos del Excel con caché"""
    eventos = []
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "EVENTOS" in wb.sheetnames:
            ws = wb["EVENTOS"]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:  # Si hay ID
                    estado_val = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                    
                    evento = {
                        "ID": row[0],
                        "Cliente": row[1] if len(row) > 1 else "",
                        "Tipo": row[2] if len(row) > 2 else "",
                        "Fecha": row[3] if len(row) > 3 else "",
                        "Estado": estado_val,
                        "Descripcion": row[5] if len(row) > 5 else "",
                        "Productos": row[6] if len(row) > 6 else "",
                        "Comercial/Agente": row[7] if len(row) > 7 else "",
                        "Número de Guía": row[8] if len(row) > 8 else "",
                        "Observaciones Envío": row[9] if len(row) > 9 else "",
                        "Productos Retornados": row[10] if len(row) > 10 else "",
                        "Fecha Retorno": row[11] if len(row) > 11 else "",
                        "Observaciones Retorno": row[12] if len(row) > 12 else ""
                    }
                    eventos.append(evento)
        wb.close()
    except Exception as e:
        print(f"Error al leer eventos: {e}")
    return eventos

@st.cache_data(ttl=30)  # Cache por 30 segundos
def leer_entregas_comerciales():
    """Leer todas las entregas a comerciales desde Excel"""
    entregas = []
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "COMERCIALES" in wb.sheetnames:
            ws = wb["COMERCIALES"]
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] and row[1]:  # ID y Comercial/Vendedor deben existir
                    entrega = {
                        'ID': row[0],
                        'Comercial/Vendedor': str(row[1]).strip() if row[1] else '',
                        'Tipo Entrega': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                        'Fecha': str(row[3]).strip() if len(row) > 3 and row[3] else '',
                        'Productos': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                        'Cliente Destino': str(row[5]).strip() if len(row) > 5 and row[5] else '',
                        'Estado': str(row[6]).strip() if len(row) > 6 and row[6] else '',
                        'Observaciones': str(row[7]).strip() if len(row) > 7 and row[7] else '',
                        'Productos Entregados': str(row[8]).strip() if len(row) > 8 and row[8] else '',
                        'Fotos': str(row[9]).strip() if len(row) > 9 and row[9] else '',
                        'ID_Entrega_Original': str(row[10]).strip() if len(row) > 10 and row[10] else ''  # Nueva columna para referenciar entrega original
                    }
                    entregas.append(entrega)
        wb.close()
    except Exception as e:
        print(f"Error al cargar entregas a comerciales: {e}")
    return entregas

def actualizar_percha_entregada(percha_id, numero_guia, medio_envio):
    """Actualizar una percha a estado 'ENTREGADO' con número de guía y medio de envío"""
    wb = None
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "PERCHAS" not in wb.sheetnames:
            print(f"❌ La hoja PERCHAS no existe")
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        ws = wb["PERCHAS"]
        percha_encontrada = False
        
        # Buscar la percha por ID
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            current_id = str(row[0].value).strip() if row[0].value is not None else ""
            buscado_id = str(percha_id).strip()
            
            if current_id == buscado_id:
                percha_encontrada = True
                # Actualizar estado (columna 5)
                ws.cell(row=idx, column=5).value = "ENTREGADO"
                
                # Asegurarse de que existen las columnas para número de guía y medio de envío
                headers = [cell.value for cell in ws[1]]
                if "Número de Guía" not in headers:
                    # Agregar columna "Número de Guía"
                    col_num_guia = ws.max_column + 1
                    ws.cell(row=1, column=col_num_guia).value = "Número de Guía"
                else:
                    col_num_guia = headers.index("Número de Guía") + 1
                
                if "Medio de Envío" not in headers:
                    # Agregar columna "Medio de Envío"
                    col_medio_envio = ws.max_column + 1
                    ws.cell(row=1, column=col_medio_envio).value = "Medio de Envío"
                else:
                    col_medio_envio = headers.index("Medio de Envío") + 1
                
                # Actualizar número de guía y medio de envío
                ws.cell(row=idx, column=col_num_guia).value = numero_guia
                ws.cell(row=idx, column=col_medio_envio).value = medio_envio
                
                print(f"✅ Actualizando percha #{percha_id} a 'ENTREGADO' con guía: {numero_guia}, medio: {medio_envio}")
                
                # Guardar el archivo
                try:
                    if safe_save_workbook(wb, EXCEL_PATH):
                        print(f"✅ Archivo guardado correctamente en {EXCEL_PATH}")
                    else:
                        print(f"❌ Error al guardar el archivo")
                        try:
                            wb.close()
                        except:
                            pass
                        return False
                except Exception as save_error:
                    print(f"❌ Error al guardar: {save_error}")
                    try:
                        wb.close()
                    except:
                        pass
                    return False
                break
        
        if not percha_encontrada:
            print(f"❌ No se encontró la percha con ID: {percha_id}")
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        if wb:
            try:
                wb.close()
            except:
                pass
        
        # Limpiar caché
        leer_perchas.clear()
        return True
        
    except Exception as e:
        print(f"❌ Error al actualizar percha entregada: {e}")
        if wb:
            try:
                wb.close()
            except:
                pass
        return False

def actualizar_estado_percha(percha_id, nuevo_estado):
    """Actualizar el estado de una percha en el Excel"""
    wb = None
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "PERCHAS" not in wb.sheetnames:
            print(f"❌ La hoja PERCHAS no existe")
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        ws = wb["PERCHAS"]
        percha_encontrada = False
        
        # Buscar la percha por ID
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # Convertir ambos a string para comparar correctamente, eliminando espacios
            current_id = str(row[0].value).strip() if row[0].value is not None else ""
            buscado_id = str(percha_id).strip()
            
            if current_id == buscado_id:
                percha_encontrada = True
                # Estado está en la columna 4 (índice 3 en el array, columna D en Excel)
                # Según HEADERS: ["ID", "Cliente", "Comercial/Vendedor", "Fecha", "Estado", ...]
                estado_actual = row[4].value if len(row) > 4 else None
                
                # Asegurarse de que la celda existe
                while ws.max_column < 5:
                    ws.cell(row=1, column=ws.max_column + 1, value="")
                
                ws.cell(row=idx, column=5).value = nuevo_estado
                print(f"✅ Actualizando percha #{percha_id} de '{estado_actual}' a '{nuevo_estado}' en fila {idx}, columna 5")
                
                # Guardar el archivo
                try:
                    if safe_save_workbook(wb, EXCEL_PATH):
                        print(f"✅ Archivo guardado correctamente en {EXCEL_PATH}")
                    else:
                        print(f"❌ Error al guardar el archivo")
                        try:
                            wb.close()
                        except:
                            pass
                        return False
                except Exception as save_error:
                    print(f"❌ Error al guardar: {save_error}")
                    try:
                        wb.close()
                    except:
                        pass
                    return False
                break
        
        if not percha_encontrada:
            print(f"❌ No se encontró la percha con ID: {percha_id}")
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        if wb:
            try:
                wb.close()
            except:
                pass
        
        # Limpiar caché
        leer_perchas.clear()
        return True
        
    except Exception as e:
        print(f"❌ Error al actualizar estado de percha: {e}")
        if wb:
            try:
                wb.close()
            except:
                pass
        return False

def actualizar_estado_evento(evento_id, nuevo_estado):
    """Actualizar el estado de un evento en el Excel"""
    wb = None
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "EVENTOS" not in wb.sheetnames:
            print(f"❌ La hoja EVENTOS no existe")
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        ws = wb["EVENTOS"]
        evento_encontrado = False
        
        # Buscar el evento por ID
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # Convertir ambos a string para comparar correctamente, eliminando espacios
            current_id = str(row[0].value).strip() if row[0].value is not None else ""
            buscado_id = str(evento_id).strip()
            
            if current_id == buscado_id:
                evento_encontrado = True
                # Estado está en la columna 5 (índice 4 en el array, columna E en Excel)
                estado_actual = row[4].value if len(row) > 4 else None
                
                # Asegurarse de que la celda existe
                while ws.max_column < 5:
                    ws.cell(row=1, column=ws.max_column + 1, value="")
                
                ws.cell(row=idx, column=5).value = nuevo_estado
                print(f"✅ Actualizando evento #{evento_id} de '{estado_actual}' a '{nuevo_estado}' en fila {idx}, columna 5")
                
                # Guardar el archivo
                try:
                    if safe_save_workbook(wb, EXCEL_PATH):
                        print(f"✅ Archivo guardado correctamente en {EXCEL_PATH}")
                    else:
                        print(f"❌ Error al guardar el archivo")
                        try:
                            wb.close()
                        except:
                            pass
                        return False
                except Exception as save_error:
                    print(f"❌ Error al guardar: {save_error}")
                    try:
                        wb.close()
                    except:
                        pass
                    return False
                
                # Cerrar el workbook después de guardar
                try:
                    wb.close()
                except:
                    pass
                
                # Limpiar caché ANTES de retornar
                leer_eventos.clear()
                time.sleep(0.3)  # Pequeño delay para asegurar que el sistema de archivos actualice
                return True
        
        if not evento_encontrado:
            print(f"⚠️ Evento #{evento_id} no encontrado en el Excel")
            if wb:
                try:
                    wb.close()
                except:
                    pass
        return False
    except Exception as e:
        print(f"❌ Error al actualizar estado del evento: {e}")
        import traceback
        traceback.print_exc()
        if wb:
            try:
                wb.close()
            except:
                pass
        return False

def actualizar_envio_evento(evento_id, numero_guia, observaciones):
    """Actualizar número de guía y observaciones de envío de un evento en el Excel"""
    wb = None
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "EVENTOS" not in wb.sheetnames:
            print(f"❌ La hoja EVENTOS no existe")
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        ws = wb["EVENTOS"]
        evento_encontrado = False
        
        # Buscar el evento por ID
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            current_id = str(row[0].value).strip() if row[0].value is not None else ""
            buscado_id = str(evento_id).strip()
            
            if current_id == buscado_id:
                evento_encontrado = True
                
                # Asegurarse de que las columnas existen
                while ws.max_column < 10:
                    ws.cell(row=1, column=ws.max_column + 1, value="")
                
                # Columna 9 es "Número de Guía" (índice 8)
                ws.cell(row=idx, column=9).value = numero_guia if numero_guia else ""
                # Columna 10 es "Observaciones Envío" (índice 9)
                ws.cell(row=idx, column=10).value = observaciones if observaciones else ""
                
                print(f"✅ Actualizando envío para evento #{evento_id}: Guía={numero_guia}, Obs={observaciones}")
                
                # Guardar el archivo
                try:
                    if safe_save_workbook(wb, EXCEL_PATH):
                        print(f"✅ Archivo guardado correctamente")
                    else:
                        print(f"❌ Error al guardar el archivo")
                        try:
                            wb.close()
                        except:
                            pass
                        return False
                except Exception as save_error:
                    print(f"❌ Error al guardar: {save_error}")
                    try:
                        wb.close()
                    except:
                        pass
                    return False
                
                # Cerrar el workbook después de guardar
                try:
                    wb.close()
                except:
                    pass
                
                # Limpiar caché
                leer_eventos.clear()
                time.sleep(0.3)
                return True
        
        if not evento_encontrado:
            print(f"⚠️ Evento #{evento_id} no encontrado en el Excel")
            if wb:
                try:
                    wb.close()
                except:
                    pass
        return False
    except Exception as e:
        print(f"❌ Error al actualizar envío del evento: {e}")
        import traceback
        traceback.print_exc()
        if wb:
            try:
                wb.close()
            except:
                pass
        return False

def actualizar_retorno_evento(evento_id, productos_retornados, fecha_retorno, observaciones_retorno):
    """Actualizar información de retorno de productos de un evento en el Excel"""
    wb = None
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "EVENTOS" not in wb.sheetnames:
            print(f"❌ La hoja EVENTOS no existe")
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        ws = wb["EVENTOS"]
        evento_encontrado = False
        
        # Buscar el evento por ID
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            current_id = str(row[0].value).strip() if row[0].value is not None else ""
            buscado_id = str(evento_id).strip()
            
            if current_id == buscado_id:
                evento_encontrado = True
                
                # Asegurarse de que las columnas existen
                while ws.max_column < 13:
                    ws.cell(row=1, column=ws.max_column + 1, value="")
                
                # Columna 11 es "Productos Retornados" (índice 10)
                ws.cell(row=idx, column=11).value = productos_retornados if productos_retornados else ""
                # Columna 12 es "Fecha Retorno" (índice 11)
                ws.cell(row=idx, column=12).value = fecha_retorno if fecha_retorno else ""
                # Columna 13 es "Observaciones Retorno" (índice 12)
                ws.cell(row=idx, column=13).value = observaciones_retorno if observaciones_retorno else ""
                
                print(f"✅ Actualizando retorno para evento #{evento_id}")
                
                # Guardar el archivo
                try:
                    if safe_save_workbook(wb, EXCEL_PATH):
                        print(f"✅ Archivo guardado correctamente")
                    else:
                        print(f"❌ Error al guardar el archivo")
                        try:
                            wb.close()
                        except:
                            pass
                        return False
                except Exception as save_error:
                    print(f"❌ Error al guardar: {save_error}")
                    try:
                        wb.close()
                    except:
                        pass
                    return False
                
                # Cerrar el workbook después de guardar
                try:
                    wb.close()
                except:
                    pass
                
                # Reintegrar stock de productos retornados
                if productos_retornados:
                    reintegrar_stock_retorno(evento_id, productos_retornados)
                
                # Limpiar caché
                leer_eventos.clear()
                time.sleep(0.3)
                return True
        
        if not evento_encontrado:
            print(f"⚠️ Evento #{evento_id} no encontrado en el Excel")
            if wb:
                try:
                    wb.close()
                except:
                    pass
        return False
    except Exception as e:
        print(f"❌ Error al actualizar retorno del evento: {e}")
        import traceback
        traceback.print_exc()
        if wb:
            try:
                wb.close()
            except:
                pass
        return False

def append_movement_global(mov):
    """Función global para registrar movimientos de inventario"""
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        INV_SHEET_MOVS = "INVENTARIO_MOVIMIENTOS"
        if INV_SHEET_MOVS not in wb.sheetnames:
            ws = wb.create_sheet(INV_SHEET_MOVS)
            ws.append(["Fecha", "Tipo", "Codigo", "Cantidad", "CostoUnit", "Proveedor", "Proceso", "Nota"])
        ws = wb[INV_SHEET_MOVS]
        ws.append([mov["Fecha"], mov["Tipo"], mov["Codigo"], mov["Cantidad"], mov["CostoUnit"], mov["Proveedor"], mov["Proceso"], mov["Nota"]])
        if safe_save_workbook(wb, EXCEL_PATH):
            wb.close()
            return True
        wb.close()
    except Exception as e:
        print(f"Error guardando movimiento: {e}")
    return False

def validar_stock_suficiente(productos_detalles):
    """Validar que hay stock suficiente para todos los productos"""
    wb = None
    try:
        # Obtener stock disponible
        stock_by_code = {}
        wb = safe_load_workbook(EXCEL_PATH)
        # base: stock inicial de productos
        if "INVENTARIO_PRODUCTOS" in wb.sheetnames:
            for row in wb["INVENTARIO_PRODUCTOS"].iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    code = str(row[0]).strip()
                    stock_by_code[code] = float(row[5]) if len(row) > 5 and row[5] else 0.0
        # movimientos
        if "INVENTARIO_MOVIMIENTOS" in wb.sheetnames:
            for row in wb["INVENTARIO_MOVIMIENTOS"].iter_rows(min_row=2, values_only=True):
                if row and len(row) > 3 and row[1] and row[2] and row[3]:
                    t = str(row[1]).strip().upper()
                    code = str(row[2]).strip()
                    qty = float(row[3]) if row[3] else 0.0
                    if code not in stock_by_code:
                        stock_by_code[code] = 0.0
                    if t == "ENTRADA" or t == "AJUSTE+":
                        stock_by_code[code] += qty
                    elif t == "SALIDA" or t == "AJUSTE-":
                        stock_by_code[code] -= qty
        if wb:
            wb.close()
        
        # Validar cada producto
        productos_insuficientes = []
        for prod in productos_detalles:
            codigo = prod.get('codigo', '')
            cantidad_solicitada = float(prod.get('cantidad', 0))
            stock_disponible = stock_by_code.get(codigo, 0.0)
            
            if cantidad_solicitada > 0 and stock_disponible < cantidad_solicitada:
                productos_insuficientes.append({
                    'codigo': codigo,
                    'nombre': prod.get('nombre', ''),
                    'solicitado': cantidad_solicitada,
                    'disponible': stock_disponible
                })
        
        return len(productos_insuficientes) == 0, productos_insuficientes
    except Exception as e:
        print(f"Error validando stock: {e}")
        if wb:
            try:
                wb.close()
            except:
                pass
        return False, []

def descontar_stock_evento(evento):
    """Descontar stock de productos del evento"""
    try:
        productos_str = evento.get('Productos', '')
        if not productos_str:
            return True  # No hay productos, no hay nada que descontar
        
        # Parsear productos desde el string
        productos_lines = productos_str.split('\n')
        movimientos_registrados = 0
        
        for line in productos_lines:
            line = line.strip()
            if not line:
                continue
            
            # Formato: "2 ES10001 - EXHIBIDOR DE BATERIAS"
            import re
            m = re.match(r"^(\d+)\s+(.+)$", line)
            if m:
                try:
                    cantidad = int(m.group(1))
                    # Extraer código del producto (primera parte antes del guión)
                    codigo_nombre = m.group(2).strip()
                    codigo_match = re.match(r"^([A-Z0-9_]+)\s*-\s*", codigo_nombre)
                    if codigo_match:
                        codigo = codigo_match.group(1)
                    else:
                        codigo = codigo_nombre.split()[0] if codigo_nombre.split() else ""
                    
                    if codigo and cantidad > 0:
                        # Registrar movimiento de SALIDA
                        mov = {
                            "Fecha": datetime.now().strftime("%Y-%m-%d"),
                            "Tipo": "SALIDA",
                            "Codigo": codigo,
                            "Cantidad": float(cantidad),
                            "CostoUnit": 0.0,
                            "Proveedor": "",
                            "Proceso": "Evento",
                            "Nota": f"Evento #{evento.get('ID', 'N/A')} - {evento.get('Cliente', 'N/A')}"
                        }
                        if append_movement_global(mov):
                            movimientos_registrados += 1
                            print(f"✅ Stock descontado: {cantidad} unidades de {codigo}")
                        else:
                            print(f"❌ Error al registrar movimiento para {codigo}")
                except Exception as e:
                    print(f"Error parseando línea de producto: {line} - {e}")
        
        return movimientos_registrados > 0 or not productos_lines
    except Exception as e:
        print(f"Error descontando stock del evento: {e}")
        return False

def reintegrar_stock_retorno(evento_id, productos_retornados_str):
    """Reintegrar stock de productos retornados al inventario"""
    try:
        if not productos_retornados_str:
            return True  # No hay productos retornados
        
        # Parsear productos desde el string
        productos_lines = productos_retornados_str.split('\n')
        movimientos_registrados = 0
        
        for line in productos_lines:
            line = line.strip()
            if not line:
                continue
            
            # Formato: "2 ES10001 - EXHIBIDOR DE BATERIAS"
            import re
            m = re.match(r"^(\d+)\s+(.+)$", line)
            if m:
                try:
                    cantidad = int(m.group(1))
                    # Extraer código del producto (primera parte antes del guión)
                    codigo_nombre = m.group(2).strip()
                    codigo_match = re.match(r"^([A-Z0-9_]+)\s*-\s*", codigo_nombre)
                    if codigo_match:
                        codigo = codigo_match.group(1)
                    else:
                        codigo = codigo_nombre.split()[0] if codigo_nombre.split() else ""
                    
                    if codigo and cantidad > 0:
                        # Registrar movimiento de ENTRADA
                        mov = {
                            "Fecha": datetime.now().strftime("%Y-%m-%d"),
                            "Tipo": "ENTRADA",
                            "Codigo": codigo,
                            "Cantidad": float(cantidad),
                            "CostoUnit": 0.0,
                            "Proveedor": "",
                            "Proceso": "Retorno Evento",
                            "Nota": f"Retorno Evento #{evento_id}"
                        }
                        if append_movement_global(mov):
                            movimientos_registrados += 1
                            print(f"✅ Stock reintegrado: {cantidad} unidades de {codigo}")
                        else:
                            print(f"❌ Error al registrar movimiento para {codigo}")
                except Exception as e:
                    print(f"Error parseando línea de producto: {line} - {e}")
        
        return movimientos_registrados > 0 or not productos_lines
    except Exception as e:
        print(f"Error reintegrando stock del retorno: {e}")
        return False

def actualizar_estado_letrero(letrero_id, nuevo_estado):
    """Actualizar el estado de un letrero en el Excel"""
    wb = None
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "LETREROS" not in wb.sheetnames:
            print(f"❌ La hoja LETREROS no existe")
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        ws = wb["LETREROS"]
        letrero_encontrado = False
        
        # Buscar el letrero por ID
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # Convertir ambos a string para comparar correctamente, eliminando espacios
            current_id = str(row[0].value).strip() if row[0].value is not None else ""
            buscado_id = str(letrero_id).strip()
            
            if current_id == buscado_id:
                letrero_encontrado = True
                # Estado está en la columna 9 (índice 8 en el array, columna I en Excel)
                estado_actual = row[8].value if len(row) > 8 else None
                
                # Asegurarse de que la celda existe
                while ws.max_column < 9:
                    ws.cell(row=1, column=ws.max_column + 1, value="")
                
                ws.cell(row=idx, column=9).value = nuevo_estado
                print(f"✅ Actualizando letrero #{letrero_id} de '{estado_actual}' a '{nuevo_estado}' en fila {idx}, columna 9")
                
                # Guardar el archivo
                try:
                    wb.save(EXCEL_PATH)
                    print(f"✅ Archivo guardado correctamente en {EXCEL_PATH}")
                except Exception as save_error:
                    print(f"❌ Error al guardar: {save_error}")
                    try:
                        wb.close()
                    except:
                        pass
                    return False
                
                # Cerrar el workbook después de guardar
                try:
                    wb.close()
                except:
                    pass
                
                # Limpiar caché ANTES de retornar
                leer_letreros.clear()
                time.sleep(0.3)  # Pequeño delay para asegurar que el sistema de archivos actualice
                return True
        
        if not letrero_encontrado:
            print(f"⚠️ Letrero #{letrero_id} no encontrado en el Excel")
            if wb:
                try:
                    wb.close()
                except:
                    pass
        return False
    except Exception as e:
        print(f"❌ Error al actualizar estado del letrero: {e}")
        import traceback
        traceback.print_exc()
        if wb:
            try:
                wb.close()
            except:
                pass
        return False

def eliminar_letrero(letrero_id):
    """Eliminar un letrero del Excel"""
    wb = None
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "LETREROS" not in wb.sheetnames:
            print(f"❌ La hoja LETREROS no existe")
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return False
        
        ws = wb["LETREROS"]
        letrero_encontrado = False
        
        # Buscar el letrero por ID
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # Convertir ambos a string para comparar correctamente, eliminando espacios
            current_id = str(row[0].value).strip() if row[0].value is not None else ""
            buscado_id = str(letrero_id).strip()
            
            if current_id == buscado_id:
                letrero_encontrado = True
                # Eliminar la fila
                ws.delete_rows(idx)
                print(f"✅ Eliminando letrero #{letrero_id} en fila {idx}")
                
                # Guardar el archivo
                try:
                    wb.save(EXCEL_PATH)
                    print(f"✅ Archivo guardado correctamente en {EXCEL_PATH}")
                except Exception as save_error:
                    print(f"❌ Error al guardar: {save_error}")
                    try:
                        wb.close()
                    except:
                        pass
                    return False
                
                # Cerrar el workbook después de guardar
                try:
                    wb.close()
                except:
                    pass
                
                # Limpiar caché ANTES de retornar
                leer_letreros.clear()
                time.sleep(0.3)  # Pequeño delay para asegurar que el sistema de archivos actualice
                return True
        
        if not letrero_encontrado:
            print(f"⚠️ Letrero #{letrero_id} no encontrado en el Excel")
            if wb:
                try:
                    wb.close()
                except:
                    pass
        return False
    except Exception as e:
        print(f"❌ Error al eliminar letrero: {e}")
        import traceback
        traceback.print_exc()
        if wb:
            try:
                wb.close()
            except:
                pass
        return False

def actualizar_fechas_fabricacion(letrero_id, fecha_envio=None, fecha_entrega=None, numero_guia=None, limpiar=False):
    """Actualizar las fechas de fabricación y número de guía de un letrero en el Excel"""
    try:
        wb = safe_load_workbook(EXCEL_PATH)
        if "LETREROS" in wb.sheetnames:
            ws = wb["LETREROS"]
            
            # Buscar el letrero por ID
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                # Convertir ambos a string para comparar correctamente
                current_id = str(row[0].value) if row[0].value is not None else ""
                buscado_id = str(letrero_id)
                
                if current_id == buscado_id:
                    # Si limpiar está activado, establecer ambos como None
                    if limpiar:
                        ws.cell(row=idx, column=10).value = None  # Fecha Envío Fabricación (columna 10)
                        ws.cell(row=idx, column=11).value = None  # Fecha Entrega Cliente (columna 11)
                        ws.cell(row=idx, column=13).value = None  # Número de Guía (columna 13)
                        print(f"Eliminando todas las fechas para letrero #{letrero_id}")
                    else:
                        # Fecha envío está en la columna 10 (Fecha Envío Fabricación)
                        if fecha_envio is not None:
                            ws.cell(row=idx, column=10).value = fecha_envio
                            print(f"Actualizando fecha de envío para letrero #{letrero_id} en columna 10")
                        
                        # Fecha entrega está en la columna 11 (Fecha Entrega Cliente)
                        if fecha_entrega is not None:
                            ws.cell(row=idx, column=11).value = fecha_entrega
                            print(f"Actualizando fecha de entrega para letrero #{letrero_id} en columna 11")
                        
                        # Número de guía está en la columna 13 (después de Observaciones que está en columna 12)
                        if numero_guia is not None:
                            ws.cell(row=idx, column=13).value = numero_guia
                            print(f"Actualizando número de guía para letrero #{letrero_id} en columna 13")
                    
                    if safe_save_workbook(wb, EXCEL_PATH):
                        print(f"✅ Fechas y guía actualizadas para letrero #{letrero_id}")
                        # Limpiar caché para refrescar datos
                        leer_letreros.clear()
                        return True
                    else:
                        print(f"❌ Error al guardar el workbook para letrero #{letrero_id}")
                        return False
        print(f"⚠️ Letrero #{letrero_id} no encontrado para actualizar fechas")
        return False
    except Exception as e:
        print(f"❌ Error al actualizar fechas del letrero: {e}")
        import traceback
        traceback.print_exc()
        return False

def generar_pdf_autorizacion_evento(evento):
    """Generar PDF de autorización para un evento"""
    try:
        # Crear carpeta para el cliente si no existe
        carpeta_cliente = Path(EXCEL_DIR) / "EVENTOS_AUTORIZACIONES" / evento['Cliente'].replace("/", "_").replace("\\", "_")
        carpeta_cliente.mkdir(parents=True, exist_ok=True)
        
        # Nombre del archivo
        nombre_archivo = carpeta_cliente / f"Autorizacion_Evento_{evento['ID']}.pdf"
        
        # Crear el PDF
        doc = SimpleDocTemplate(str(nombre_archivo), pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        
        # Logo EXTREMEMAX
        ruta_logo = Path(EXCEL_DIR) / "logo_extrememax.png"
        if ruta_logo.exists():
            logo = Image(str(ruta_logo), width=2.5*inch, height=0.8*inch)
            elements.append(logo)
        else:
            # Fallback si no existe el logo
            header_text = '<font size="24" color="#FF0000"><b>EXTREMEMAX</b></font><br/><font size="12">OEM PARTS</font>'
            elements.append(Paragraph(header_text, styles['Heading1']))
        
        elements.append(Spacer(1, 0.3*inch))
        
        # Información del cliente
        cliente_val = evento.get('Cliente') or 'N/A'
        agente_val = evento.get('Comercial/Agente') or 'N/A'
        fecha_val = str(evento.get('Fecha')) if evento.get('Fecha') else 'N/A'
        tipo_val = evento.get('Tipo') or 'N/A'
        
        cliente_info = [
            [Paragraph('<b>Cliente:</b>', styles['Normal']), Paragraph(str(cliente_val), styles['Normal'])],
            [Paragraph('<b>Comercial/Agente:</b>', styles['Normal']), Paragraph(str(agente_val), styles['Normal'])],
            [Paragraph('<b>Tipo de Evento:</b>', styles['Normal']), Paragraph(str(tipo_val), styles['Normal'])],
            [Paragraph('<b>Fecha:</b>', styles['Normal']), Paragraph(str(fecha_val), styles['Normal'])]
        ]
        
        tabla_cliente = Table(cliente_info, colWidths=[1.5*inch, 4.5*inch])
        tabla_cliente.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
        ]))
        elements.append(tabla_cliente)
        elements.append(Spacer(1, 0.3*inch))
        
        # Productos solicitados
        elements.append(Paragraph("<b>Productos Solicitados:</b>", styles['Heading3']))
        elements.append(Spacer(1, 0.1*inch))
        
        productos_val = evento.get('Productos') or 'N/A'
        # Dividir productos en líneas y crear tabla
        productos_lines = str(productos_val).split('\n')
        productos_table_data = [[Paragraph('<b>Producto</b>', styles['Normal'])]]
        for prod in productos_lines:
            if prod.strip():
                productos_table_data.append([Paragraph(prod.strip(), styles['Normal'])])
        
        if len(productos_table_data) > 1:
            tabla_productos = Table(productos_table_data, colWidths=[6*inch])
            tabla_productos.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('BACKGROUND', (0, 0), (0, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
            ]))
            elements.append(tabla_productos)
        else:
            elements.append(Paragraph(productos_val, styles['Normal']))
        
        elements.append(Spacer(1, 0.4*inch))
        
        # Descripción si existe
        if evento.get('Descripcion'):
            elements.append(Paragraph("<b>Descripción:</b>", styles['Heading3']))
            elements.append(Spacer(1, 0.1*inch))
            elements.append(Paragraph(str(evento.get('Descripcion')), styles['Normal']))
            elements.append(Spacer(1, 0.4*inch))
        
        # Sección de autorización y firma
        elements.append(Paragraph("<b>Autorización:</b>", styles['Heading3']))
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph(f"Por medio de la presente se autoriza el evento descrito arriba para el cliente {cliente_val}.", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Línea para firma
        elements.append(Paragraph("__________________________", styles['Normal']))
        elements.append(Spacer(1, 0.05*inch))
        elements.append(Paragraph("<b>GERENTE GENERAL DE EXTREMEMAX IMPORTACIONES S.A.S</b>", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        elements.append(Paragraph("Atentamente:", styles['Normal']))
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
        
        # Construir el PDF
        doc.build(elements)
        
        # Intentar subir a Google Drive
        try:
            from gdrive_helper import save_pdf_evento_drive
            drive_result = save_pdf_evento_drive(nombre_archivo, evento['ID'], evento['Cliente'])
            if drive_result and drive_result.get('url'):
                return drive_result['url']
        except Exception as e:
            print(f"Google Drive no disponible para PDF de evento: {e}")
        
        return nombre_archivo
    except Exception as e:
        print(f"Error al generar PDF: {e}")
        return None

def generar_pdf_autorizacion(letrero):
    """Generar PDF de autorización para un letrero"""
    try:
        # Crear carpeta para el cliente si no existe
        carpeta_cliente = Path(EXCEL_DIR) / "LETREROS_AUTORIZACIONES" / letrero['Cliente'].replace("/", "_").replace("\\", "_")
        carpeta_cliente.mkdir(parents=True, exist_ok=True)
        
        # Nombre del archivo
        nombre_archivo = carpeta_cliente / f"Autorizacion_Letrero_{letrero['ID']}.pdf"
        
        # Crear el PDF
        doc = SimpleDocTemplate(str(nombre_archivo), pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        
        # Logo EXTREMEMAX
        ruta_logo = Path(EXCEL_DIR) / "logo_extrememax.png"
        if ruta_logo.exists():
            logo = Image(str(ruta_logo), width=2.5*inch, height=0.8*inch)
            elements.append(logo)
        else:
            # Fallback si no existe el logo
            header_text = '<font size="24" color="#FF0000"><b>EXTREMEMAX</b></font><br/><font size="12">OEM PARTS</font>'
            elements.append(Paragraph(header_text, styles['Heading1']))
        
        elements.append(Spacer(1, 0.3*inch))
        
        # Información del cliente - Usar Paragraph para que <b> funcione
        cliente_val = letrero.get('Cliente') or 'N/A'
        agente_val = letrero.get('Comercial') or 'N/A'
        fecha_val = str(letrero.get('Fecha')) if letrero.get('Fecha') else 'N/A'
        
        # Buscar monto de ventas del cliente
        monto_ventas = 0
        try:
            if EXCEL_PATH.exists():
                wb_ventas = safe_load_workbook(EXCEL_PATH, tries=3, wait=0.1)
                if "VENTAS_VINCULADAS" in wb_ventas.sheetnames:
                    ws_ventas = wb_ventas["VENTAS_VINCULADAS"]
                    cliente_norm = normalizar_nombre(cliente_val)
                    
                    for row in ws_ventas.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:  # Si hay cliente
                            cliente_venta_norm = normalizar_nombre(str(row[0]).strip())
                            if cliente_norm == cliente_venta_norm:
                                venta_valor = row[1] if len(row) > 1 else 0
                                try:
                                    monto_ventas += float(venta_valor) if venta_valor else 0
                                except:
                                    pass
                wb_ventas.close()
        except:
            pass
        
        # Formatear monto de ventas
        monto_ventas_str = f"${monto_ventas:,.2f}" if monto_ventas > 0 else "No disponible"
        
        cliente_info = [
            [Paragraph('<b>Cliente:</b>', styles['Normal']), Paragraph(str(cliente_val), styles['Normal'])],
            [Paragraph('<b>Agente:</b>', styles['Normal']), Paragraph(str(agente_val), styles['Normal'])],
            [Paragraph('<b>Fecha:</b>', styles['Normal']), Paragraph(str(fecha_val), styles['Normal'])]
        ]
        
        # Agregar monto de ventas si está disponible
        if monto_ventas > 0:
            cliente_info.append([
                Paragraph('<b>Venta Total Cliente:</b>', styles['Normal']), 
                Paragraph(monto_ventas_str, styles['Normal'])
            ])
        
        tabla_cliente = Table(cliente_info, colWidths=[1.5*inch, 4.5*inch])
        tabla_cliente.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
        ]))
        elements.append(tabla_cliente)
        elements.append(Spacer(1, 0.3*inch))
        
        # Solicitud del letrero
        elements.append(Paragraph("<b>Solicitud:</b>", styles['Heading3']))
        elements.append(Spacer(1, 0.1*inch))
        
        tipo_solicitud_val = letrero.get('Tipo Solicitud', 'NUEVO')
        tipo_val = letrero.get('Tipo') or 'N/A'
        medidas_val = letrero.get('Medidas') or 'N/A'
        texto_solicitud = f"Entregar letrero tipo {tipo_val} ({tipo_solicitud_val}) con medidas de {medidas_val} al cliente mencionado."
        elements.append(Paragraph(texto_solicitud, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Detalles del letrero
        tipo_val = letrero.get('Tipo') or 'N/A'
        medidas_val = letrero.get('Medidas') or 'N/A'
        costo_val = letrero.get('Costo') or 0
        obs_val = letrero.get('Observaciones') or 'Ninguna'
        
        detalles = [
            [Paragraph('<b>Tipo de Letrero:</b>', styles['Normal']), Paragraph(str(tipo_val), styles['Normal'])],
            [Paragraph('<b>Medidas:</b>', styles['Normal']), Paragraph(str(medidas_val), styles['Normal'])],
            [Paragraph('<b>Costo:</b>', styles['Normal']), Paragraph(f"${costo_val}", styles['Normal'])],
            [Paragraph('<b>Observaciones:</b>', styles['Normal']), Paragraph(str(obs_val), styles['Normal'])]
        ]
        
        tabla_detalles = Table(detalles, colWidths=[2*inch, 4*inch])
        tabla_detalles.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6)
        ]))
        elements.append(tabla_detalles)
        elements.append(Spacer(1, 0.4*inch))
        
        # Sección de autorización y firma
        elements.append(Paragraph("<b>ARTE FINAL:</b>", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        elements.append(Paragraph("<b>Autorización:</b>", styles['Heading3']))
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph(f"Por medio de la presente se autoriza la entrega del letrero descrito arriba al cliente {cliente_val}.", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Línea para firma
        elements.append(Paragraph("__________________________", styles['Normal']))
        elements.append(Spacer(1, 0.05*inch))
        elements.append(Paragraph("<b>GERENTE GENERAL DE EXTREMEMAX IMPORTACIONES S.A.S</b>", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        elements.append(Paragraph("Atentamente:", styles['Normal']))
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
        
        # Construir el PDF
        doc.build(elements)
        
        # Intentar subir a Google Drive
        try:
            from gdrive_helper import save_pdf_letrero_drive
            drive_result = save_pdf_letrero_drive(nombre_archivo, letrero['ID'], letrero['Cliente'])
            if drive_result and drive_result.get('url'):
                return drive_result['url']
        except Exception as e:
            print(f"Google Drive no disponible para PDF de letrero: {e}")
        
        return nombre_archivo
    except Exception as e:
        print(f"Error al generar PDF: {e}")
        return None

def append_row(sheet_name: str, values: list) -> bool:
    """Agregar fila al Excel"""
    ensure_workbook(EXCEL_PATH)
    wb = None
    try:
        wb = safe_load_workbook(EXCEL_PATH)
    except PermissionError as e:
        st.error(f"🔒 **Error de permisos**: El archivo Excel está siendo usado por otro programa (OneDrive o Excel abierto). Por favor:\n1. Cierra Excel si está abierto\n2. Pausa sincronización de OneDrive temporalmente\n3. Espera unos segundos y vuelve a intentar")
        print(f"❌ Error de permisos: {e}")
        return False
    except Exception as e:
        st.error(f"🔒 **Error al abrir Excel**: {e}\n\n**Soluciones:**\n1. Cierra el archivo si está abierto\n2. Verifica permisos de escritura en la carpeta\n3. Pausa OneDrive temporalmente")
        print(f"❌ Error inesperado al abrir: {e}")
        return False
    
    if wb is None:
        st.error("❌ No se pudo cargar el archivo Excel")
        return False
    
    try:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(HEADERS[sheet_name])
        else:
            ws = wb[sheet_name]
        
        # Asegurar que todas las columnas del HEADER existan
        headers_existentes = [cell.value for cell in ws[1]]
        headers_requeridos = HEADERS.get(sheet_name, [])
        
        # Si falta la columna ID_Entrega_Original, agregarla
        if sheet_name == "COMERCIALES" and "ID_Entrega_Original" not in headers_existentes:
            col_idx = len(headers_existentes) + 1
            ws.cell(row=1, column=col_idx).value = "ID_Entrega_Original"
            print(f"✅ Columna ID_Entrega_Original agregada a la hoja {sheet_name} en columna {col_idx}")
            # Guardar antes de continuar para que la columna exista
            safe_save_workbook(wb, EXCEL_PATH)
        
        # Asegurar que values tenga el mismo número de elementos que headers
        headers_actuales = [cell.value for cell in ws[1]]
        
        # Si es COMERCIALES, asegurar que el ID_Entrega_Original esté en la posición correcta
        if sheet_name == "COMERCIALES":
            # Encontrar el índice de ID_Entrega_Original en los headers
            idx_id_original = None
            for idx, header in enumerate(headers_actuales):
                if header == "ID_Entrega_Original":
                    idx_id_original = idx
                    break
            
            # Si la columna existe, asegurar que el valor esté en la posición correcta
            if idx_id_original is not None:
                # Asegurar que values tenga el tamaño correcto
                while len(values) < len(headers_actuales):
                    values.append("")
                
                # Buscar el ID_Entrega_Original en values (puede estar al final o en otra posición)
                id_original_val = None
                # Primero verificar si ya está en la posición correcta
                if len(values) > idx_id_original:
                    id_original_val = values[idx_id_original]
                    # Si está vacío o es None, buscar en otras posiciones
                    if not id_original_val or str(id_original_val).strip() == "":
                        id_original_val = None
                
                # Si no está en la posición correcta, buscarlo en el último elemento (donde se agregó)
                if id_original_val is None and len(values) > 0:
                    # Verificar el último elemento
                    ultimo_val = values[-1] if values else None
                    if ultimo_val and str(ultimo_val).strip() != "":
                        # Verificar si es un número o string que podría ser el ID
                        try:
                            # Intentar convertir a string para comparar
                            id_original_val = str(ultimo_val).strip()
                        except:
                            pass
                
                # Si encontramos el valor, colocarlo en la posición correcta
                if id_original_val is not None:
                    # Asegurar que values tenga el tamaño correcto
                    while len(values) <= idx_id_original:
                        values.append("")
                    values[idx_id_original] = id_original_val
                    # Limpiar elementos extra después de la posición correcta
                    if len(values) > idx_id_original + 1:
                        # Mantener solo hasta idx_id_original + 1
                        values = values[:idx_id_original+1]
                        # Rellenar el resto con vacíos si es necesario
                        while len(values) < len(headers_actuales):
                            values.append("")
        
        # Debug solo en consola
        print(f"🔍 Debug append_row: Sheet={sheet_name}, Headers={len(headers_actuales)}, Headers={headers_actuales}")
        print(f"🔍 Debug append_row: Values={len(values)}, Values={values}")
        
        # Si es COMERCIALES, verificar que ID_Entrega_Original esté en la posición correcta ANTES de append
        if sheet_name == "COMERCIALES":
            idx_id_original = None
            for idx, header in enumerate(headers_actuales):
                if header == "ID_Entrega_Original":
                    idx_id_original = idx
                    break
            
            if idx_id_original is not None:
                # Asegurar que values tenga el tamaño correcto (al menos hasta idx_id_original + 1)
                while len(values) <= idx_id_original:
                    values.append("")
                
                # Buscar el ID_Entrega_Original en values (puede estar en cualquier posición)
                id_original_val = None
                id_original_pos = None
                
                # Primero verificar si ya está en la posición correcta
                if len(values) > idx_id_original:
                    valor_actual = values[idx_id_original]
                    if valor_actual and str(valor_actual).strip() != "":
                        id_original_val = str(valor_actual).strip()
                        id_original_pos = idx_id_original
                        print(f"🔍 Debug append_row: ID_Entrega_Original ya está en posición correcta {idx_id_original}: '{id_original_val}'")
                
                # Si no está en la posición correcta, buscarlo en todas las posiciones
                if id_original_val is None:
                    # Buscar en todas las posiciones de values
                    for pos, val in enumerate(values):
                        if val and str(val).strip() != "":
                            # Verificar si podría ser un ID (número o string numérico)
                            try:
                                val_str = str(val).strip()
                                # Si es un número o string que parece un ID
                                if val_str.isdigit() or (isinstance(val, (int, float)) and val > 0):
                                    # Verificar que no sea el ID principal (posición 0)
                                    if pos != 0:
                                        id_original_val = val_str
                                        id_original_pos = pos
                                        print(f"🔍 Debug append_row: ID_Entrega_Original encontrado en posición {pos}: '{id_original_val}', moviendo a posición {idx_id_original}")
                                        break
                            except:
                                pass
                
                # Si encontramos el valor, colocarlo en la posición correcta
                if id_original_val is not None and id_original_pos != idx_id_original:
                    # Asegurar que values tenga el tamaño correcto
                    while len(values) <= idx_id_original:
                        values.append("")
                    
                    # Colocar el ID en la posición correcta
                    values[idx_id_original] = id_original_val
                    
                    # Si estaba en otra posición, limpiar esa posición
                    if id_original_pos is not None and id_original_pos < len(values) and id_original_pos != idx_id_original:
                        values[id_original_pos] = ""
                    
                    # Asegurar que values tenga exactamente el tamaño de headers
                    while len(values) < len(headers_actuales):
                        values.append("")
                    # Si tiene más elementos, truncar
                    if len(values) > len(headers_actuales):
                        values = values[:len(headers_actuales)]
                    
                    print(f"🔍 Debug append_row: ID_Entrega_Original movido a posición {idx_id_original}: '{id_original_val}'")
                elif id_original_val is None:
                    print(f"⚠️ Debug append_row: ADVERTENCIA - No se encontró ID_Entrega_Original en values. Values: {values}")
        
        ws.append(values)
        success = safe_save_workbook(wb, EXCEL_PATH)
        
        if success:
            print(f"✅ Registro guardado exitosamente en {sheet_name}")
            # Limpiar caché cuando se actualizan datos
            if sheet_name == "LETREROS":
                leer_letreros.clear()
            obtener_estadisticas_registros.clear()
            # Cerrar el workbook explícitamente
            try:
                wb.close()
            except:
                pass
        else:
            print(f"⚠️ Registro NO se pudo guardar en {sheet_name}")
        
        return success
    except PermissionError as e:
        st.error(f"🔒 **Error al guardar**: No se pudo guardar el archivo Excel. OneDrive puede estar sincronizando.\n\n**Intenta:**\n1. Cierra todas las ventanas de Excel\n2. Pausa OneDrive unos segundos\n3. Vuelve a intentar")
        print(f"❌ Error de permisos al guardar: {e}")
        return False
    except Exception as e:
        st.error(f"❗ **Error inesperado al guardar**: {e}\n\nPor favor, intenta de nuevo en unos segundos.")
        print(f"❌ Error inesperado: {e}")
        return False

@st.cache_data(ttl=120)  # Cache por 2 minutos
def obtener_estadisticas_registros():
    """Obtener estadísticas de registros guardados con caché"""
    stats = {}
    try:
        if EXCEL_PATH.exists():
            wb = safe_load_workbook(EXCEL_PATH)
            for sheet_name in HEADERS.keys():
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    stats[sheet_name] = ws.max_row - 1
                else:
                    stats[sheet_name] = 0
            wb.close()
    except:
        pass
    return stats

def normalizar_nombre(nombre):
    """Normalizar nombre para comparación"""
    if not nombre:
        return ""
    nombre_str = str(nombre).upper()
    # Eliminar espacios extra
    nombre_str = ' '.join(nombre_str.split())
    return nombre_str

def buscar_vinculo_cliente(nombre_venta, clientes):
    """Buscar el cliente que más coincida con el nombre de venta"""
    if not nombre_venta or not clientes:
        return None
    
    nombre_norm = normalizar_nombre(nombre_venta)
    mejor_coincidencia = None
    mejor_score = 0
    
    for cliente in clientes:
        # Comparar con nombre fiscal
        nombre_cliente = normalizar_nombre(cliente.get('nombre', ''))
        if nombre_norm == nombre_cliente or nombre_norm in nombre_cliente or nombre_cliente in nombre_norm:
            mejor_coincidencia = cliente['nombre']
            mejor_score = 100
            break
        # Calcular similitud simple
        palabras_venta = set(nombre_norm.split())
        palabras_cliente = set(nombre_cliente.split())
        if palabras_venta and palabras_cliente:
            score = len(palabras_venta & palabras_cliente) / len(palabras_venta | palabras_cliente) * 100
            if score > mejor_score and score > 50:  # Al menos 50% de coincidencia
                mejor_score = score
                mejor_coincidencia = cliente['nombre']
    
    return mejor_coincidencia if mejor_score > 50 else None

def vincular_ventas_con_clientes(file_path=None):
    """Vincular datos de ventas con clientes y crear reporte"""
    try:
        # Cargar clientes
        clientes = cargar_clientes()
        if not clientes:
            return False, "No se encontraron clientes en la base de datos"
        
        # Cargar datos de ventas
        if EXCEL_PATH.exists():
            wb = safe_load_workbook(EXCEL_PATH)
            if "VENTAS" not in wb.sheetnames:
                wb.close()
                return False, "Primero importa los datos de ventas"
            
            ws_ventas = wb["VENTAS"]
            
            # Leer datos de ventas
            datos_ventas = []
            for row in ws_ventas.iter_rows(min_row=2, values_only=True):
                if row and row[0]:  # Si hay empresa
                    datos_ventas.append({
                        'Empresa': str(row[0]).strip() if row[0] else "",
                        'Venta_Bruta': row[1] if len(row) > 1 else 0
                    })
            
            wb.close()
            
            if not datos_ventas:
                return False, "No hay datos de ventas para vincular"
            
            # Crear reporte de vinculación
            reporte = []
            clientes_vincular = []
            total_ventas = 0
            
            for venta in datos_ventas:
                empresa_venta = venta['Empresa']
                monto = venta['Venta_Bruta']
                
                # Buscar cliente vinculado
                cliente_encontrado = buscar_vinculo_cliente(empresa_venta, clientes)
                
                if cliente_encontrado:
                    clientes_vincular.append({
                        'Cliente': cliente_encontrado,
                        'Empresa_Venta': empresa_venta,
                        'Venta_Bruta': monto
                    })
                    total_ventas += float(monto) if monto else 0
                else:
                    reporte.append(f"⚠️ No encontrado: {empresa_venta}")
            
            # Guardar en hoja VENTAS_VINCULADAS
            wb_registro = safe_load_workbook(EXCEL_PATH)
            
            if "VENTAS_VINCULADAS" in wb_registro.sheetnames:
                wb_registro.remove(wb_registro["VENTAS_VINCULADAS"])
            
            ws_vinculadas = wb_registro.create_sheet("VENTAS_VINCULADAS")
            ws_vinculadas.append(['Cliente', 'Venta_Bruta', 'Empresa_Venta'])
            
            for vinculo in clientes_vincular:
                ws_vinculadas.append([
                    vinculo['Cliente'],
                    vinculo['Venta_Bruta'],
                    vinculo['Empresa_Venta']
                ])
            
            if safe_save_workbook(wb_registro, EXCEL_PATH, tries=20, wait=0.2):
                wb_registro.close()
                obtener_estadisticas_registros.clear()
                
                mensaje = f"✅ {len(clientes_vincular)} ventas vinculadas con clientes existentes\n"
                mensaje += f"💰 Total de ventas: ${total_ventas:,.2f}\n"
                if reporte:
                    mensaje += f"\n⚠️ {len(reporte)} empresas no pudieron vincularse automáticamente"
                
                return True, mensaje
            else:
                wb_registro.close()
                return False, "Error al guardar el archivo"
        else:
            return False, "No se encontró el archivo REGISTRO_MARKETING.xlsx"
            
    except Exception as e:
        return False, f"Error: {str(e)}"

def copiar_datos_ventas(file_path=None):
    """Copiar datos de VENTAS POR PRODUCTOS.xlsx a REGISTRO_MARKETING.xlsx"""
    try:
        # Usar archivo subido o archivo del sistema
        if file_path:
            archivo_ventas = file_path
        elif VENTAS_PATH.exists():
            archivo_ventas = VENTAS_PATH
        else:
            return False, "No se encontró el archivo 'VENTAS POR PRODUCTOS.xlsx'. Por favor súbelo primero."
        
        # Leer con openpyxl para buscar headers
        wb_temp = None
        idx_nombre_fiscal = None
        idx_venta_bruta = None
        header_row = 1
        datos_ventas = []
        
        try:
            if file_path:
                wb_temp = load_workbook(file_path, read_only=True)
            else:
                wb_temp = load_workbook(VENTAS_PATH, read_only=True)
            
            if "VENTA POR PRECIO" not in wb_temp.sheetnames:
                wb_temp.close()
                return False, "No se encontró la hoja 'VENTA POR PRECIO' en el archivo"
            
            ws_temp = wb_temp["VENTA POR PRECIO"]
            
            # Buscar headers - las columnas A y B donde están los datos
            # Columna A (índice 0) = nombre_fiscal
            # Columna B (índice 1) = suma de venta_bruta
            for row_num in range(1, min(11, ws_temp.max_row + 1)):  # Buscar hasta fila 10
                row = list(ws_temp.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
                for col_idx, cell_val in enumerate(row[:5]):  # Solo revisar primeras 5 columnas
                    if cell_val and isinstance(cell_val, str):
                        cell_lower = cell_val.lower()
                        # Buscar header de nombre_fiscal o v_empresa en columna A
                        if col_idx == 0 and ('nombre_fiscal' in cell_lower or 'v_empresa' in cell_lower or 'empresa' in cell_lower):
                            idx_nombre_fiscal = 0  # Columna A
                            header_row = row_num
                        # Buscar header de venta_bruta o suma en columna B
                        elif col_idx == 1 and ('venta_bruta' in cell_lower or 'suma' in cell_lower or 'bruta' in cell_lower):
                            idx_venta_bruta = 1  # Columna B
                            header_row = row_num
                
                if idx_nombre_fiscal is not None and idx_venta_bruta is not None:
                    break
            
            # Si no encontramos por nombre, usar las columnas A y B directamente
            if idx_nombre_fiscal is None:
                idx_nombre_fiscal = 0  # Columna A
            if idx_venta_bruta is None:
                idx_venta_bruta = 1  # Columna B
            
            # Si encontramos los índices, leer los datos
            if idx_nombre_fiscal is not None and idx_venta_bruta is not None:
                for row in ws_temp.iter_rows(min_row=header_row + 1, max_row=min(ws_temp.max_row + 1, header_row + 10001)):
                    try:
                        empresa = row[idx_nombre_fiscal].value
                        venta = row[idx_venta_bruta].value
                        if empresa and venta:
                            datos_ventas.append({
                                'Empresa': str(empresa).strip(),
                                'Venta_Bruta': venta
                            })
                    except:
                        continue
            
            wb_temp.close()
            
        except Exception as e:
            return False, f"Error al leer el archivo: {str(e)}"
        
        if not datos_ventas:
            return False, f"No se encontraron datos válidos. Verifica que las columnas 'v_empresa.nombre_fiscal' y 'Suma de venta_bruta' existan en la hoja 'VENTA POR PRECIO'."
        
        # Abrir archivo de registro de marketing (con menos reintentos)
        wb_registro = safe_load_workbook(EXCEL_PATH, tries=3, wait=0.1)
        
        # Crear o limpiar hoja VENTAS
        if "VENTAS" in wb_registro.sheetnames:
            wb_registro.remove(wb_registro["VENTAS"])
        
        ws_ventas_registro = wb_registro.create_sheet("VENTAS")
        
        # Escribir headers
        ws_ventas_registro.append(['Empresa', 'Venta_Bruta'])
        
        # Escribir datos (por lotes para mejor rendimiento)
        for dato in datos_ventas:
            ws_ventas_registro.append([dato['Empresa'], dato['Venta_Bruta']])
        
        # Guardar con menos reintentos
        if safe_save_workbook(wb_registro, EXCEL_PATH, tries=20, wait=0.2):
            wb_registro.close()
            # Limpiar caché
            obtener_estadisticas_registros.clear()
            return True, f"✅ {len(datos_ventas)} registros copiados exitosamente"
        else:
            wb_registro.close()
            return False, "Error al guardar el archivo. Verifica que el archivo no esté abierto."
            
    except PermissionError as e:
        return False, "El archivo está bloqueado. Cierra Excel y OneDrive, luego intenta de nuevo."
    except Exception as e:
        print(f"Error al copiar datos de ventas: {e}")
        import traceback
        traceback.print_exc()
        return False, f"Error: {str(e)}"

# ===== SIDEBAR MENÚ =====
with st.sidebar:
    # Header del sidebar con logo
    logo_path = Path(EXCEL_DIR) / "logo_extrememax.png"
    logo_shown = False
    
    # Intentar cargar el logo usando HTML
    if logo_path.exists() and logo_path.is_file():
        try:
            # Leer imagen y convertir a base64
            with open(logo_path, "rb") as img_file:
                img_data = img_file.read()
                img_base64 = base64.b64encode(img_data).decode()
                st.markdown(f"""
                <div style="text-align: center; margin-bottom: 2rem;">
                    <img src="data:image/png;base64,{img_base64}" style="max-width: 300px; width: 100%; height: auto;" />
                </div>
                """, unsafe_allow_html=True)
                logo_shown = True
        except:
            pass
    
    # Si no se pudo mostrar el logo, mostrar el banner
    if not logo_shown:
        st.markdown("""
    <div style="
        background: linear-gradient(135deg, #FF9A00 0%, #FF6D00 100%);
        padding: 1.5rem;
        border-radius: 15px;
        margin-bottom: 2rem;
        text-align: center;
        box-shadow: 0 5px 15px rgba(255, 154, 0, 0.3);
    ">
        <h2 style="
            color: #ffffff;
            font-family: 'Poppins', sans-serif;
            font-weight: 800;
            font-size: 1.5rem;
            margin: 0;
        ">🎯 Marketing Extrememax</h2>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    
    # Menú de navegación
    st.markdown("#### 📋 Menú Principal")
    
    menu_options = [
        "👥 CLIENTES",
        "📋 GESTION LETREROS",
        "🎉 ACTIVACIONES Y EVENTOS",
        "📰 ENTREGA DE PUBLICIDAD",
        "📦 ENTREGA DE PERCHAS/EXHIBIDORES",
        "💼 ENTREGA A COMERCIALES",
        "📊 INVENTARIO",
        "📊 REPORTES GENERALES"
    ]
    
    selected_menu = st.radio(
        "Selecciona una sección:",
        menu_options,
        index=0
    )
    
    st.markdown("---")
    
    # Sección de administración
    st.markdown("#### 🔧 Administración de Datos")
    
    # Instrucciones
    with st.expander("ℹ️ ¿Cómo funciona el proceso?", expanded=False):
        st.markdown("""
        **📋 Pasos para actualizar las ventas de clientes:**
        
        1. **📤 Sube el Excel** "VENTAS POR PRODUCTOS.xlsx" (botón arriba)
        2. **📥 Haz clic en "Importar Datos"** - Copia los datos a REGISTRO_MARKETING.xlsx
        3. **🔗 Haz clic en "Vincular con Clientes"** - Compara nombres y vincula las ventas
        4. **✅ Automático:** La columna "Venta Total" se actualiza en la tabla de clientes
        
        **💡 Tip:** Puedes repetir este proceso cada vez que tengas nuevos datos de ventas.
        """)
    
    # Uploader para archivo de ventas
    uploaded_ventas = st.file_uploader(
        "📤 Subir VENTAS POR PRODUCTOS.xlsx",
        type=['xlsx'],
        help="Paso 1: Sube el archivo Excel con los datos de ventas de la hoja 'VENTA POR PRECIO'",
        key="upload_ventas"
    )
    
    if uploaded_ventas is not None:
        # Guardar temporalmente el archivo
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(uploaded_ventas.getbuffer())
            tmp_path = tmp_file.name
        
        col_import, col_vincular = st.columns(2)
        
        with col_import:
            if st.button("📥 Paso 2: Importar Datos", use_container_width=True, type="primary"):
                with st.spinner("Importando datos de ventas (esto puede tardar unos segundos)..."):
                    try:
                        success, message = copiar_datos_ventas(tmp_path)
                        if success:
                            st.success(message)
                            st.info("✅ Datos importados. Ahora haz clic en 'Vincular con Clientes' para actualizar las ventas.")
                            st.rerun()
                        else:
                            st.error(message)
                    finally:
                        try:
                            os.unlink(tmp_path)
                        except:
                            pass
        
        with col_vincular:
            if st.button("🔗 Paso 3: Vincular con Clientes", use_container_width=True):
                with st.spinner("Vinculando ventas con clientes existentes..."):
                    success, message = vincular_ventas_con_clientes()
                    if success:
                        st.success(message)
                        st.info("🎉 ¡Actualización completa! Ve a la sección 'CLIENTES' para ver los totales actualizados.")
                        st.rerun()
                    else:
                        st.error(message)
    else:
        # Si no hay archivo subido, intentar con el del sistema
        col_import, col_vincular = st.columns(2)
        
        with col_import:
            if st.button("📥 Paso 2: Importar Datos (Local)", use_container_width=True, type="primary"):
                with st.spinner("Importando datos de ventas (esto puede tardar unos segundos)..."):
                    success, message = copiar_datos_ventas()
                    if success:
                        st.success(message)
                        st.info("✅ Datos importados. Ahora haz clic en 'Vincular con Clientes' para actualizar las ventas.")
                        st.rerun()
                    else:
                        st.error(message)
        
        with col_vincular:
            if st.button("🔗 Paso 3: Vincular con Clientes", use_container_width=True):
                with st.spinner("Vinculando ventas con clientes existentes..."):
                    success, message = vincular_ventas_con_clientes()
                    if success:
                        st.success(message)
                        st.info("🎉 ¡Actualización completa! Ve a la sección 'CLIENTES' para ver los totales actualizados.")
                        st.rerun()
                    else:
                        st.error(message)
    
    st.markdown("---")
    
    # Sección de archivos
    st.markdown("#### 📁 Gestión de Archivos")
    
    uploaded_file = st.file_uploader(
        "Cargar EMPRESAS.xlsx",
        type=['xlsx'],
        help="Sube tu archivo Excel con los datos de clientes"
    )
    
    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file)
            
            st.success(f"✅ **{len(df)}** registros cargados")
            
            if st.button("📥 Copiar a Base de Datos", type="primary", use_container_width=True):
                with st.spinner("Procesando..."):
                    uploaded_file.seek(0)
                    file_bytes = uploaded_file.read()
                    with open(EMPRESAS_PATH, 'wb') as f:
                        f.write(file_bytes)
                    
                    copiar_base_datos_clientes()
                
                st.success("✅ Base de datos actualizada!")
                st.rerun()
        except Exception as e:
            st.error(f"Error: {str(e)}")
    
    st.divider()
    
    # Botón para recargar
    if st.button("🔄 Copiar Base de Datos a Excel", use_container_width=True):
        with st.spinner("Copiando datos de EMPRESAS.xlsx a REGISTRO_MARKETING.xlsx..."):
            if copiar_base_datos_clientes():
                st.success("✅ Base de datos copiada exitosamente!")
            else:
                st.error("❌ Error al copiar la base de datos. Verifica que el archivo no esté abierto.")
    
    st.caption("💡 Copia los datos de clientes al Excel para tenerlos disponibles en todas las hojas")
    
    st.divider()
    
    # Información importante
    st.markdown("#### ⚠️ Solución de Problemas")
    with st.expander("🔓 Si no se guardan los datos..."):
        st.markdown("""
        **Si intentas guardar información y no se guarda:**
        
        1. **Cierra Excel** si está abierto
        2. **Espera 5-10 segundos** para que OneDrive termine de sincronizar
        3. **Vuelve a intentar** guardar
        4. El sistema automáticamente reintentará hasta 100 veces
        
        **Si el problema persiste:**
        - Pausa la sincronización de OneDrive temporalmente
        - O espera a que OneDrive termine de sincronizar completamente
        """)
    
    st.divider()
    
    # Descargar Excel
    if EXCEL_PATH.exists():
        try:
            with open(EXCEL_PATH, 'rb') as f:
                st.download_button(
                    label="⬇️ Descargar REGISTRO_MARKETING.xlsx",
                    data=f.read(),
                    file_name=EXCEL_FILE,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True
                )
        except Exception as e:
            st.error(f"Error al leer archivo: {e}")
    
    st.divider()
    
    # Info del sistema
    st.markdown("#### ℹ️ Información")
    if EMPRESAS_PATH.exists():
        st.success("✅ EMPRESAS.xlsx conectado")
    else:
        st.warning("⚠️ Archivo no encontrado")

# Cargar clientes
clientes = cargar_clientes()

# Cargar agentes
agentes = cargar_agentes()

# Solo copiar base de datos si es necesario (no en cada recarga)
# El usuario puede activar esto manualmente desde el sidebar

# Banner de sección actual
st.markdown(f"""
<div style="
    background: linear-gradient(135deg, rgba(255, 154, 0, 0.1), rgba(255, 154, 0, 0.05));
    padding: 1rem 1.5rem;
    border-radius: 12px;
    border-left: 5px solid #FF9A00;
    margin-bottom: 2rem;
    box-shadow: 0 4px 15px rgba(255, 154, 0, 0.1);
">
    <h2 style="
        color: #1a1a1a;
        font-family: 'Poppins', sans-serif;
        font-weight: 700;
        font-size: 1.8rem;
        margin: 0;
    ">{selected_menu}</h2>
</div>
""", unsafe_allow_html=True)

# Función auxiliar para formularios de cliente
def form_cliente_section(clientes, key_prefix, con_buscador=False):
    """Mostrar selector de cliente con información"""
    if clientes:
        # Solo mostrar buscador si se solicita
        if con_buscador:
            search_input = st.text_input(
                "¿Qué estás buscando?", 
                key=f"{key_prefix}_search_input",
                placeholder="Nombre, ID, RUC, cédula, teléfono...",
                help="Busca por cualquier campo del cliente"
            )
            
            # Filtrar clientes según búsqueda
            if search_input:
                clientes_filtrados = [c for c in clientes if buscar_cliente_inteligente(c, search_input)]
            else:
                clientes_filtrados = clientes
            
            # Mostrar contador de resultados si hay búsqueda
            if search_input:
                st.caption(f"📊 {len(clientes_filtrados)} cliente(s) encontrado(s) de {len(clientes)}")
        else:
            clientes_filtrados = clientes
        
        # Crear opciones con nombre fiscal y ciudad
        opciones_clientes = [""] + [f"{c['id']} - {c['nombre']} ({c['ciudad']})" for c in clientes_filtrados]
        
        cliente_seleccionado = st.selectbox("Selecciona el Cliente *", opciones_clientes, key=f"{key_prefix}_cliente")
        
        # Extraer información del cliente correctamente y validar
        if cliente_seleccionado and len(cliente_seleccionado) > 0:
            try:
                # Extraer el ID (antes del " - ")
                cliente_id = cliente_seleccionado.split(" - ")[0].strip()
                
                # Validar que el cliente existe en la lista completa de clientes
                cliente_info = next((c for c in clientes if c['id'] == cliente_id), None)
                
                if cliente_info:
                    # Buscar monto de ventas del cliente
                    monto_ventas = 0
                    try:
                        if EXCEL_PATH.exists():
                            wb_ventas = safe_load_workbook(EXCEL_PATH, tries=3, wait=0.1)
                            cliente_norm = normalizar_nombre(cliente_info['nombre'])
                            # Preferir VENTAS_VINCULADAS
                            if "VENTAS_VINCULADAS" in wb_ventas.sheetnames:
                                ws_v = wb_ventas["VENTAS_VINCULADAS"]
                                for row in ws_v.iter_rows(min_row=2, values_only=True):
                                    if row and row[0]:
                                        if cliente_norm == normalizar_nombre(str(row[0]).strip()):
                                            try:
                                                monto_ventas += float(row[1]) if row[1] else 0
                                            except:
                                                pass
                            # Fallback: VENTAS (Empresa, Venta_Bruta)
                            elif "VENTAS" in wb_ventas.sheetnames:
                                ws_v = wb_ventas["VENTAS"]
                                for row in ws_v.iter_rows(min_row=2, values_only=True):
                                    if row and row[0]:
                                        if cliente_norm == normalizar_nombre(str(row[0]).strip()):
                                            try:
                                                monto_ventas += float(row[1]) if len(row) > 1 and row[1] else 0
                                            except:
                                                pass
                            wb_ventas.close()
                    except:
                        pass
                    
                    # Mostrar monto de compra visiblemente al seleccionar (también si es 0)
                    st.info(f"💰 **Venta Total del Cliente:** ${monto_ventas:,.2f}")
                    
                    # Mostrar información expandible
                    with st.expander("📋 Información del Cliente", expanded=False):
                        col_a, col_b = st.columns(2)
                        col_a.markdown(f"**ID:** {cliente_info['id']}")
                        col_a.markdown(f"**Nombre Fiscal:** {cliente_info['nombre']}")
                        if cliente_info.get('nombre_comercial') and cliente_info['nombre_comercial'] != 'nan' and cliente_info['nombre_comercial']:
                            col_a.markdown(f"**Local/Negocio:** {cliente_info['nombre_comercial']}")
                        if cliente_info.get('identificacion'):
                            col_a.markdown(f"**Identificación:** {cliente_info['identificacion']}")
                        col_a.markdown(f"**Teléfono:** {cliente_info['telefono']}")
                        col_b.markdown(f"**Ciudad:** {cliente_info['ciudad']}")
                        col_b.markdown(f"**Provincia:** {cliente_info['provincia']}")
                        if cliente_info.get('direccion'):
                            col_b.markdown(f"**Dirección:** {cliente_info['direccion']}")
                        col_b.markdown(f"**Agente:** {cliente_info['agente']}")
                        
                        # Agregar monto de ventas
                        st.markdown("---")
                        st.markdown(f"**💰 Venta Total del Cliente:** ${monto_ventas:,.2f}")
                    
                    # Guardar el agente del cliente en session_state para uso en otros campos
                    agente_cliente = cliente_info.get('agente', '')
                    if agente_cliente and agente_cliente.strip() and agente_cliente != 'nan':
                        # Guardar en session_state con la key_prefix para evitar conflictos entre formularios
                        st.session_state[f"{key_prefix}_cliente_agente"] = agente_cliente.strip()
                    else:
                        st.session_state[f"{key_prefix}_cliente_agente"] = ""
                    
                    # Retornar el NOMBRE del cliente en lugar del ID
                    cliente = cliente_info['nombre']
                else:
                    # Si no se encuentra, extraer el nombre del texto seleccionado
                    try:
                        # El formato es "ID - Nombre (Ciudad)"
                        partes = cliente_seleccionado.split(" - ")
                        if len(partes) > 1:
                            nombre_completo = partes[1].split(" (")[0].strip()
                            cliente = nombre_completo
                        else:
                            cliente = cliente_id
                    except:
                        cliente = cliente_id
            except Exception as e:
                st.error(f"⚠️ Error al procesar: {str(e)}")
                cliente = ""
        else:
            cliente = ""
            # Limpiar el agente si no hay cliente seleccionado
            if f"{key_prefix}_cliente_agente" in st.session_state:
                st.session_state[f"{key_prefix}_cliente_agente"] = ""
    else:
        cliente = st.text_input("Nombre del Cliente *", key=f"{key_prefix}_cliente")
        # Limpiar el agente si no hay lista de clientes
        if f"{key_prefix}_cliente_agente" in st.session_state:
            st.session_state[f"{key_prefix}_cliente_agente"] = ""
    
    return cliente

# ===== CONTENIDO SEGÚN MENÚ SELECCIONADO =====
if selected_menu == "👥 CLIENTES":
    st.markdown("### 📊 Total de CLIENTES")
    
    if clientes:
        # Estadísticas principales
        ciudades = {}
        provincias = {}
        agentes = {}
        
        for cliente in clientes:
            ciudad = cliente.get('ciudad', 'Sin ciudad')
            provincia = cliente.get('provincia', 'Sin provincia')
            agente = cliente.get('agente', 'Sin agente')
            
            if ciudad and ciudad != 'nan' and ciudad != '':
                ciudades[ciudad] = ciudades.get(ciudad, 0) + 1
            if provincia and provincia != 'nan' and provincia != '':
                provincias[provincia] = provincias.get(provincia, 0) + 1
            if agente and agente != 'nan' and agente != '':
                agentes[agente] = agentes.get(agente, 0) + 1
        
        # Buscador y Filtros - Expandible (ARRIBA)
        with st.expander("🔍 Buscar y Filtrar CLIENTES", expanded=False):
            col_search, col_filter1, col_filter2, col_filter3 = st.columns(4)
            
            # Buscador inteligente
            search_term = col_search.text_input("¿Qué estás buscando?", key="search_clientes", help="Busca por nombre, RUC, cédula, teléfono, ID...")
            
            # Filtros
            ciudades_list = [""] + sorted(list(set([c.get('ciudad', '') for c in clientes if c.get('ciudad', '') and c.get('ciudad', '') != 'nan'])))
            provincias_list = [""] + sorted(list(set([c.get('provincia', '') for c in clientes if c.get('provincia', '') and c.get('provincia', '') != 'nan'])))
            agentes_list = [""] + sorted(list(set([c.get('agente', '') for c in clientes if c.get('agente', '') and c.get('agente', '') != 'nan'])))
            
            filter_ciudad = col_filter1.selectbox("📍 Filtrar por Ciudad", ciudades_list, key="filter_ciudad")
            filter_provincia = col_filter2.selectbox("🌍 Filtrar por Provincia", provincias_list, key="filter_provincia")
            filter_agente = col_filter3.selectbox("👤 Filtrar por Agente", agentes_list, key="filter_agente")
            
            # Aplicar filtros
            clientes_filtrados = clientes.copy()
            
            # Aplicar búsqueda inteligente
            if search_term:
                clientes_filtrados = [c for c in clientes_filtrados if buscar_cliente_inteligente(c, search_term)]
            
            if filter_ciudad:
                clientes_filtrados = [c for c in clientes_filtrados if c.get('ciudad', '') == filter_ciudad]
            
            if filter_provincia:
                clientes_filtrados = [c for c in clientes_filtrados if c.get('provincia', '') == filter_provincia]
            
            if filter_agente:
                clientes_filtrados = [c for c in clientes_filtrados if c.get('agente', '') == filter_agente]
            
            # Mostrar resultados
            st.markdown(f"**Mostrando {len(clientes_filtrados)} de {len(clientes)} clientes**")
            
            if clientes_filtrados:
                # Cargar ventas vinculadas para cada cliente
                ventas_por_cliente = {}
                try:
                    if EXCEL_PATH.exists():
                        wb_ventas = safe_load_workbook(EXCEL_PATH, tries=3, wait=0.1)
                        if "VENTAS_VINCULADAS" in wb_ventas.sheetnames:
                            ws_ventas = wb_ventas["VENTAS_VINCULADAS"]
                            for row in ws_ventas.iter_rows(min_row=2, values_only=True):
                                if row and row[0]:  # Si hay cliente
                                    cliente_nombre = str(row[0]).strip()
                                    venta_valor = row[1] if len(row) > 1 else 0
                                    try:
                                        venta_valor = float(venta_valor) if venta_valor else 0
                                    except:
                                        venta_valor = 0
                                    
                                    # Normalizar nombre para comparar
                                    cliente_norm = normalizar_nombre(cliente_nombre)
                                    if cliente_norm not in ventas_por_cliente:
                                        ventas_por_cliente[cliente_norm] = 0
                                    ventas_por_cliente[cliente_norm] += venta_valor
                        wb_ventas.close()
                except:
                    pass
                
                # Agregar venta total a cada cliente
                for cliente in clientes_filtrados:
                    nombre_cliente = normalizar_nombre(cliente.get('nombre', ''))
                    cliente['venta_total'] = ventas_por_cliente.get(nombre_cliente, 0)
                
                # Mostrar clientes en tabla
                df_clientes = pd.DataFrame(clientes_filtrados)
                
                # Reordenar columnas para que venta_total esté al final (si existe)
                columnas_orden = ['id', 'nombre', 'nombre_comercial', 'identificacion', 'telefono', 
                                 'ciudad', 'provincia', 'direccion', 'agente', 'venta_total']
                columnas_existentes = [col for col in columnas_orden if col in df_clientes.columns]
                df_clientes = df_clientes[columnas_existentes]
                
                st.dataframe(
                    df_clientes,
                    column_config={
                        "id": "ID",
                        "nombre": "Nombre Fiscal",
                        "nombre_comercial": "Local/Negocio",
                        "identificacion": "Identificacion",
                        "telefono": "Teléfono",
                        "ciudad": "Ciudad",
                        "provincia": "Provincia",
                        "direccion": "Direccion",
                        "agente": "Agente",
                        "venta_total": st.column_config.NumberColumn(
                            "💰 Venta Total",
                            help="Total de ventas brutas del cliente",
                            format="$%.2f"
                        )
                    },
                    hide_index=True,
                    use_container_width=True
                )
            else:
                st.info("No se encontraron clientes con los filtros aplicados")
        
        st.markdown("---")
        
        # Métricas principales
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("👥 Total CLIENTES", len(clientes), delta=None)
        col2.metric("🏙️ Ciudades", len(ciudades), delta=None)
        col3.metric("📍 Provincias", len(provincias), delta=None)
        col4.metric("👤 Agentes", len(agentes), delta=None)
        
        st.markdown("---")
        
        # Top ciudades
        col_left, col_right = st.columns([2, 1])
        
        with col_left:
            st.markdown("#### 📍 Top Ciudades")
            if ciudades:
                ciudades_sorted = sorted(ciudades.items(), key=lambda x: x[1], reverse=True)[:10]
                for ciudad, cantidad in ciudades_sorted:
                    porcentaje = (cantidad / len(clientes)) * 100
                    st.progress(porcentaje / 100, text=f"**{ciudad}**: {cantidad} clientes ({porcentaje:.1f}%)")
            else:
                st.info("No hay datos de ciudades")
        
        with col_right:
            st.markdown("#### 📈 Estadísticas de Registros")
            stats = obtener_estadisticas_registros()
            st.metric("📦 Perchas", stats.get("PERCHAS", 0))
            st.metric("📋 Letreros", stats.get("LETREROS", 0))
            st.metric("🎉 Eventos", stats.get("EVENTOS", 0))
        
        # Mostrar datos de ventas si existen
        st.markdown("---")
        try:
            if EXCEL_PATH.exists():
                wb = safe_load_workbook(EXCEL_PATH)
                if "VENTAS" in wb.sheetnames:
                    ws_ventas = wb["VENTAS"]
                    
                    # Leer datos
                    datos_ventas = []
                    for row in ws_ventas.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:  # Si hay empresa
                            datos_ventas.append({
                                'Empresa': row[0],
                                'Venta_Bruta': row[1] if len(row) > 1 else 0
                            })
                    wb.close()
                    
                    if datos_ventas:
                        st.markdown("#### 💰 Datos de Ventas Importados")
                        df_ventas = pd.DataFrame(datos_ventas)
                        st.dataframe(df_ventas, use_container_width=True, height=300)
                        
                        total_ventas = sum([float(str(v.get('Venta_Bruta', 0)).replace(',', '').replace('$', '')) for v in datos_ventas if v.get('Venta_Bruta')])
                        st.metric("💰 Total en Ventas Brutas", f"${total_ventas:,.2f}")
        except Exception as e:
            pass
            
    else:
        st.warning("⚠️ No se encontró EMPRESAS.xlsx. Por favor, carga el archivo desde el sidebar.")
        st.info("💡 **Instrucciones:** \n1. Haz clic en 'Cargar EMPRESAS.xlsx' en el sidebar \n2. Selecciona tu archivo \n3. Haz clic en 'Copiar a Base de Datos'")

# ===== SECCIÓN ENTREGA DE LETREROS =====
elif selected_menu == "📋 GESTION LETREROS":
    # Leer letreros primero (se usa en ambas secciones)
    letreros = leer_letreros()
    
    # Filtrar letreros por estado (normalizando para evitar problemas con espacios)
    letreros_proceso = [l for l in letreros if str(l.get("Estado", "")).strip().upper() == "EN PROCESO"]
    letreros_pendientes = [l for l in letreros if str(l.get("Estado", "")).strip().upper() == "PENDIENTE"]
    letreros_entregados = [l for l in letreros if str(l.get("Estado", "")).strip().upper() == "ENTREGADO"]
    
    # Tabs para organizar
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["⏳ En Proceso", "⏸️ Pendientes", "📅 Fecha de Fabricación", "✅ Finalizados", "📈 Reportes"])
    
    with tab1:
        # Formulario de solicitud de letrero
        st.markdown("### 📋 SOLICITUD DE REALIZACION DE LETRERO")
        
        with st.form("form_letreros", clear_on_submit=False):
            col1, col2 = st.columns(2)
            
            with col1:
                cliente = form_cliente_section(clientes, "l")
            
            with col2:
                tipo_solicitud = st.selectbox(
                    "Tipo de Solicitud *",
                    ["", "NUEVO", "ACTUALIZACIÓN"],
                    key="l_tipo_solicitud",
                    help="¿Es un letrero nuevo o una actualización/reemplazo de uno existente?"
                )
            
            col2_1, col2_2 = st.columns(2)
            with col2_1:
                tipo_letrero_selected = st.selectbox(
                    "Tipo de Letrero *",
                    ["", "LETRERO CON ESTRUCTURA", "MICROPERFORADO", "VINIL", "LONA", "ROMPETRAFICO", "PINTURA DE LOCAL", "OTRO"],
                    key="l_tipo_letrero"
                )
            
            with col2_2:
                fecha = st.date_input("Fecha *", key="l_fecha")
            
            # Campo para escribir tipo personalizado
            otro_tipo = st.text_input("Otro (especificar tipo)", key="l_otro_tipo", placeholder="Si no encuentras el tipo, escríbelo aquí")
            
            # Determinar el valor final del tipo de letrero
            if otro_tipo and otro_tipo.strip():
                tipo_letrero = otro_tipo
            else:
                tipo_letrero = tipo_letrero_selected
            
            col3, col4 = st.columns(2)
            medidas = col3.text_input("Medidas (cm) *", placeholder="Ej: 100x50cm", key="l_medidas")
            costo = col4.number_input("Costo ($)", min_value=0.0, value=0.0, step=0.01, key="l_costo")

            col5, col6 = st.columns(2)
            
            # Obtener el agente del cliente seleccionado si existe
            agente_default = ""
            if "l_cliente_agente" in st.session_state and st.session_state["l_cliente_agente"]:
                agente_cliente = st.session_state["l_cliente_agente"]
                # Buscar si el agente está en la lista de agentes
                if agente_cliente in agentes:
                    agente_default = agente_cliente
                # También buscar por coincidencia parcial o similar
                elif agentes:
                    for agente in agentes:
                        if agente and (agente_cliente.lower() in str(agente).lower() or str(agente).lower() in agente_cliente.lower()):
                            agente_default = agente
                            break
            
            # Configurar el agente por defecto basado en el cliente seleccionado
            # Si hay un agente del cliente y existe en la lista, actualizar session_state
            if agente_default and agente_default in agentes:
                # Solo actualizar si no hay un valor previo o si el cliente cambió
                if "l_comercial" not in st.session_state or st.session_state.get("l_cliente_anterior") != cliente:
                    st.session_state["l_comercial"] = agente_default
                    st.session_state["l_cliente_anterior"] = cliente
            elif not cliente:
                # Si no hay cliente seleccionado, limpiar el agente
                if "l_comercial" in st.session_state and st.session_state.get("l_cliente_anterior"):
                    st.session_state["l_comercial"] = agentes[0] if agentes else ""
                    st.session_state["l_cliente_anterior"] = ""
            
            # Autocompletar Comercial/Agente en base al cliente seleccionado.
            # Usamos una clave dinámica para forzar que el selectbox se regenere cuando cambie el cliente
            agente_cliente = st.session_state.get("l_cliente_agente", "")
            indice_agente = 0
            if agentes:
                # Coincidencia exacta
                try:
                    if agente_cliente and agente_cliente in agentes:
                        indice_agente = agentes.index(agente_cliente)
                    else:
                        # Coincidencia parcial, sin sensibilidad a mayúsculas
                        agente_norm = str(agente_cliente).strip().lower()
                        for i, a in enumerate(agentes):
                            a_norm = str(a).strip().lower()
                            if agente_norm and (agente_norm in a_norm or a_norm in agente_norm):
                                indice_agente = i
                                break
                except Exception:
                    indice_agente = 0
            comercial_key = f"l_comercial_{(agente_cliente or 'vacio')[:24]}"
            comercial = col5.selectbox("Comercial/Agente", agentes, key=comercial_key, index=indice_agente if agentes else 0)
            estado = "EN PROCESO"  # Estado inicial fijo
            
            observaciones = st.text_area("Observaciones", key="l_obs", height=100)
            st.info("ℹ️ Los letreros ingresados comienzan con estado 'EN PROCESO'. Puedes gestionar su estado después desde la sección de gestión.")
            
            # Verificar letreros existentes antes de mostrar el botón (solo informativo)
            letreros_cliente = []
            
            if cliente:
                letreros_existentes = leer_letreros()
                letreros_cliente = [l for l in letreros_existentes if normalizar_nombre(str(l.get('Cliente', ''))) == normalizar_nombre(cliente)]
                if letreros_cliente:
                    st.info(f"📋 **Información:** Este cliente tiene {len(letreros_cliente)} letrero(s) registrado(s) previamente. Puedes agregar otro letrero sin problema.")
                    
                    # Mostrar lista de letreros existentes en un expander compacto
                    with st.expander(f"Ver letreros existentes ({len(letreros_cliente)})", expanded=False):
                        for let_existente in letreros_cliente[:10]:  # Mostrar máximo 10
                            tipo_solicitud_existente = let_existente.get('Tipo Solicitud', 'NUEVO')
                            st.write(f"  • **#{let_existente['ID']}** - {let_existente['Tipo']} ({tipo_solicitud_existente}) - Estado: {let_existente['Estado']}")
                        if len(letreros_cliente) > 10:
                            st.caption(f"  ... y {len(letreros_cliente) - 10} más")
            
            submit = st.form_submit_button("✅ Guardar Letrero", use_container_width=True)
            
            if submit:
                # Validar campos obligatorios
                validacion = True
                if not cliente:
                    st.error("❌ Selecciona un cliente")
                    validacion = False
                if not tipo_letrero or tipo_letrero.strip() == "":
                    st.error("❌ Selecciona un tipo de letrero o especifica otro tipo en el campo 'Otro'")
                    validacion = False
                if not medidas:
                    st.error("❌ Ingresa las medidas")
                    validacion = False
                if not tipo_solicitud or tipo_solicitud.strip() == "":
                    st.error("❌ Selecciona el tipo de solicitud (NUEVO o ACTUALIZACIÓN)")
                    validacion = False
                
                if validacion:
                    # Guardar el letrero
                    next_id_val = next_id("LETREROS")
                    
                    # Fechas de fabricación vacías (se puede actualizar después)
                    fecha_envio_str = ""
                    fecha_entrega_str = ""
                    numero_guia_str = ""  # Número de guía vacío al crear
                    
                    values = [
                        next_id_val, 
                        cliente, 
                        tipo_solicitud,  # Tipo de Solicitud (NUEVO/ACTUALIZACIÓN)
                        tipo_letrero, 
                        medidas,
                        costo,
                        comercial,
                        fecha.strftime("%Y-%m-%d"), 
                        estado,
                        fecha_envio_str,
                        fecha_entrega_str,
                        observaciones,
                        numero_guia_str
                    ]
                    if append_row("LETREROS", values):
                        st.success("✅ **Letrero guardado exitosamente con estado 'EN PROCESO'**")
                        # Actualizar la caché de letreros para reflejar el nuevo registro
                        leer_letreros.clear()
                        # Recargar la página para actualizar la vista con el nuevo letrero
                        st.rerun()
                    else:
                        st.error("❌ No se pudo guardar el letrero. Revisa los mensajes de error arriba.")
        
        st.markdown("---")
        st.markdown("### 📊 Letreros en Proceso")
        
        if letreros_proceso:
            st.info(f"📋 {len(letreros_proceso)} letrero(s) en proceso")
            for letrero in letreros_proceso:
                # Buscar monto de ventas del cliente
                monto_ventas_cliente = 0
                try:
                    if EXCEL_PATH.exists():
                        wb_ventas = safe_load_workbook(EXCEL_PATH, tries=3, wait=0.1)
                        if "VENTAS_VINCULADAS" in wb_ventas.sheetnames:
                            ws_ventas = wb_ventas["VENTAS_VINCULADAS"]
                            cliente_norm = normalizar_nombre(letrero.get('Cliente', ''))
                            
                            for row in ws_ventas.iter_rows(min_row=2, values_only=True):
                                if row and row[0]:
                                    cliente_venta_norm = normalizar_nombre(str(row[0]).strip())
                                    if cliente_norm == cliente_venta_norm:
                                        venta_valor = row[1] if len(row) > 1 else 0
                                        try:
                                            monto_ventas_cliente += float(venta_valor) if venta_valor else 0
                                        except:
                                            pass
                        wb_ventas.close()
                except:
                    pass
                
                with st.expander(f"🔨 Letrero #{letrero['ID']} - {letrero['Cliente']}", expanded=False):
                    col1, col2 = st.columns(2)
                    col1.write(f"**Tipo Solicitud:** {letrero.get('Tipo Solicitud', 'NUEVO')}")
                    col1.write(f"**Tipo de Letrero:** {letrero['Tipo']}")
                    col1.write(f"**Medidas:** {letrero['Medidas']}")
                    col1.write(f"**Costo:** ${letrero['Costo']}")
                    col2.write(f"**Comercial:** {letrero['Comercial']}")
                    col2.write(f"**Fecha:** {letrero['Fecha']}")
                    col2.write(f"**Observaciones:** {letrero['Observaciones']}")
                    
                    # Mostrar monto de ventas del cliente si existe
                    if monto_ventas_cliente > 0:
                        st.markdown("---")
                        col_venta = st.columns(1)[0]
                        col_venta.markdown(f"**💰 Venta Total del Cliente:** ${monto_ventas_cliente:,.2f}")
                    
                    # Mostrar fechas de fabricación si existen
                    if letrero.get('Fecha_Envio_Fab') or letrero.get('Fecha_Entrega_Cliente') or letrero.get('Numero_Guia'):
                        st.markdown("**📅 Fechas de Fabricación:**")
                        col3, col4 = st.columns(2)
                        if letrero.get('Fecha_Envio_Fab'):
                            col3.write(f"🚚 Envío a fabricación: {letrero.get('Fecha_Envio_Fab')}")
                        if letrero.get('Fecha_Entrega_Cliente'):
                            col4.write(f"📦 Entrega a cliente: {letrero.get('Fecha_Entrega_Cliente')}")
                        if letrero.get('Numero_Guia'):
                            col3.write(f"📋 Número de Guía: {letrero.get('Numero_Guia')}")
                    
                    # Botones de acción
                    col_btn1, col_btn2 = st.columns(2)
                    
                    with col_btn1:
                        if st.button(f"✅ DISEÑO REALIZADO", key=f"cambiar_proceso_{letrero['ID']}", use_container_width=True):
                            with st.spinner("Actualizando estado..."):
                                if actualizar_estado_letrero(letrero['ID'], "PENDIENTE"):
                                    st.success(f"✅ Letrero #{letrero['ID']} movido a 'PENDIENTE'")
                                    time.sleep(1.5)  # Delay más largo para asegurar que se guardó completamente
                                    st.rerun()
                                else:
                                    st.error("❌ Error al actualizar el estado. Por favor verifica los logs.")
                    
                    with col_btn2:
                        # Verificar si ya se confirmó la eliminación
                        confirm_key = f"confirm_eliminar_{letrero['ID']}"
                        if confirm_key not in st.session_state:
                            st.session_state[confirm_key] = False
                        
                        if not st.session_state[confirm_key]:
                            if st.button(f"🗑️ Eliminar", key=f"eliminar_{letrero['ID']}", use_container_width=True, type="secondary"):
                                st.session_state[confirm_key] = True
                                st.warning(f"⚠️ ¿Estás seguro de eliminar el letrero #{letrero['ID']}? Esta acción no se puede deshacer.")
                                st.rerun()
                        else:
                            col_conf1, col_conf2 = st.columns(2)
                            with col_conf1:
                                if st.button(f"✅ Confirmar", key=f"confirmar_eliminar_{letrero['ID']}", use_container_width=True):
                                    with st.spinner("Eliminando letrero..."):
                                        if eliminar_letrero(letrero['ID']):
                                            st.success(f"✅ Letrero #{letrero['ID']} eliminado correctamente")
                                            del st.session_state[confirm_key]
                                            time.sleep(1.0)
                                            st.rerun()
                                        else:
                                            st.error("❌ Error al eliminar el letrero. Por favor verifica los logs.")
                                            del st.session_state[confirm_key]
                            with col_conf2:
                                if st.button(f"❌ Cancelar", key=f"cancelar_eliminar_{letrero['ID']}", use_container_width=True):
                                    del st.session_state[confirm_key]
                                    st.rerun()
        else:
            st.info("No hay letreros en proceso")
    
    with tab2:
        if letreros_pendientes:
            st.info(f"📋 {len(letreros_pendientes)} letrero(s) pendientes")
            for letrero in letreros_pendientes:
                with st.expander(f"⏸️ Letrero #{letrero['ID']} - {letrero['Cliente']}", expanded=False):
                    col1, col2 = st.columns(2)
                    col1.write(f"**Tipo Solicitud:** {letrero.get('Tipo Solicitud', 'NUEVO')}")
                    col1.write(f"**Tipo de Letrero:** {letrero['Tipo']}")
                    col1.write(f"**Medidas:** {letrero['Medidas']}")
                    col1.write(f"**Costo:** ${letrero['Costo']}")
                    col2.write(f"**Comercial:** {letrero['Comercial']}")
                    col2.write(f"**Fecha:** {letrero['Fecha']}")
                    col2.write(f"**Observaciones:** {letrero['Observaciones']}")
                    
                    # Mostrar monto total de compras del cliente cuando está PENDIENTE
                    monto_ventas_cliente = 0
                    try:
                        if EXCEL_PATH.exists():
                            wb_ventas = safe_load_workbook(EXCEL_PATH, tries=3, wait=0.1)
                            cliente_norm = normalizar_nombre(letrero.get('Cliente', ''))
                            if "VENTAS_VINCULADAS" in wb_ventas.sheetnames:
                                ws_v = wb_ventas["VENTAS_VINCULADAS"]
                                for row in ws_v.iter_rows(min_row=2, values_only=True):
                                    if row and row[0]:
                                        if cliente_norm == normalizar_nombre(str(row[0]).strip()):
                                            try:
                                                monto_ventas_cliente += float(row[1]) if row[1] else 0
                                            except:
                                                pass
                            elif "VENTAS" in wb_ventas.sheetnames:
                                ws_v = wb_ventas["VENTAS"]
                                for row in ws_v.iter_rows(min_row=2, values_only=True):
                                    if row and row[0]:
                                        if cliente_norm == normalizar_nombre(str(row[0]).strip()):
                                            try:
                                                monto_ventas_cliente += float(row[1]) if len(row) > 1 and row[1] else 0
                                            except:
                                                pass
                            wb_ventas.close()
                    except:
                        pass

                    st.markdown("---")
                    st.markdown(f"**💰 Venta Total del Cliente:** ${monto_ventas_cliente:,.2f}")
                    
                    # Mostrar fechas de fabricación si existen (solo lectura)
                    if letrero.get('Fecha_Envio_Fab') or letrero.get('Fecha_Entrega_Cliente') or letrero.get('Numero_Guia'):
                        st.markdown("**📅 Fechas de Fabricación:**")
                        col3, col4 = st.columns(2)
                        if letrero.get('Fecha_Envio_Fab'):
                            col3.write(f"🚚 Envío a fabricación: {letrero.get('Fecha_Envio_Fab')}")
                        if letrero.get('Fecha_Entrega_Cliente'):
                            col4.write(f"📦 Entrega a cliente: {letrero.get('Fecha_Entrega_Cliente')}")
                        if letrero.get('Numero_Guia'):
                            col3.write(f"📋 Número de Guía: {letrero.get('Numero_Guia')}")
                    
                    st.divider()
                    
                    # Botón para generar PDF de autorización
                    col_gen, col_down = st.columns(2)
                    if col_gen.button(f"📄 Generar PDF de Autorización", key=f"gen_pdf_{letrero['ID']}"):
                        archivo_pdf = generar_pdf_autorizacion(letrero)
                        if archivo_pdf:
                            st.success(f"✅ PDF generado: {archivo_pdf.name}")
                            try:
                                with open(str(archivo_pdf), 'rb') as f:
                                    pdf_data = f.read()
                                col_down.download_button(
                                    label="⬇️ Descargar PDF",
                                    data=pdf_data,
                                    file_name=archivo_pdf.name,
                                    mime="application/pdf",
                                    key=f"download_{letrero['ID']}"
                                )
                            except Exception as e:
                                st.error(f"Error al leer PDF: {e}")
                        else:
                            st.error("❌ Error al generar el PDF")
                    
                    st.divider()
                    
                    # Sección para subir PDF o Foto firmada
                    st.markdown("**📤 Subir PDF o Foto Firmada por Gerencia**")
                    archivo_firmado = st.file_uploader(
                        "Selecciona el PDF o foto firmada escaneada",
                        type=['pdf', 'jpg', 'jpeg', 'png'],
                        key=f"upload_pdf_{letrero['ID']}"
                    )
                    
                    if archivo_firmado:
                        # Guardar archivo en carpeta del cliente
                        carpeta_cliente = Path(EXCEL_DIR) / "LETREROS_AUTORIZACIONES" / letrero['Cliente'].replace("/", "_").replace("\\", "_")
                        carpeta_cliente.mkdir(parents=True, exist_ok=True)
                        
                        # Determinar extensión según tipo de archivo
                        tipo_archivo = archivo_firmado.type
                        if 'pdf' in tipo_archivo:
                            extension = '.pdf'
                        elif 'jpeg' in tipo_archivo or 'jpg' in tipo_archivo:
                            extension = '.jpg'
                        elif 'png' in tipo_archivo:
                            extension = '.png'
                        else:
                            extension = '.pdf'  # Default
                        
                        nombre_archivo = carpeta_cliente / f"Autorizacion_Firmada_{letrero['ID']}{extension}"
                        
                        with open(nombre_archivo, "wb") as f:
                            f.write(archivo_firmado.getbuffer())
                        st.success(f"✅ Archivo guardado en: {nombre_archivo.name}")
                    
                    st.divider()
                    
                    # Botón para marcar como ENTREGADO
                    if st.button(f"✅ Marcar como ENTREGADO", key=f"cambiar_entregado_{letrero['ID']}"):
                        if actualizar_estado_letrero(letrero['ID'], "ENTREGADO"):
                            st.success("✅ Estado actualizado a 'ENTREGADO'")
                            st.rerun()
                        else:
                            st.error("❌ Error al actualizar el estado")
        else:
            st.info("No hay letreros pendientes")
    
    with tab3:
        # Tab de Fecha de Fabricación
        st.markdown("### 📅 Gestión de Fechas de Fabricación")
        
        # Crear lista de letreros para selección
        letreros_lista = []
        for l in letreros:
            estado_display = ""
            if l["Estado"] == "EN PROCESO":
                estado_display = " (⏳ En Proceso)"
            elif l["Estado"] == "PENDIENTE":
                estado_display = " (⏸️ Pendiente)"
            elif l["Estado"] == "ENTREGADO":
                estado_display = " (✅ Entregado)"
            
            letreros_lista.append(f"#{l['ID']} - {l['Cliente']}{estado_display}")
        
        if letreros_lista:
            letrero_seleccionado = st.selectbox("Selecciona un letrero", [""] + letreros_lista, key="select_letrero_fecha")
            
            col_fecha_envio, col_fecha_entrega = st.columns(2)
            
            # Extraer ID del letrero seleccionado
            if letrero_seleccionado and letrero_seleccionado != "":
                letrero_id_seleccionado = letrero_seleccionado.split(" - ")[0].replace("#", "")
                
                # Buscar el letrero para mostrar fechas actuales
                letrero_actual = next((l for l in letreros if l['ID'] == letrero_id_seleccionado), None)
                
                # Mostrar información del letrero
                if letrero_actual:
                    col_info1, col_info2 = st.columns(2)
                    col_info1.markdown(f"**Cliente:** {letrero_actual.get('Cliente', 'N/A')}")
                    col_info1.markdown(f"**Tipo:** {letrero_actual.get('Tipo', 'N/A')}")
                    col_info2.markdown(f"**Medidas:** {letrero_actual.get('Medidas', 'N/A')}")
                    col_info2.markdown(f"**Costo:** ${letrero_actual.get('Costo', '0')}")
                    
                    st.divider()
                
                # Convertir fechas si existen
                try:
                    if letrero_actual and letrero_actual.get('Fecha_Envio_Fab'):
                        fecha_envio_actual = datetime.strptime(str(letrero_actual['Fecha_Envio_Fab']), '%Y-%m-%d').date() if isinstance(letrero_actual['Fecha_Envio_Fab'], str) else letrero_actual['Fecha_Envio_Fab']
                    else:
                        fecha_envio_actual = None
                except:
                    fecha_envio_actual = None
                
                try:
                    if letrero_actual and letrero_actual.get('Fecha_Entrega_Cliente'):
                        fecha_entrega_actual = datetime.strptime(str(letrero_actual['Fecha_Entrega_Cliente']), '%Y-%m-%d').date() if isinstance(letrero_actual['Fecha_Entrega_Cliente'], str) else letrero_actual['Fecha_Entrega_Cliente']
                    else:
                        fecha_entrega_actual = None
                except:
                    fecha_entrega_actual = None
                
                fecha_envio_fab = col_fecha_envio.date_input("Fecha de envío a fabricación", key="gestion_fecha_envio", value=fecha_envio_actual)
                fecha_entrega_cli = col_fecha_entrega.date_input("Fecha de entrega a cliente", key="gestion_fecha_entrega", value=fecha_entrega_actual)
                
                # Campo para número de guía
                st.divider()
                numero_guia_valor = ""
                if letrero_actual and letrero_actual.get('Numero_Guia'):
                    numero_guia_valor = str(letrero_actual.get('Numero_Guia', ''))
                numero_guia = st.text_input("Número de Guía", key="gestion_numero_guia", value=numero_guia_valor)
                
                # Botón para guardar fechas
                col_btn1, col_btn2 = st.columns(2)
                if col_btn1.button("💾 Guardar Fechas", key="guardar_fechas_global", use_container_width=True):
                    if actualizar_fechas_fabricacion(letrero_id_seleccionado, fecha_envio_fab, fecha_entrega_cli, numero_guia if numero_guia.strip() else None):
                        st.success("✅ Fechas y número de guía guardados correctamente")
                        st.rerun()
                    else:
                        st.error("❌ Error al guardar las fechas")
            else:
                fecha_envio_fab = col_fecha_envio.date_input("Fecha de envío a fabricación", key="gestion_fecha_envio", value=None)
                fecha_entrega_cli = col_fecha_entrega.date_input("Fecha de entrega a cliente", key="gestion_fecha_entrega", value=None)
                st.divider()
                numero_guia = st.text_input("Número de Guía", key="gestion_numero_guia", value="")
    
    with tab4:
        if letreros_entregados:
            st.info(f"✅ {len(letreros_entregados)} letrero(s) finalizado(s)")
            for letrero in letreros_entregados:
                with st.expander(f"✅ Letrero #{letrero['ID']} - {letrero['Cliente']}", expanded=False):
                    col1, col2 = st.columns(2)
                    col1.write(f"**Tipo Solicitud:** {letrero.get('Tipo Solicitud', 'NUEVO')}")
                    col1.write(f"**Tipo de Letrero:** {letrero['Tipo']}")
                    col1.write(f"**Medidas:** {letrero['Medidas']}")
                    col1.write(f"**Costo:** ${letrero['Costo']}")
                    col2.write(f"**Comercial:** {letrero['Comercial']}")
                    col2.write(f"**Fecha:** {letrero['Fecha']}")
                    col2.write(f"**Observaciones:** {letrero['Observaciones']}")
                    
                    # Mostrar fechas de fabricación si existen
                    if letrero.get('Fecha_Envio_Fab') or letrero.get('Fecha_Entrega_Cliente') or letrero.get('Numero_Guia'):
                        st.markdown("**📅 Fechas de Fabricación:**")
                        col3, col4 = st.columns(2)
                        if letrero.get('Fecha_Envio_Fab'):
                            col3.write(f"🚚 Envío a fabricación: {letrero.get('Fecha_Envio_Fab')}")
                        if letrero.get('Fecha_Entrega_Cliente'):
                            col4.write(f"📦 Entrega a cliente: {letrero.get('Fecha_Entrega_Cliente')}")
                        if letrero.get('Numero_Guia'):
                            col3.write(f"📋 Número de Guía: {letrero.get('Numero_Guia')}")
                    
                    st.markdown("---")
                    st.markdown("**📷 Foto del Letrero Colocado:**")
                    
                    # Usar la misma carpeta donde están las autorizaciones
                    cliente_nombre_limpio = str(letrero['Cliente']).replace("/", "_").replace("\\", "_")
                    carpeta_cliente = Path(EXCEL_DIR) / "LETREROS_AUTORIZACIONES" / cliente_nombre_limpio
                    carpeta_cliente.mkdir(parents=True, exist_ok=True)
                    
                    # Buscar foto existente con diferentes variaciones
                    foto_existente = None
                    letrero_id = str(letrero['ID'])
                    
                    # Buscar con diferentes patrones de nombre
                    patrones = [
                        f"Foto_Letrero_{letrero_id}",
                        f"Letrero_{letrero_id}",
                        f"foto_{letrero_id}",
                    ]
                    
                    extensiones = ['.jpg', '.jpeg', '.png', '.JPG', '.JPEG', '.PNG']
                    
                    for patron in patrones:
                        for ext in extensiones:
                            foto_path = carpeta_cliente / f"{patron}{ext}"
                            if foto_path.exists():
                                foto_existente = foto_path
                                break
                        if foto_existente:
                            break
                    
                    # Mostrar foto si existe
                    if foto_existente:
                        st.success(f"✅ 1 foto cargada")
                        st.image(str(foto_existente), width=400, caption=f"Foto del Letrero #{letrero['ID']}")
                    else:
                        st.info("ℹ️ No hay foto cargada aún")
                    
                    # Campo para subir nueva foto
                    foto_subida = st.file_uploader(
                        "Subir foto del letrero colocado",
                        type=['jpg', 'jpeg', 'png'],
                        key=f"upload_foto_letrero_{letrero['ID']}",
                        help="Sube una foto del letrero ya instalado/colocado"
                    )
                    
                    # Verificar si ya hay una foto en el sistema antes de procesar
                    foto_upload_key = f"upload_foto_letrero_{letrero['ID']}"
                    foto_procesada_key = f"foto_procesada_{letrero['ID']}"
                    
                    if foto_subida:
                        # Obtener información única del archivo
                        foto_info = f"{foto_subida.name}_{foto_subida.size}"
                        
                        # Verificar si esta foto ya fue procesada en esta sesión
                        if foto_procesada_key not in st.session_state or st.session_state[foto_procesada_key] != foto_info:
                            try:
                                # Determinar extensión
                                tipo_archivo = foto_subida.type
                                if 'jpeg' in tipo_archivo or 'jpg' in tipo_archivo:
                                    extension = '.jpg'
                                elif 'png' in tipo_archivo:
                                    extension = '.png'
                                else:
                                    extension = '.jpg'
                                
                                # Intentar guardar en Google Drive primero
                                foto_url = None
                                try:
                                    from gdrive_helper import save_photo_letrero_drive
                                    drive_result = save_photo_letrero_drive(letrero['ID'], foto_subida)
                                    if drive_result and drive_result.get('url'):
                                        foto_url = drive_result['url']
                                except Exception as e:
                                    print(f"Google Drive no disponible para foto de letrero: {e}")
                                
                                # Si Google Drive no está disponible, guardar localmente
                                if not foto_url:
                                    nombre_foto = carpeta_cliente / f"Foto_Letrero_{letrero['ID']}{extension}"
                                    with open(nombre_foto, "wb") as f:
                                        f.write(foto_subida.getbuffer())
                                    foto_url = str(nombre_foto) if nombre_foto.exists() else None
                                
                                # Verificar que el archivo se guardó correctamente
                                if foto_url:
                                    # Marcar esta foto como procesada
                                    st.session_state[foto_procesada_key] = foto_info
                                    
                                    # Limpiar el cache
                                    leer_letreros.clear()
                                    
                                    st.success("✅ Foto subida correctamente")
                                    
                                    # Agregar botón para refrescar y ver la foto
                                    if st.button("🔄 Ver foto actualizada", key=f"refresh_foto_{letrero['ID']}"):
                                        st.rerun()
                                else:
                                    st.error("❌ Error: La foto no se guardó correctamente")
                            except Exception as e:
                                st.error(f"❌ Error al guardar la foto: {e}")
        else:
            st.info("No hay letreros finalizados")
    
    with tab5:
        st.markdown("### 📈 Reportes de Letreros")
        if letreros:
            # Convertir a DataFrame para facilitar filtros
            df_letreros = pd.DataFrame(letreros)
            # Normalizar tipos
            try:
                df_letreros['Fecha'] = pd.to_datetime(df_letreros['Fecha'], errors='coerce')
            except Exception:
                pass
            try:
                df_letreros['Costo'] = pd.to_numeric(df_letreros['Costo'], errors='coerce').fillna(0.0)
            except Exception:
                df_letreros['Costo'] = 0.0

            col_f1, col_f2, col_f3, col_f4 = st.columns([2,2,2,1])
            cliente_f = col_f1.selectbox("Filtrar por Cliente", [""] + sorted(df_letreros['Cliente'].dropna().unique().tolist()), index=0)
            comercial_f = col_f2.selectbox("Filtrar por Comercial", [""] + sorted(df_letreros['Comercial'].dropna().unique().tolist()), index=0)
            # Rango de fechas
            min_fecha = df_letreros['Fecha'].min()
            max_fecha = df_letreros['Fecha'].max()
            rango_f = col_f3.date_input("Rango de Fechas", value=(min_fecha.date() if pd.notna(min_fecha) else datetime.now().date(),
                                                                max_fecha.date() if pd.notna(max_fecha) else datetime.now().date()))
            estado_f = col_f4.selectbox("Estado", ["", "EN PROCESO", "PENDIENTE", "ENTREGADO"], index=0)

            df_rep = df_letreros.copy()
            if cliente_f:
                df_rep = df_rep[df_rep['Cliente'] == cliente_f]
            if comercial_f:
                df_rep = df_rep[df_rep['Comercial'] == comercial_f]
            if isinstance(rango_f, tuple) and len(rango_f) == 2:
                inicio, fin = pd.to_datetime(rango_f[0]), pd.to_datetime(rango_f[1])
                df_rep = df_rep[(df_rep['Fecha'] >= inicio) & (df_rep['Fecha'] <= fin)]
            if estado_f:
                df_rep = df_rep[df_rep['Estado'] == estado_f]

            # Métricas
            col_m1, col_m2, col_m3 = st.columns(3)
            col_m1.metric("Total de Letreros", len(df_rep))
            col_m2.metric("Costo Total", f"${df_rep['Costo'].sum():,.2f}")
            col_m3.metric("Clientes Únicos", df_rep['Cliente'].nunique())

            # Tabla
            st.dataframe(df_rep.sort_values(by=['Fecha','ID'], ascending=[False, True]), use_container_width=True, hide_index=True)

            # Descarga en formato Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_rep.to_excel(writer, index=False, sheet_name='Reporte Letreros')
            excel_bytes = output.getvalue()
            st.download_button("⬇️ Descargar Excel", excel_bytes, file_name="reporte_letreros.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("No hay letreros registrados para generar reportes.")

# ===== SECCIÓN ACTIVACIONES Y EVENTOS =====
elif selected_menu == "🎉 ACTIVACIONES Y EVENTOS":
    # Leer eventos
    eventos = leer_eventos()
    
    # Función para cargar productos del inventario
    def load_inv_products():
        productos = []
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if "INVENTARIO_PRODUCTOS" in wb.sheetnames:
                ws = wb["INVENTARIO_PRODUCTOS"]
                headers = [cell.value for cell in ws[1]]
                has_precio_venta = "PrecioVenta" in headers
                is_old_format = "StockInicial" in headers
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        prod = {
                            "Codigo": str(row[0]).strip(),
                            "Nombre": str(row[1]).strip() if row[1] else "",
                            "Categoria": str(row[2]).strip() if row[2] else "",
                            "Unidad": str(row[3]).strip() if len(row) > 3 and row[3] else "pc",
                        }
                        if is_old_format and len(row) >= 9:
                            # Estructura antigua: PrecioVenta en posición 3
                            prod["PrecioVenta"] = float(row[3]) if row[3] else 0.0
                            prod["Unidad"] = str(row[4]).strip() if len(row) > 4 and row[4] else "pc"
                        elif has_precio_venta:
                            # Estructura nueva: PrecioVenta en posición 5
                            prod["PrecioVenta"] = float(row[5]) if len(row) > 5 and row[5] else 0.0
                        else:
                            # Sin PrecioVenta
                            prod["PrecioVenta"] = 0.0
                        productos.append(prod)
            wb.close()
        except Exception:
            pass
        return productos
    
    # Filtrar eventos por estado
    eventos_programados = [e for e in eventos if str(e.get("Estado", "")).strip().upper() == "PROGRAMADO"]
    eventos_aprobados = [e for e in eventos if str(e.get("Estado", "")).strip().upper() == "APROBADO"]
    eventos_realizados = [e for e in eventos if str(e.get("Estado", "")).strip().upper() == "REALIZADO"]
    
    # Tabs para organizar
    tab1, tab2, tab3, tab4 = st.tabs(["⏳ En Proceso", "✅ Aprobados", "✅ Finalizados", "📊 Reportes"])
    
    with tab1:
        st.markdown("#### ⏳ Activaciones en Proceso")
        
        if eventos_programados:
            st.info(f"📋 {len(eventos_programados)} activación(es) programada(s)")
            for evento in eventos_programados:
                with st.expander(f"📅 {evento['Tipo']} - {evento['Cliente']} (#{evento['ID']})", expanded=False):
                    col1, col2 = st.columns(2)
                    col1.write(f"**Cliente:** {evento.get('Cliente', 'N/A')}")
                    col1.write(f"**Tipo:** {evento.get('Tipo', 'N/A')}")
                    col1.write(f"**Fecha:** {evento.get('Fecha', 'N/A')}")
                    col2.write(f"**Estado:** {evento.get('Estado', 'N/A')}")
                    if evento.get('Descripcion'):
                        st.write(f"**Descripción:** {evento.get('Descripcion')}")
                    
                    # Botón para enviar a aprobado
                    if st.button(f"✅ Enviar a Aprobado", key=f"aprobado_{evento['ID']}", use_container_width=True, type="primary"):
                        with st.spinner("Aprobando evento y descontando stock..."):
                            # Primero descontar stock
                            stock_descontado = descontar_stock_evento(evento)
                            
                            # Luego actualizar el estado
                            if actualizar_estado_evento(evento['ID'], "APROBADO"):
                                if stock_descontado:
                                    st.success(f"✅ Activación #{evento['ID']} aprobada y stock descontado exitosamente")
                                else:
                                    st.success(f"✅ Activación #{evento['ID']} enviada a 'APROBADO'")
                                time.sleep(1.5)
                                st.rerun()
                            else:
                                st.error("❌ Error al actualizar el estado. Por favor verifica los logs.")
        else:
            st.info("No hay activaciones programadas (en proceso)")
        
        st.markdown("---")
        
        # Cargar productos del inventario
        inv_productos = load_inv_products()

        # Definir proformas por tipo de evento (códigos de productos por defecto)
        def get_default_products_for_type(tipo):
            """Devuelve lista de códigos de productos por defecto según el tipo de evento"""
            # Cargar proformas desde un archivo de configuración o usar defaults hardcodeados
            proformas = {
                "ACTIVACION EXTREME": [
                    "CARPA_EX", "INFLABLE_ARCO_EX", "MESA_MANTEL", "PARLANTE_MICROFONO",
                    "BANDERAS_EX", "INFLABLE_BATERIAS_EX", "QR_EX", "VESTIDO_EX"
                ],
                "ACTIVACION PANTRO": ["PROD1_PT", "PROD2_PT", "PROD3_PT"],
                "ACTIVACION EXTREME Y PANTRO": [
                    "CARPA_EX", "INFLABLE_ARCO_EX", "PROD1_PT", "PROD2_PT"
                ],
                "CARRERA PANTRO": ["CARRERA1_PT", "CARRERA2_PT"],
                "CARRERA EXTREME": ["CARRERA1_EX", "CARRERA2_EX"],
                "PRESTAR LAS COSAS": ["PRESTADO1", "PRESTADO2", "PRESTADO3"]
            }
            return proformas.get(tipo, [])
        
        def load_products_for_event_type(tipo, inv_productos):
            """Carga productos guardados para un tipo de evento, o productos por defecto si no hay guardados"""
            # Primero intentar cargar productos guardados
            productos_guardados = load_saved_products_for_type(tipo)
            if productos_guardados:
                # Verificar que los productos guardados aún existen en el inventario
                productos_validos = []
                for prod_guardado in productos_guardados:
                    codigo = prod_guardado.get('codigo', '')
                    producto_info = next((p for p in inv_productos if p['Codigo'] == codigo), None)
                    if producto_info:
                        # Actualizar con información actualizada del inventario pero mantener cantidad, precio y descuento
                        productos_validos.append({
                            'codigo': producto_info['Codigo'],
                            'nombre': producto_info['Nombre'],
                            'categoria': producto_info.get('Categoria', prod_guardado.get('categoria', '')),
                            'unidad': producto_info.get('Unidad', prod_guardado.get('unidad', 'pc')),
                            'cantidad': float(prod_guardado.get('cantidad', 1.0)),
                            'precio_unit': float(prod_guardado.get('precio_unit', producto_info.get('PrecioVenta', 0.0))),
                            'descuento': float(prod_guardado.get('descuento', 0.0)),
                            'nota': str(prod_guardado.get('nota', ''))
                        })
                if productos_validos:
                    return productos_validos
            
            # Si no hay productos guardados, usar productos por defecto
            default_codes = get_default_products_for_type(tipo)
            productos_defecto = []
            for codigo in default_codes:
                producto_info = next((p for p in inv_productos if p['Codigo'] == codigo), None)
                if producto_info:
                    productos_defecto.append({
                            'codigo': producto_info['Codigo'],
                            'nombre': producto_info['Nombre'],
                            'categoria': producto_info.get('Categoria', ''),
                            'unidad': producto_info.get('Unidad', 'pc'),
                            'cantidad': 1.0,
                            'precio_unit': float(producto_info.get('PrecioVenta', 0.0)),
                            'descuento': 0.0,
                            'nota': ''
                        })
            return productos_defecto

        # Pesos por defecto para ACTIVACION EXTREME (valores totales por línea)
        pesos_default_extreme = {
            "Mesa y mantel": 12.95,
            "Parlante y micrófono": 18.75,
            "Banderas Extrememax": 5.10,  # total de la línea; se divide por la cantidad
            "Inflable tipo baterías voltmax": 8.05,
            "QR Extrememax": 0.0,
            "Vestido Extrememax": 0.0,
        }
        
        # Mostrar tipos de eventos como botones en forma horizontal
        st.markdown("### Tipos de Eventos Disponibles:")
        col_tipo1, col_tipo2, col_tipo3, col_tipo4, col_tipo5, col_tipo6 = st.columns(6)
        
        # CSS personalizado para botones sutiles con texto blanco
        st.markdown("""
        <style>
        /* Estilos específicos para los botones de tipos de eventos usando sus keys */
        button[key="btn_extreme"],
        button[key="btn_pantro"],
        button[key="btn_extreme_pantro"],
        button[key="btn_carrera_pantro"],
        button[key="btn_carrera_extreme"],
        button[key="btn_prestar"] {
            background-color: #6B7280 !important;
            background: #6B7280 !important;
            background-image: none !important;
            color: #FFFFFF !important;
            border: 1px solid #9CA3AF !important;
            border-radius: 3px !important;
            padding: 0.05rem 0.15rem !important;
            font-size: 0.4rem !important;
            font-weight: 500 !important;
            min-height: 16px !important;
            max-height: 20px !important;
            height: auto !important;
            line-height: 0.9 !important;
            box-shadow: 0 1px 2px rgba(0,0,0,0.1) !important;
            text-shadow: none !important;
            white-space: pre-line !important;
            transition: all 0.2s ease !important;
        }
        button[key="btn_extreme"]:hover,
        button[key="btn_pantro"]:hover,
        button[key="btn_extreme_pantro"]:hover,
        button[key="btn_carrera_pantro"]:hover,
        button[key="btn_carrera_extreme"]:hover,
        button[key="btn_prestar"]:hover {
            background-color: #4B5563 !important;
            background: #4B5563 !important;
            border-color: #6B7280 !important;
            color: #FFFFFF !important;
            box-shadow: 0 2px 4px rgba(0,0,0,0.15) !important;
        }
        /* Forzar color de texto blanco en todos los elementos internos */
        button[key="btn_extreme"] *,
        button[key="btn_pantro"] *,
        button[key="btn_extreme_pantro"] *,
        button[key="btn_carrera_pantro"] *,
        button[key="btn_carrera_extreme"] *,
        button[key="btn_prestar"] *,
        button[key="btn_extreme"] span,
        button[key="btn_pantro"] span,
        button[key="btn_extreme_pantro"] span,
        button[key="btn_carrera_pantro"] span,
        button[key="btn_carrera_extreme"] span,
        button[key="btn_prestar"] span,
        button[key="btn_extreme"] p,
        button[key="btn_pantro"] p,
        button[key="btn_extreme_pantro"] p,
        button[key="btn_carrera_pantro"] p,
        button[key="btn_carrera_extreme"] p,
        button[key="btn_prestar"] p {
            color: #FFFFFF !important;
            fill: #FFFFFF !important;
        }
        </style>
        <script>
        function aplicarEstilosBotones() {
            const keys = ['btn_extreme', 'btn_pantro', 'btn_extreme_pantro', 'btn_carrera_pantro', 'btn_carrera_extreme', 'btn_prestar'];
            keys.forEach(function(key) {
                const buttons = document.querySelectorAll('button[key="' + key + '"]');
                buttons.forEach(function(btn) {
                    if(btn) {
                        btn.style.setProperty('background-color', '#6B7280', 'important');
                        btn.style.setProperty('background', '#6B7280', 'important');
                        btn.style.setProperty('color', '#FFFFFF', 'important');
                        btn.style.setProperty('border', '1px solid #9CA3AF', 'important');
                        btn.style.setProperty('border-radius', '3px', 'important');
                        btn.style.setProperty('padding', '0.4rem 0.6rem', 'important');
                        btn.style.setProperty('font-size', '0.65rem', 'important');
                        btn.style.setProperty('font-weight', '600', 'important');
                        btn.style.setProperty('min-height', '38px', 'important');
                        btn.style.setProperty('max-height', '45px', 'important');
                        btn.style.setProperty('background', 'linear-gradient(180deg, #4A5568 0%, #2D3748 100%)', 'important');
                        btn.style.setProperty('box-shadow', '0 2px 4px rgba(0,0,0,0.15)', 'important');
                        btn.style.setProperty('border-radius', '6px', 'important');
                        btn.style.setProperty('text-transform', 'uppercase', 'important');
                        btn.style.setProperty('letter-spacing', '0.5px', 'important');
                        btn.style.setProperty('box-shadow', '0 1px 2px rgba(0,0,0,0.1)', 'important');
                        btn.style.setProperty('transition', 'all 0.2s ease', 'important');
                        btn.style.setProperty('line-height', '1.0', 'important');
                        btn.style.setProperty('white-space', 'pre-line', 'important');
                        // Forzar color blanco en todos los elementos hijos
                        var elementos = btn.querySelectorAll('*');
                        for(var i = 0; i < elementos.length; i++) {
                            elementos[i].style.setProperty('color', '#FFFFFF', 'important');
                            elementos[i].style.setProperty('fill', '#FFFFFF', 'important');
                        }
                        // También aplicar directamente al texto del botón
                        if(btn.textContent) {
                            btn.style.setProperty('color', '#FFFFFF', 'important');
                        }
                    }
                });
            });
        }
        // Ejecutar múltiples veces para asegurar que se aplique
        aplicarEstilosBotones();
        setTimeout(aplicarEstilosBotones, 50);
        setTimeout(aplicarEstilosBotones, 100);
        setTimeout(aplicarEstilosBotones, 200);
        setTimeout(aplicarEstilosBotones, 400);
        setTimeout(aplicarEstilosBotones, 800);
        setTimeout(aplicarEstilosBotones, 1500);
        // Observar cambios en el DOM de forma más agresiva
        if (window.MutationObserver) {
            var observer = new MutationObserver(function(mutations) {
                aplicarEstilosBotones();
            });
            observer.observe(document.body, { 
                childList: true, 
                subtree: true,
                attributes: true,
                attributeFilter: ['style', 'class']
            });
        }
        // También ejecutar cuando se carga completamente y periódicamente
        window.addEventListener('load', aplicarEstilosBotones);
        setInterval(aplicarEstilosBotones, 1000);
        </script>
        """, unsafe_allow_html=True)
        
        tipo_seleccionado_boton = ""
        
        # Crear botones HTML personalizados más pequeños
        btn_html = """
        <style>
        /* Estilos mejorados para botones más atractivos */
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"] button,
        button[key="btn_extreme"],
        button[key="btn_pantro"],
        button[key="btn_extreme_pantro"],
        button[key="btn_carrera_pantro"],
        button[key="btn_carrera_extreme"],
        button[key="btn_prestar"],
        button[data-baseweb="button"] {
            background-color: #4A5568 !important;
            background: linear-gradient(180deg, #4A5568 0%, #2D3748 100%) !important;
            color: #FFFFFF !important;
            border: 1px solid #718096 !important;
            border-radius: 6px !important;
            padding: 0.4rem 0.6rem !important;
            font-size: 0.65rem !important;
            font-weight: 600 !important;
            min-height: 38px !important;
            max-height: 45px !important;
            height: auto !important;
            line-height: 1.3 !important;
            box-shadow: 0 2px 4px rgba(0,0,0,0.15) !important;
            cursor: pointer !important;
            white-space: pre-line !important;
            width: 100% !important;
            transition: all 0.2s ease !important;
            text-transform: uppercase !important;
            letter-spacing: 0.5px !important;
        }
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"] button:hover,
        button[key="btn_extreme"]:hover,
        button[key="btn_pantro"]:hover,
        button[key="btn_extreme_pantro"]:hover,
        button[key="btn_carrera_pantro"]:hover,
        button[key="btn_carrera_extreme"]:hover,
        button[key="btn_prestar"]:hover {
            background: linear-gradient(180deg, #2D3748 0%, #1A202C 100%) !important;
            border-color: #4A5568 !important;
            box-shadow: 0 4px 8px rgba(0,0,0,0.25) !important;
            transform: translateY(-1px) !important;
        }
        /* Forzar color de texto blanco en todos los elementos */
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"] button span,
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"] button p,
        div[data-testid="stHorizontalBlock"] > div[data-testid="column"] button *,
        button[key="btn_extreme"] *,
        button[key="btn_pantro"] *,
        button[key="btn_extreme_pantro"] *,
        button[key="btn_carrera_pantro"] *,
        button[key="btn_carrera_extreme"] *,
        button[key="btn_prestar"] * {
            color: #FFFFFF !important;
            font-size: 0.65rem !important;
            font-weight: 600 !important;
        }
        </style>
        """
        st.markdown(btn_html, unsafe_allow_html=True)
        
        with col_tipo1:
            if st.button("ACTIVACION\nEXTREME", key="btn_extreme_tab2", use_container_width=True):
                st.session_state["act_tipo"] = "ACTIVACION EXTREME"
                # Cargar productos guardados o por defecto para este tipo
                st.session_state["act_productos_detalles"] = load_products_for_event_type("ACTIVACION EXTREME", inv_productos)
                st.session_state["act_productos_text"] = ""
                st.rerun()
        with col_tipo2:
            if st.button("ACTIVACION\nPANTRO", key="btn_pantro_tab2", use_container_width=True):
                st.session_state["act_tipo"] = "ACTIVACION PANTRO"
                # Cargar productos guardados o por defecto para este tipo
                st.session_state["act_productos_detalles"] = load_products_for_event_type("ACTIVACION PANTRO", inv_productos)
                st.session_state["act_productos_text"] = ""
                st.rerun()
        with col_tipo3:
            if st.button("ACTIVACION\nEXTREME Y\nPANTRO", key="btn_extreme_pantro_tab2", use_container_width=True):
                st.session_state["act_tipo"] = "ACTIVACION EXTREME Y PANTRO"
                # Cargar productos guardados o por defecto para este tipo
                st.session_state["act_productos_detalles"] = load_products_for_event_type("ACTIVACION EXTREME Y PANTRO", inv_productos)
                st.session_state["act_productos_text"] = ""
                st.rerun()
        with col_tipo4:
            if st.button("CARRERA\nPANTRO", key="btn_carrera_pantro_tab2", use_container_width=True):
                st.session_state["act_tipo"] = "CARRERA PANTRO"
                # Cargar productos guardados o por defecto para este tipo
                st.session_state["act_productos_detalles"] = load_products_for_event_type("CARRERA PANTRO", inv_productos)
                st.session_state["act_productos_text"] = ""
                st.rerun()
        with col_tipo5:
            if st.button("CARRERA\nEXTREME", key="btn_carrera_extreme_tab2", use_container_width=True):
                st.session_state["act_tipo"] = "CARRERA EXTREME"
                # Cargar productos guardados o por defecto para este tipo
                st.session_state["act_productos_detalles"] = load_products_for_event_type("CARRERA EXTREME", inv_productos)
                st.session_state["act_productos_text"] = ""
                st.rerun()
        with col_tipo6:
            if st.button("PRESTAR\nLAS COSAS", key="btn_prestar_tab2", use_container_width=True):
                st.session_state["act_tipo"] = "PRESTAR LAS COSAS"
                # Cargar productos guardados o por defecto para este tipo
                st.session_state["act_productos_detalles"] = load_products_for_event_type("PRESTAR LAS COSAS", inv_productos)
                st.session_state["act_productos_text"] = ""
                st.rerun()
        
        st.markdown("---")
        
        # Campo para productos - Tabla interactiva tipo proforma (ANTES del formulario)
        st.markdown("#### 📦 Productos")
        
        if inv_productos:
            # Inicializar lista de productos seleccionados con detalles
            if "act_productos_detalles" not in st.session_state:
                st.session_state["act_productos_detalles"] = []
            
            # Si hay un tipo de evento seleccionado pero no hay productos, intentar cargar productos guardados
            tipo_actual = st.session_state.get("act_tipo", "")
            if tipo_actual and not st.session_state["act_productos_detalles"]:
                productos_cargados = load_products_for_event_type(tipo_actual, inv_productos)
                if productos_cargados:
                    st.session_state["act_productos_detalles"] = productos_cargados
            
            # Selector para agregar nuevos productos
            productos_lista = [(f"{p['Codigo']} - {p['Nombre']}", p['Codigo']) for p in inv_productos]
            productos_disponibles = [opt[1] for opt in productos_lista]
            productos_ya_seleccionados = [prod['codigo'] for prod in st.session_state["act_productos_detalles"]]
            productos_para_agregar = [p for p in productos_disponibles if p not in productos_ya_seleccionados]
            
            col_add1, col_add2 = st.columns([3, 1])
            with col_add1:
                # Verificar si el valor actual del selectbox sigue siendo válido
                valor_actual = st.session_state.get("act_select_new_product", "")
                if valor_actual and valor_actual not in productos_para_agregar:
                    # Limpiar valor inválido
                    del st.session_state["act_select_new_product"]
                producto_nuevo = st.selectbox(
                    "Agregar Producto",
                    options=[""] + productos_para_agregar,
                    format_func=lambda x: "" if x == "" else next((opt[0] for opt in productos_lista if opt[1] == x), x),
                    key="act_select_new_product",
                    index=0
                )
            with col_add2:
                if st.button("➕ Agregar", key="act_add_product", use_container_width=True):
                    if producto_nuevo:
                        producto_info = next((p for p in inv_productos if p['Codigo'] == producto_nuevo), None)
                        if producto_info:
                            nuevo_producto = {
                                'codigo': producto_info['Codigo'],
                                'nombre': producto_info['Nombre'],
                                'categoria': producto_info.get('Categoria', ''),
                                'unidad': producto_info.get('Unidad', 'pc'),
                                'cantidad': 1.0,
                                'precio_unit': float(producto_info.get('PrecioVenta', 0.0)),
                                'descuento': 0.0,
                                'nota': ''
                            }
                            st.session_state["act_productos_detalles"].append(nuevo_producto)
                            st.rerun()
            
            # Mostrar tabla de productos seleccionados
            if st.session_state["act_productos_detalles"]:
                st.markdown("---")
                
                # Obtener stock de productos
                def get_stock_for_product(codigo):
                    """Obtener stock disponible de un producto"""
                    try:
                        stock_by_code = {}
                        wb = safe_load_workbook(EXCEL_PATH)
                        # base: stock inicial de productos
                        if "INVENTARIO_PRODUCTOS" in wb.sheetnames:
                            for row in wb["INVENTARIO_PRODUCTOS"].iter_rows(min_row=2, values_only=True):
                                if row and row[0]:
                                    code = str(row[0]).strip()
                                    stock_by_code[code] = float(row[5]) if len(row) > 5 and row[5] else 0.0
                        # movimientos
                        if "INVENTARIO_MOVIMIENTOS" in wb.sheetnames:
                            for row in wb["INVENTARIO_MOVIMIENTOS"].iter_rows(min_row=2, values_only=True):
                                if row and len(row) > 3 and row[1] and row[2] and row[3]:
                                    t = str(row[1]).strip().upper()
                                    code = str(row[2]).strip()
                                    qty = float(row[3]) if row[3] else 0.0
                                    if code not in stock_by_code:
                                        stock_by_code[code] = 0.0
                                    if t == "ENTRADA" or t == "AJUSTE+":
                                        stock_by_code[code] += qty
                                    elif t == "SALIDA" or t == "AJUSTE-":
                                        stock_by_code[code] -= qty
                        wb.close()
                        return stock_by_code.get(codigo, 0.0)
                    except Exception as e:
                        print(f"Error obteniendo stock: {e}")
                        return 0.0
                
                # Encabezados de la tabla
                col_h1, col_h2, col_h3, col_h4, col_h5, col_h6, col_h7 = st.columns([1, 3, 1, 1, 1, 1, 1])
                with col_h1:
                    st.markdown("**Código**")
                with col_h2:
                    st.markdown("**Producto**")
                with col_h3:
                    st.markdown("**Cantidad**")
                with col_h4:
                    st.markdown("**Precio (U)**")
                with col_h5:
                    st.markdown("**Stock**")
                with col_h6:
                    st.markdown("**Total**")
                with col_h7:
                    st.markdown("**Acción**")
                
                # Filas de productos
                productos_a_eliminar = []
                total_general = 0.0
                
                for idx, prod in enumerate(st.session_state["act_productos_detalles"]):
                    col_p1, col_p2, col_p3, col_p4, col_p5, col_p6, col_p7 = st.columns([1, 3, 1, 1, 1, 1, 1])
                    
                    with col_p1:
                        st.write(prod['codigo'])
                    with col_p2:
                        st.write(prod['nombre'])
                    with col_p3:
                        cantidad = st.number_input(
                            "",
                            min_value=0.0,
                            step=1.0,
                            value=float(prod['cantidad']),
                            key=f"act_qty_{idx}",
                            label_visibility="collapsed"
                        )
                        prod['cantidad'] = cantidad
                    with col_p4:
                        precio = st.number_input(
                            "",
                            min_value=0.0,
                            step=0.01,
                            value=float(prod['precio_unit']),
                            key=f"act_price_{idx}",
                            label_visibility="collapsed",
                            format="%.2f"
                        )
                        prod['precio_unit'] = precio
                    with col_p5:
                        # Mostrar stock disponible en lugar de descuento
                        stock_disponible = get_stock_for_product(prod['codigo'])
                        # Usar color para indicar si hay stock disponible
                        if stock_disponible >= cantidad:
                            st.write(f"✅ {stock_disponible:.0f}")
                        elif stock_disponible > 0:
                            st.write(f"⚠️ {stock_disponible:.0f}")
                        else:
                            st.write(f"❌ {stock_disponible:.0f}")
                    with col_p6:
                        total_linea = precio * cantidad  # Sin descuento
                        total_general += total_linea
                        st.write(f"${total_linea:.2f}")
                    with col_p7:
                        if st.button("🗑️", key=f"act_del_{idx}", help="Eliminar"):
                            # Marcar producto para eliminar después del formulario
                            if "act_productos_a_eliminar" not in st.session_state:
                                st.session_state["act_productos_a_eliminar"] = []
                            st.session_state["act_productos_a_eliminar"].append(idx)
                
                # Eliminar productos marcados (fuera del loop para evitar modificar durante iteración)
                if "act_productos_a_eliminar" in st.session_state and st.session_state["act_productos_a_eliminar"]:
                    indices_a_eliminar = sorted(set(st.session_state["act_productos_a_eliminar"]), reverse=True)
                    for idx in indices_a_eliminar:
                        if 0 <= idx < len(st.session_state["act_productos_detalles"]):
                            st.session_state["act_productos_detalles"].pop(idx)
                    st.session_state["act_productos_a_eliminar"] = []
                    st.rerun()
                
                # Resumen de totales
                st.markdown("---")
                col_sum1, col_sum2 = st.columns([2, 1])
                with col_sum1:
                    st.markdown(f"**Total: ${total_general:.2f}**")
            else:
                st.info("💡 Agrega productos usando el selector arriba")
        else:
            st.warning("⚠️ No hay productos en el inventario. Por favor, primero registra productos en la sección de Inventario.")
            st.text_area(
                "Productos",
                key="act_productos_text",
                height=150,
                placeholder="Ingresa los productos manualmente (uno por línea)"
            )
        
        st.markdown("---")
        
        with st.form("form_activacion", clear_on_submit=True):
            # Mostrar el tipo seleccionado por los botones (arriba)
            tipo = st.session_state.get("act_tipo", "")
            st.markdown(f"**Tipo seleccionado:** {tipo if tipo else '—'}")

            # Fila: Cliente y Comercial/Agente lado a lado
            col_cli, col_ag = st.columns(2)
            with col_cli:
                cliente = form_cliente_section(clientes, "act")
            with col_ag:
                # Campo de Comercial/Agente autocompletado según cliente
                agente_default = ""
                if "act_cliente_agente" in st.session_state and st.session_state["act_cliente_agente"]:
                    agente_cliente = st.session_state["act_cliente_agente"]
                    if agente_cliente in agentes:
                        agente_default = agente_cliente
                    elif agentes:
                        for ag in agentes:
                            if ag and (agente_cliente.lower() in str(ag).lower() or str(ag).lower() in agente_cliente.lower()):
                                agente_default = ag
                                break
                # Selección con índice del agente detectado
                indice_agente = 0
                if agentes:
                    try:
                        if agente_default and agente_default in agentes:
                            indice_agente = agentes.index(agente_default)
                        else:
                            agente_norm = str(agente_default).strip().lower()
                            for i, a in enumerate(agentes):
                                a_norm = str(a).strip().lower()
                                if agente_norm and (agente_norm in a_norm or a_norm in agente_norm):
                                    indice_agente = i
                                    break
                    except Exception:
                        indice_agente = 0
                # Usar una key estable para mantener el valor en la sesión
                st.selectbox("Comercial/Agente", agentes, key="act_comercial", index=indice_agente if agentes else 0)
            
            col3, col4 = st.columns(2)
            fecha = col3.date_input("Fecha *", key="act_fecha")
            # Usar index 1 para que "PROGRAMADO" sea el valor por defecto
            estado = col4.selectbox("Estado *", ["", "PROGRAMADO", "APROBADO", "REALIZADO", "CANCELADO"], key="act_estado", index=1)
        
            descripcion = st.text_area("Descripción", key="act_desc", height=100)
            
            # Botón de acción
            submit = st.form_submit_button("✅ Guardar Solicitud de Activación", use_container_width=True)
        
        # Procesar botón de Guardar
        if submit:
            # Obtener productos desde la tabla de detalles CON consolidación
            if "act_productos_detalles" in st.session_state and st.session_state["act_productos_detalles"]:
                # Consolidar productos por código
                productos_consolidados_save = {}
                for prod in st.session_state["act_productos_detalles"]:
                    if prod['cantidad'] > 0:
                        codigo = prod['codigo']
                        if codigo in productos_consolidados_save:
                            productos_consolidados_save[codigo]['cantidad'] += int(prod['cantidad'])
                        else:
                            productos_consolidados_save[codigo] = {
                                'cantidad': int(prod['cantidad']),
                                'nombre': prod['nombre']
                            }
                productos = "\n".join([
                    f"{info['cantidad']} {cod} - {info['nombre']}"
                    for cod, info in productos_consolidados_save.items()
                ])
            else:
                productos = st.session_state.get("act_productos_text", "")
            
            # Obtener comercial/agente del session state
            comercial = st.session_state.get("act_comercial", "")
            
            if not all([cliente, tipo, estado]):
                    st.error("❌ Completa los campos obligatorios (Cliente, Tipo de Evento y Estado)")
            else:
                # Validar stock antes de guardar
                productos_validar = []
                if "act_productos_detalles" in st.session_state and st.session_state["act_productos_detalles"]:
                    productos_validar = [prod for prod in st.session_state["act_productos_detalles"] if prod.get('cantidad', 0) > 0]
                
                if productos_validar:
                    stock_suficiente, productos_insuficientes = validar_stock_suficiente(productos_validar)
                    if not stock_suficiente:
                        mensaje_error = "❌ **Stock insuficiente para los siguientes productos:**\n\n"
                        for p in productos_insuficientes:
                            mensaje_error += f"- **{p['codigo']} - {p['nombre']}**: Solicitado: {int(p['solicitado'])}, Disponible: {int(p['disponible'])}\n"
                        st.error(mensaje_error)
                    else:
                        # Stock suficiente, guardar evento
                        next_id_val = next_id("EVENTOS")
                        values = [next_id_val, cliente, tipo, fecha.strftime("%Y-%m-%d"), estado, descripcion, productos, comercial, "", "", "", "", ""]
                        if append_row("EVENTOS", values):
                            # Guardar productos para este tipo de evento antes de limpiar
                            tipo_actual = st.session_state.get("act_tipo", tipo)
                            if "act_productos_detalles" in st.session_state and st.session_state["act_productos_detalles"]:
                                productos_a_guardar = [prod.copy() for prod in st.session_state["act_productos_detalles"]]
                                save_saved_products_for_type(tipo_actual, productos_a_guardar)
                            st.success("✅ **Solicitud de Activación guardada exitosamente**")
                            leer_eventos.clear()  # Limpiar caché
                            st.rerun()
                        else:
                            st.error("❌ No se pudo guardar la solicitud. Revisa los mensajes de error arriba.")
                else:
                    # No hay productos para validar, guardar normalmente
                    next_id_val = next_id("EVENTOS")
                    values = [next_id_val, cliente, tipo, fecha.strftime("%Y-%m-%d"), estado, descripcion, productos, comercial, "", "", "", "", ""]
                    if append_row("EVENTOS", values):
                        # Guardar productos para este tipo de evento antes de limpiar
                        tipo_actual = st.session_state.get("act_tipo", tipo)
                        if "act_productos_detalles" in st.session_state and st.session_state["act_productos_detalles"]:
                            productos_a_guardar = [prod.copy() for prod in st.session_state["act_productos_detalles"]]
                            save_saved_products_for_type(tipo_actual, productos_a_guardar)
                        st.success("✅ **Solicitud de Activación guardada exitosamente**")
                        leer_eventos.clear()  # Limpiar caché
                        st.rerun()
                    else:
                        st.error("❌ No se pudo guardar la solicitud. Revisa los mensajes de error arriba.")

        # Panel de pesos fuera del formulario (como antes)
        def parse_items(productos_text):
            items = []
            for raw in (productos_text or "").splitlines():
                s = raw.strip()
                if not s:
                    continue
                if s.startswith("-"):
                    s = s[1:].strip()
                qty = 1
                name = s
                m = re.match(r"^(\d+)\s*[xX]?\s*(.+)$", s)
                if m:
                    try:
                        qty = int(m.group(1))
                    except:
                        qty = 1
                    name = m.group(2).strip()
                items.append((name, qty))
            return items

        # Actualizar act_productos_text desde act_productos_detalles antes de calcular pesos
        if "act_productos_detalles" in st.session_state and st.session_state["act_productos_detalles"]:
            # Consolidar productos por código para evitar duplicados
            productos_consolidados = {}
            for prod in st.session_state["act_productos_detalles"]:
                if prod['cantidad'] > 0:
                    codigo = prod['codigo']
                    if codigo in productos_consolidados:
                        productos_consolidados[codigo]['cantidad'] += int(prod['cantidad'])
                    else:
                        productos_consolidados[codigo] = {
                            'cantidad': int(prod['cantidad']),
                            'nombre': prod['nombre']
                        }
            
            # Generar texto consolidado
            productos_text_updated = "\n".join([
                f"{info['cantidad']} {cod} - {info['nombre']}"
                for cod, info in productos_consolidados.items()
            ])
            st.session_state["act_productos_text"] = productos_text_updated

        items_out = parse_items(st.session_state.get("act_productos_text", ""))
        if items_out:
            with st.expander("⚖️ Pesos de artículos (kg)", expanded=False):
                col_btn_pesos, _ = st.columns([1,3])
                actualizar_click = col_btn_pesos.button("Actualizar pesos", key="btn_actualizar_pesos_tab2")

                # Nombres únicos
                nombres_unicos = []
                for name, _ in items_out:
                    if name not in nombres_unicos:
                        nombres_unicos.append(name)

                # Cargar pesos persistidos por TIPO de evento
                tipo_actual = st.session_state.get('act_tipo', '')
                if 'act_pesos_guardados' not in st.session_state or st.session_state.get('act_pesos_tipo') != tipo_actual:
                    st.session_state['act_pesos_guardados'] = load_saved_weights_for_type(tipo_actual)
                    st.session_state['act_pesos_tipo'] = tipo_actual

                # Aplicar persistidos
                for name in nombres_unicos:
                    if name in st.session_state['act_pesos_guardados']:
                        # Usar solo el código del producto en el slug para evitar colisiones
                        codigo_match = re.search(r'^([A-Z0-9_]+)', name)
                        codigo = codigo_match.group(1) if codigo_match else name[:20]
                        slug_set = f"peso_{codigo}"
                        if slug_set not in st.session_state:
                            st.session_state[slug_set] = float(st.session_state['act_pesos_guardados'][name])

                # Defaults por tipo si existen y no hay pesos aún o si se pulsa actualizar
                defaults_por_tipo = {
                    "ACTIVACION EXTREME": pesos_default_extreme,
                }
                if tipo_actual in defaults_por_tipo:
                    hay_pesos = any(
                        (f"peso_{re.search(r'^([A-Z0-9_]+)', n).group(1) if re.search(r'^([A-Z0-9_]+)', n) else n[:20]}" in st.session_state)
                        for n in nombres_unicos
                    )
                    if actualizar_click or not hay_pesos:
                        for name in nombres_unicos:
                            if name in defaults_por_tipo[tipo_actual]:
                                cantidad = sum(q for n, q in items_out if n == name) or 1
                                per_unit = (defaults_por_tipo[tipo_actual][name] / cantidad) if cantidad > 0 else defaults_por_tipo[tipo_actual][name]
                                codigo_match = re.search(r'^([A-Z0-9_]+)', name)
                                codigo = codigo_match.group(1) if codigo_match else name[:20]
                                slug_set = f"peso_{codigo}"
                                st.session_state[slug_set] = round(per_unit, 2)

                total_kg = 0.0
                for name in nombres_unicos:
                    codigo_match = re.search(r'^([A-Z0-9_]+)', name)
                    codigo = codigo_match.group(1) if codigo_match else name[:20]
                    slug = f"peso_{codigo}"
                    peso_val = st.number_input(
                        f"Peso kg - {name}",
                        min_value=0.0,
                        step=0.1,
                        key=slug
                    )
                    # Solo sumar el peso por unidad, sin multiplicar por cantidad
                    total_kg += (peso_val or 0.0)

                # Campo adicional para peso extra no listado (cajas, cables, etc.) por TIPO
                tipo_slug = re.sub(r'[^a-zA-Z0-9_]+', '_', tipo_actual or 'SIN_TIPO')[:40]
                extra_key = f"act_peso_extra_{tipo_slug}"
                # Precargar desde guardados del tipo
                if '__PESO_EXTRA__' in st.session_state.get('act_pesos_guardados', {}) and extra_key not in st.session_state:
                    try:
                        st.session_state[extra_key] = float(st.session_state['act_pesos_guardados']['__PESO_EXTRA__'])
                    except:
                        st.session_state[extra_key] = 0.0
                peso_extra = st.number_input(
                    "Peso extra (kg)",
                    min_value=0.0,
                    step=0.1,
                    key=extra_key
                )
                total_kg += (peso_extra or 0.0)

                st.metric("Peso total", f"{total_kg:.1f} kg")

                # Guardar cambios del usuario (no cuando se hace clic en Actualizar pesos)
                if not actualizar_click:
                    hubo_cambios = False
                    nuevos = dict(st.session_state.get('act_pesos_guardados', {}))
                    for name in nombres_unicos:
                        codigo_match = re.search(r'^([A-Z0-9_]+)', name)
                        codigo = codigo_match.group(1) if codigo_match else name[:20]
                        slug = f"peso_{codigo}"
                        val = float(st.session_state.get(slug, 0.0) or 0.0)
                        prev = float(nuevos.get(name, -1.0)) if name in nuevos else -1.0
                        if val >= 0.0 and (name not in nuevos or abs(val - prev) > 1e-9):
                            nuevos[name] = val
                            hubo_cambios = True
                    # Guardar peso extra por tipo
                    extra_val = float(st.session_state.get(extra_key, 0.0) or 0.0)
                    prev_extra = float(nuevos.get('__PESO_EXTRA__', -1.0)) if '__PESO_EXTRA__' in nuevos else -1.0
                    if extra_val >= 0.0 and abs(extra_val - prev_extra) > 1e-9:
                        nuevos['__PESO_EXTRA__'] = extra_val
                        hubo_cambios = True
                    if hubo_cambios:
                        if save_saved_weights_for_type(tipo_actual, nuevos):
                            st.session_state['act_pesos_guardados'] = nuevos
    
    with tab2:
        st.markdown("#### ✅ Solicitudes Aprobadas")
        
        if eventos_aprobados:
            st.info(f"📋 {len(eventos_aprobados)} solicitud(es) aprobada(s)")
            for evento in eventos_aprobados:
                with st.expander(f"✅ {evento['Tipo']} - {evento['Cliente']} (#{evento['ID']})", expanded=False):
                    col1, col2 = st.columns(2)
                    col1.write(f"**Cliente:** {evento.get('Cliente', 'N/A')}")
                    col1.write(f"**Tipo:** {evento.get('Tipo', 'N/A')}")
                    col1.write(f"**Fecha:** {evento.get('Fecha', 'N/A')}")
                    col2.write(f"**Estado:** {evento.get('Estado', 'N/A')}")
                    col2.write(f"**Comercial/Agente:** {evento.get('Comercial/Agente', 'N/A')}")
                    if evento.get('Descripcion'):
                        st.write(f"**Descripción:** {evento.get('Descripcion')}")
                    
                    # Mostrar productos
                    if evento.get('Productos'):
                        st.markdown("**📦 Productos Solicitados:**")
                        productos_lista = evento.get('Productos', '').split('\n')
                        for prod in productos_lista:
                            if prod.strip():
                                st.markdown(f"- {prod.strip()}")
                    
                    st.markdown("---")
                    
                    # Botón para generar PDF de autorización
                    col_gen, col_down = st.columns(2)
                    pdf_generado = False
                    pdf_data = None
                    pdf_nombre = None
                    
                    if col_gen.button(f"📄 Generar PDF de Autorización", key=f"gen_pdf_evento_{evento['ID']}"):
                        archivo_pdf = generar_pdf_autorizacion_evento(evento)
                        if archivo_pdf:
                            st.success(f"✅ PDF generado: {archivo_pdf.name}")
                            pdf_generado = True
                            try:
                                with open(str(archivo_pdf), 'rb') as f:
                                    pdf_data = f.read()
                                pdf_nombre = archivo_pdf.name
                            except Exception as e:
                                st.error(f"Error al leer PDF: {e}")
                        else:
                            st.error("❌ Error al generar el PDF")
                    
                    if pdf_generado and pdf_data:
                        col_down.download_button(
                            label="⬇️ Descargar PDF",
                            data=pdf_data,
                            file_name=pdf_nombre,
                            mime="application/pdf",
                            key=f"download_evento_{evento['ID']}"
                        )
                    
                    st.markdown("---")
                    
                    # Sección para subir foto autorizada
                    st.markdown("**📤 Subir Foto de Autorización Firmada**")
                    archivo_firmado = st.file_uploader(
                        "Selecciona la foto de autorización firmada",
                        type=['jpg', 'jpeg', 'png', 'pdf'],
                        key=f"upload_auth_{evento['ID']}"
                    )
                    
                    if archivo_firmado:
                        # Guardar archivo en carpeta del cliente
                        carpeta_cliente = Path(EXCEL_DIR) / "EVENTOS_AUTORIZACIONES" / evento['Cliente'].replace("/", "_").replace("\\", "_")
                        carpeta_cliente.mkdir(parents=True, exist_ok=True)
                        
                        # Determinar extensión según tipo de archivo
                        tipo_archivo = archivo_firmado.type
                        if 'pdf' in tipo_archivo:
                            extension = '.pdf'
                        elif 'jpeg' in tipo_archivo or 'jpg' in tipo_archivo:
                            extension = '.jpg'
                        elif 'png' in tipo_archivo:
                            extension = '.png'
                        else:
                            extension = '.pdf'
                        
                        nombre_archivo = carpeta_cliente / f"Autorizacion_Firmada_{evento['ID']}{extension}"
                        
                        with open(nombre_archivo, "wb") as f:
                            f.write(archivo_firmado.getbuffer())
                        st.success(f"✅ Archivo guardado: {nombre_archivo.name}")
                    
                    st.markdown("---")
                    
                    # Sección para información de envío
                    st.markdown("**📦 Información de Envío**")
                    col_env1, col_env2 = st.columns(2)
                    
                    with col_env1:
                        numero_guia = st.text_input(
                            "Número de Guía",
                            value=evento.get('Número de Guía', ''),
                            key=f"guia_{evento['ID']}"
                        )
                    
                    with col_env2:
                        observaciones_envio = st.text_area(
                            "Observaciones del Envío",
                            value=evento.get('Observaciones Envío', ''),
                            key=f"obs_env_{evento['ID']}",
                            height=80
                        )
                    
                    # Botón para finalizar evento (guarda envío y marca como finalizado)
                    if st.button(f"✅ Finalizar Evento", key=f"finalizar_{evento['ID']}", use_container_width=True, type="primary"):
                        with st.spinner("Finalizando evento..."):
                            # Primero guardar información de envío si hay datos
                            envio_guardado = True
                            if numero_guia or observaciones_envio:
                                envio_guardado = actualizar_envio_evento(evento['ID'], numero_guia, observaciones_envio)
                            
                            # Luego actualizar estado a REALIZADO
                            if envio_guardado:
                                if actualizar_estado_evento(evento['ID'], "REALIZADO"):
                                    st.success(f"✅ Evento #{evento['ID']} finalizado exitosamente")
                                    time.sleep(1.5)
                                    st.rerun()
                                else:
                                    st.error("❌ Error al finalizar el evento. Por favor verifica los logs.")
                            else:
                                st.error("❌ Error al guardar la información de envío.")
        else:
            st.info("No hay solicitudes aprobadas")
    
    with tab3:
        st.markdown("#### ✅ Activaciones Finalizadas")
        
        # Cargar productos del inventario para el control de retorno
        def load_inv_products():
            productos = []
            try:
                wb = safe_load_workbook(EXCEL_PATH)
                if "INVENTARIO_PRODUCTOS" in wb.sheetnames:
                    ws = wb["INVENTARIO_PRODUCTOS"]
                    headers = [cell.value for cell in ws[1]]
                    has_precio_venta = "PrecioVenta" in headers
                    is_old_format = "StockInicial" in headers
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            prod = {
                                "Codigo": str(row[0]).strip(),
                                "Nombre": str(row[1]).strip() if row[1] else "",
                                "Categoria": str(row[2]).strip() if row[2] else "",
                                "Unidad": str(row[3]).strip() if len(row) > 3 and row[3] else "pc",
                            }
                            if is_old_format and len(row) >= 9:
                                # Estructura antigua: PrecioVenta en posición 3
                                prod["PrecioVenta"] = float(row[3]) if row[3] else 0.0
                                prod["Unidad"] = str(row[4]).strip() if len(row) > 4 and row[4] else "pc"
                            elif has_precio_venta:
                                # Estructura nueva: PrecioVenta en la posición correspondiente
                                idx_precio = headers.index("PrecioVenta")
                                prod["PrecioVenta"] = float(row[idx_precio]) if len(row) > idx_precio and row[idx_precio] else 0.0
                            else:
                                prod["PrecioVenta"] = 0.0
                            productos.append(prod)
                    wb.close()
                else:
                    print("⚠️ No se encontró la hoja INVENTARIO_PRODUCTOS")
            except Exception as e:
                print(f"❌ Error al cargar productos: {e}")
            return productos
        
        inv_productos = load_inv_products()
        
        if eventos_realizados:
            st.info(f"📋 {len(eventos_realizados)} activación(es) finalizada(s)")
            for evento in eventos_realizados:
                with st.expander(f"✅ {evento['Tipo']} - {evento['Cliente']} (#{evento['ID']})", expanded=False):
                    col1, col2 = st.columns(2)
                    col1.write(f"**Cliente:** {evento.get('Cliente', 'N/A')}")
                    col1.write(f"**Tipo:** {evento.get('Tipo', 'N/A')}")
                    col1.write(f"**Fecha:** {evento.get('Fecha', 'N/A')}")
                    col2.write(f"**Estado:** {evento.get('Estado', 'N/A')}")
                    col2.write(f"**Comercial/Agente:** {evento.get('Comercial/Agente', 'N/A')}")
                    if evento.get('Descripcion'):
                        st.write(f"**Descripción:** {evento.get('Descripcion')}")
                    
                    # Mostrar productos
                    if evento.get('Productos'):
                        st.markdown("**📦 Productos Solicitados:**")
                        productos_lista = evento.get('Productos', '').split('\n')
                        for prod in productos_lista:
                            if prod.strip():
                                st.markdown(f"- {prod.strip()}")
                    
                    st.markdown("---")
                    
                    # Información de envío
                    if evento.get('Número de Guía') or evento.get('Observaciones Envío'):
                        st.markdown("**📦 Información de Envío:**")
                        if evento.get('Número de Guía'):
                            st.write(f"**Número de Guía:** {evento.get('Número de Guía')}")
                        if evento.get('Observaciones Envío'):
                            st.write(f"**Observaciones del Envío:** {evento.get('Observaciones Envío')}")
                    
                    st.markdown("---")
                    
                    # Sección de Control de Retorno
                    st.markdown("**🔄 Control de Retorno de Productos**")
                    
                    # Inicializar productos retornados en session_state si no existen
                    retorno_key = f"ret_productos_detalles_{evento['ID']}"
                    if retorno_key not in st.session_state:
                        st.session_state[retorno_key] = []
                        # Si hay productos retornados guardados, parsearlos
                        productos_ret_guardados = evento.get('Productos Retornados', '')
                        if productos_ret_guardados:
                            import re
                            for line in productos_ret_guardados.split('\n'):
                                line = line.strip()
                                if line:
                                    m = re.match(r"^(\d+)\s+(.+)$", line)
                                    if m:
                                        qty = int(m.group(1))
                                        nombre = m.group(2).strip()
                                        # Buscar el producto en el inventario
                                        codigo_match = re.match(r"^([A-Z0-9_]+)\s*-\s*", nombre)
                                        if codigo_match:
                                            codigo = codigo_match.group(1)
                                            producto_info = next((p for p in inv_productos if p['Codigo'] == codigo), None)
                                            if producto_info:
                                                st.session_state[retorno_key].append({
                                                    'codigo': codigo,
                                                    'nombre': producto_info['Nombre'],
                                                    'categoria': producto_info.get('Categoria', ''),
                                                    'cantidad': float(qty)
                                                })
                    
                    # Selector para agregar nuevos productos retornados
                    if inv_productos:
                        productos_lista = [(f"{p['Codigo']} - {p['Nombre']}", p['Codigo']) for p in inv_productos]
                        productos_disponibles = [opt[1] for opt in productos_lista]
                        productos_ya_retornados = [prod['codigo'] for prod in st.session_state[retorno_key]]
                        productos_para_agregar = [p for p in productos_disponibles if p not in productos_ya_retornados]
                        
                        col_add_ret1, col_add_ret2 = st.columns([3, 1])
                        with col_add_ret1:
                            valor_actual_ret = st.session_state.get(f"ret_select_new_product_{evento['ID']}", "")
                            if valor_actual_ret and valor_actual_ret not in productos_para_agregar:
                                del st.session_state[f"ret_select_new_product_{evento['ID']}"]
                            producto_nuevo_ret = st.selectbox(
                                "Agregar Producto Retornado",
                                options=[""] + productos_para_agregar,
                                format_func=lambda x: "" if x == "" else next((opt[0] for opt in productos_lista if opt[1] == x), x),
                                key=f"ret_select_new_product_{evento['ID']}",
                                index=0
                            )
                        with col_add_ret2:
                            if st.button("➕ Agregar", key=f"ret_add_product_{evento['ID']}", use_container_width=True):
                                if producto_nuevo_ret:
                                    producto_info = next((p for p in inv_productos if p['Codigo'] == producto_nuevo_ret), None)
                                    if producto_info:
                                        nuevo_producto_ret = {
                                            'codigo': producto_info['Codigo'],
                                            'nombre': producto_info['Nombre'],
                                            'categoria': producto_info.get('Categoria', ''),
                                            'cantidad': 1.0
                                        }
                                        st.session_state[retorno_key].append(nuevo_producto_ret)
                                        st.rerun()
                        
                        # Mostrar tabla de productos retornados
                        if st.session_state[retorno_key]:
                            st.markdown("---")
                            
                            # Encabezados de la tabla
                            col_h_ret1, col_h_ret2, col_h_ret3, col_h_ret4 = st.columns([1, 3, 1, 1])
                            with col_h_ret1:
                                st.markdown("**Código**")
                            with col_h_ret2:
                                st.markdown("**Producto**")
                            with col_h_ret3:
                                st.markdown("**Cantidad**")
                            with col_h_ret4:
                                st.markdown("**Acción**")
                            
                            # Filas de productos
                            productos_ret_a_eliminar = []
                            
                            for idx, prod in enumerate(st.session_state[retorno_key]):
                                col_r1, col_r2, col_r3, col_r4 = st.columns([1, 3, 1, 1])
                                
                                with col_r1:
                                    st.write(prod['codigo'])
                                with col_r2:
                                    st.write(prod['nombre'])
                                with col_r3:
                                    cantidad_ret = st.number_input(
                                        "",
                                        min_value=0.0,
                                        step=1.0,
                                        value=float(prod['cantidad']),
                                        key=f"ret_qty_{evento['ID']}_{idx}",
                                        label_visibility="collapsed"
                                    )
                                    prod['cantidad'] = cantidad_ret
                                with col_r4:
                                    if st.button("🗑️", key=f"ret_del_{evento['ID']}_{idx}", help="Eliminar"):
                                        productos_ret_a_eliminar.append(idx)
                            
                            # Eliminar productos marcados
                            if productos_ret_a_eliminar:
                                indices_a_eliminar = sorted(set(productos_ret_a_eliminar), reverse=True)
                                for idx in indices_a_eliminar:
                                    if 0 <= idx < len(st.session_state[retorno_key]):
                                        st.session_state[retorno_key].pop(idx)
                                st.rerun()
                    
                    st.markdown("---")
                    
                    # Fecha y observaciones en columnas
                    col_fecha_obs1, col_fecha_obs2 = st.columns(2)
                    with col_fecha_obs1:
                        fecha_retorno_val = None
                        if evento.get('Fecha Retorno'):
                            try:
                                fecha_retorno_val = datetime.strptime(str(evento.get('Fecha Retorno', '')), '%Y-%m-%d').date()
                            except:
                                fecha_retorno_val = None
                        fecha_retorno = st.date_input(
                            "Fecha de Retorno",
                            value=fecha_retorno_val,
                            key=f"ret_fecha_{evento['ID']}"
                        )
                    
                    with col_fecha_obs2:
                        pass  # Espacio en blanco para alineación
                    
                    observaciones_retorno = st.text_area(
                        "Observaciones del Retorno",
                        value=evento.get('Observaciones Retorno', ''),
                        key=f"ret_obs_{evento['ID']}",
                        height=80,
                        placeholder="Notas sobre el estado de los productos retornados..."
                    )
                    
                    # Botón para guardar información de retorno
                    if st.button(f"💾 Guardar Información de Retorno", key=f"guardar_retorno_{evento['ID']}", use_container_width=True):
                        with st.spinner("Guardando información de retorno..."):
                            # Convertir productos de la tabla interactiva a formato de texto
                            productos_retornados_text = "\n".join([
                                f"{int(prod['cantidad'])} {prod['codigo']} - {prod['nombre']}"
                                for prod in st.session_state[retorno_key]
                                if prod['cantidad'] > 0
                            ])
                            fecha_ret_str = fecha_retorno.strftime("%Y-%m-%d") if fecha_retorno else ""
                            if actualizar_retorno_evento(evento['ID'], productos_retornados_text, fecha_ret_str, observaciones_retorno):
                                st.success(f"✅ Información de retorno guardada para evento #{evento['ID']}")
                                time.sleep(1.5)
                                st.rerun()
                            else:
                                st.error("❌ Error al guardar la información de retorno.")
                    
                    # Comparación automática de productos enviados vs retornados
                    if evento.get('Productos') and st.session_state.get(retorno_key):
                        st.markdown("---")
                        # Parsear productos enviados
                        import re
                        productos_enviados = {}
                        for line in evento.get('Productos', '').split('\n'):
                            line = line.strip()
                            if line:
                                m = re.match(r"^(\d+)\s+(.+)$", line)
                                if m:
                                    qty = int(m.group(1))
                                    nombre = m.group(2).strip()
                                    productos_enviados[nombre] = qty
                        
                        # Parsear productos retornados de la tabla interactiva
                        productos_ret_parsed = {}
                        for prod in st.session_state[retorno_key]:
                            if prod['cantidad'] > 0:
                                nombre_completo = f"{prod['codigo']} - {prod['nombre']}"
                                productos_ret_parsed[nombre_completo] = int(prod['cantidad'])
                        
                        # Comparar
                        diferencias = []
                        productos_faltantes = []
                        productos_extra = []
                        
                        # Verificar productos enviados
                        for nombre, qty_env in productos_enviados.items():
                            if nombre in productos_ret_parsed:
                                qty_ret = productos_ret_parsed[nombre]
                                if qty_env != qty_ret:
                                    diferencias.append((nombre, qty_env, qty_ret))
                            else:
                                productos_faltantes.append((nombre, qty_env))
                        
                        # Verificar productos extra retornados
                        for nombre, qty_ret in productos_ret_parsed.items():
                            if nombre not in productos_enviados:
                                productos_extra.append((nombre, qty_ret))
                        
                        # Mostrar alertas
                        if diferencias or productos_faltantes or productos_extra:
                            st.markdown("**⚠️ Diferencias Detectadas:**")
                            if diferencias:
                                st.warning("📊 **Cantidades Diferentes:**")
                                for nombre, qty_env, qty_ret in diferencias:
                                    st.write(f"- **{nombre}**: Enviado: {qty_env}, Retornado: {qty_ret}")
                            if productos_faltantes:
                                st.error("❌ **Productos No Retornados:**")
                                for nombre, qty_env in productos_faltantes:
                                    st.write(f"- **{nombre}**: Enviado: {qty_env}, Retornado: 0")
                            if productos_extra:
                                st.warning("➕ **Productos Extra Retornados:**")
                                for nombre, qty_ret in productos_extra:
                                    st.write(f"- **{nombre}**: Cantidad: {qty_ret}")
                        else:
                            st.success("✅ Todos los productos coinciden correctamente")
                    
                    # Mostrar información de retorno ya guardada si existe
                    if evento.get('Productos Retornados') or evento.get('Fecha Retorno') or evento.get('Observaciones Retorno'):
                        st.markdown("---")
                        st.markdown("**✅ Información de Retorno Registrada:**")
                        if evento.get('Productos Retornados'):
                            st.markdown("**📦 Productos Retornados:**")
                            productos_ret_lista = evento.get('Productos Retornados', '').split('\n')
                            for prod in productos_ret_lista:
                                if prod.strip():
                                    st.markdown(f"- {prod.strip()}")
                        if evento.get('Fecha Retorno'):
                            st.write(f"**Fecha de Retorno:** {evento.get('Fecha Retorno')}")
                        if evento.get('Observaciones Retorno'):
                            st.write(f"**Observaciones:** {evento.get('Observaciones Retorno')}")
        else:
            st.info("No hay activaciones finalizadas")
    
    with tab4:
        st.markdown("#### 📊 Reportes de Eventos")
        
        # Filtros
        st.markdown("### 🔍 Filtros de Búsqueda")
        col_filtro1, col_filtro2 = st.columns(2)
        
        with col_filtro1:
            estado_filtro = st.selectbox(
                "Filtrar por Estado",
                ["Todos", "PROGRAMADO", "APROBADO", "REALIZADO", "CANCELADO"],
                key="reporte_estado"
            )
        
        with col_filtro2:
            buscar_texto = st.text_input(
                "Buscar por Cliente o Comercial",
                key="reporte_buscar",
                placeholder="Ingresa nombre de cliente o comercial..."
            )
        
        # Aplicar filtros
        eventos_filtrados = eventos.copy()
        
        if estado_filtro != "Todos":
            eventos_filtrados = [e for e in eventos_filtrados if str(e.get("Estado", "")).strip().upper() == estado_filtro.upper()]
        
        if buscar_texto:
            texto_buscar = buscar_texto.lower()
            eventos_filtrados = [e for e in eventos_filtrados if 
                                texto_buscar in str(e.get("Cliente", "")).lower() or 
                                texto_buscar in str(e.get("Comercial/Agente", "")).lower()]
        
        # Botón para descargar reporte en Excel
        st.markdown("---")
        col_desc1, col_desc2 = st.columns([1, 3])
        with col_desc1:
            if eventos_filtrados:
                # Crear DataFrame con los eventos filtrados
                df_report = pd.DataFrame(eventos_filtrados)
                
                # Convertir a bytes para descarga
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df_report.to_excel(writer, index=False, sheet_name='Reporte Eventos')
                
                excel_bytes = excel_buffer.getvalue()
                
                st.download_button(
                    "⬇️ Descargar Reporte Excel",
                    excel_bytes,
                    file_name=f"reporte_eventos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_reporte_eventos"
                )
        
        with col_desc2:
            st.markdown(f"### 📋 Resultados ({len(eventos_filtrados)} evento(s))")
        
        st.markdown("---")
        
        if eventos_filtrados:
            for evento in eventos_filtrados:
                # Determinar color según estado
                if evento.get("Estado") == "PROGRAMADO":
                    estado_color = "🔵"
                    estado_texto = "⏳ PROGRAMADO"
                elif evento.get("Estado") == "APROBADO":
                    estado_color = "🟡"
                    estado_texto = "✅ APROBADO"
                elif evento.get("Estado") == "REALIZADO":
                    estado_color = "🟢"
                    estado_texto = "✅ REALIZADO"
                else:
                    estado_color = "⚫"
                    estado_texto = evento.get("Estado", "N/A")
                
                with st.expander(f"{estado_texto} - {evento['Cliente']} (#{evento['ID']})", expanded=False):
                    col_info1, col_info2, col_info3 = st.columns(3)
                    
                    with col_info1:
                        st.write(f"**Cliente:** {evento.get('Cliente', 'N/A')}")
                        st.write(f"**Comercial/Agente:** {evento.get('Comercial/Agente', 'N/A')}")
                    
                    with col_info2:
                        st.write(f"**Tipo de Evento:** {evento.get('Tipo', 'N/A')}")
                        st.write(f"**Fecha:** {evento.get('Fecha', 'N/A')}")
                    
                    with col_info3:
                        st.write(f"**Estado:** {evento.get('Estado', 'N/A')}")
                        if evento.get('Número de Guía'):
                            st.write(f"**Número de Guía:** {evento.get('Número de Guía')}")
                    
                    if evento.get('Descripcion'):
                        st.write(f"**Descripción:** {evento.get('Descripcion')}")
                    
                    # Mostrar productos
                    if evento.get('Productos'):
                        st.markdown("**📦 Productos:**")
                        productos_lista = evento.get('Productos', '').split('\n')
                        for prod in productos_lista:
                            if prod.strip():
                                st.markdown(f"- {prod.strip()}")
                    
                    st.markdown("---")
                    
                    # Sección para fotos del evento (solo si está REALIZADO)
                    if evento.get("Estado") == "REALIZADO":
                        st.markdown("**📸 Fotos del Evento Realizado**")
                        
                        # Crear carpeta para las fotos del evento
                        carpeta_evento = Path(EXCEL_DIR) / "EVENTOS_FOTOS" / evento['Cliente'].replace("/", "_").replace("\\", "_")
                        carpeta_evento.mkdir(parents=True, exist_ok=True)
                        
                        # Buscar fotos existentes
                        fotos_existentes = []
                        fotos_patterns = [f"Evento_{evento['ID']}_*.jpg", f"Evento_{evento['ID']}_*.png", f"Evento_{evento['ID']}_*.jpeg"]
                        for pattern in fotos_patterns:
                            fotos_existentes.extend(list(carpeta_evento.glob(pattern)))
                        
                        # Mostrar fotos existentes
                        if fotos_existentes:
                            st.markdown(f"**Fotos existentes ({len(fotos_existentes)}):**")
                            fotos_cols = st.columns(min(len(fotos_existentes), 4))
                            for idx, foto_path in enumerate(fotos_existentes):
                                with fotos_cols[idx % 4]:
                                    st.image(str(foto_path), caption=foto_path.name, use_container_width=True)
                        
                        # Campo para subir nuevas fotos
                        st.markdown("**Subir Nueva Foto:**")
                        foto_subida = st.file_uploader(
                            f"Selecciona foto del evento #{evento['ID']}",
                            type=['jpg', 'jpeg', 'png'],
                            key=f"upload_foto_evento_{evento['ID']}"
                        )
                        
                        if foto_subida:
                            # Contar fotos existentes para numerar la nueva
                            num_fotos = len(fotos_existentes) + 1
                            
                            # Determinar extensión
                            tipo_archivo = foto_subida.type
                            if 'jpeg' in tipo_archivo or 'jpg' in tipo_archivo:
                                extension = '.jpg'
                            elif 'png' in tipo_archivo:
                                extension = '.png'
                            else:
                                extension = '.jpg'
                            
                            # Guardar foto
                            nombre_foto = carpeta_evento / f"Evento_{evento['ID']}_{num_fotos:03d}{extension}"
                            
                            foto_upload_key = f"foto_subida_{evento['ID']}_{num_fotos}"
                            if foto_upload_key not in st.session_state:
                                with open(nombre_foto, "wb") as f:
                                    f.write(foto_subida.getbuffer())
                                st.session_state[foto_upload_key] = True
                                st.success(f"✅ Foto guardada: {nombre_foto.name}")
                                st.rerun()
        else:
            st.info("No se encontraron eventos con los filtros aplicados")

# ===== SECCIÓN ENTREGA DE PUBLICIDAD =====
elif selected_menu == "📰 ENTREGA DE PUBLICIDAD":
    # Cargar productos del inventario
    def load_inv_products_pub():
        productos = []
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if "INVENTARIO_PRODUCTOS" in wb.sheetnames:
                ws = wb["INVENTARIO_PRODUCTOS"]
                headers = [cell.value for cell in ws[1]]
                has_precio_venta = "PrecioVenta" in headers
                is_old_format = "StockInicial" in headers
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        prod = {
                            "Codigo": str(row[0]).strip(),
                            "Nombre": str(row[1]).strip() if row[1] else "",
                            "Categoria": str(row[2]).strip() if row[2] else "",
                            "Unidad": str(row[3]).strip() if len(row) > 3 and row[3] else "pc",
                        }
                        if is_old_format and len(row) >= 9:
                            # Estructura antigua: PrecioVenta en posición 3
                            prod["PrecioVenta"] = float(row[3]) if row[3] else 0.0
                            prod["Unidad"] = str(row[4]).strip() if len(row) > 4 and row[4] else "pc"
                        elif has_precio_venta:
                            # Estructura nueva: PrecioVenta en posición 5
                            prod["PrecioVenta"] = float(row[5]) if len(row) > 5 and row[5] else 0.0
                        else:
                            # Sin PrecioVenta
                            prod["PrecioVenta"] = 0.0
                        productos.append(prod)
                wb.close()
        except Exception as e:
            print(f"Error al cargar productos: {e}")
        return productos
    
    inv_productos_pub = load_inv_products_pub()
    
    # Inicializar lista de productos
    if "pub_productos_detalles" not in st.session_state:
        st.session_state["pub_productos_detalles"] = []
    
    # Función para obtener stock de productos
    def get_stock_for_product_pub(codigo):
        """Obtener stock disponible de un producto"""
        try:
            stock_by_code = {}
            wb = safe_load_workbook(EXCEL_PATH)
            # base: stock inicial de productos
            if "INVENTARIO_PRODUCTOS" in wb.sheetnames:
                for row in wb["INVENTARIO_PRODUCTOS"].iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        code = str(row[0]).strip()
                        stock_by_code[code] = float(row[5]) if len(row) > 5 and row[5] else 0.0
            # movimientos
            if "INVENTARIO_MOVIMIENTOS" in wb.sheetnames:
                for row in wb["INVENTARIO_MOVIMIENTOS"].iter_rows(min_row=2, values_only=True):
                    if row and len(row) > 3 and row[1] and row[2] and row[3]:
                        t = str(row[1]).strip().upper()
                        code = str(row[2]).strip()
                        qty = float(row[3]) if row[3] else 0.0
                        if code not in stock_by_code:
                            stock_by_code[code] = 0.0
                        if t == "ENTRADA" or t == "AJUSTE+":
                            stock_by_code[code] += qty
                        elif t == "SALIDA" or t == "AJUSTE-":
                            stock_by_code[code] -= qty
            wb.close()
            return stock_by_code.get(codigo, 0.0)
        except Exception as e:
            print(f"Error obteniendo stock: {e}")
            return 0.0
    
    # Función para cargar entregas de publicidad
    def cargar_entregas_publicidad():
        """Cargar todas las entregas de publicidad desde Excel"""
        entregas = []
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if "PUBLICIDAD" in wb.sheetnames:
                ws = wb["PUBLICIDAD"]
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0] and row[1]:  # ID y Cliente deben existir
                        entrega = {
                            'ID': row[0],
                            'Cliente': str(row[1]).strip() if row[1] else '',
                            'Comercial': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                            'Fecha': str(row[3]).strip() if len(row) > 3 and row[3] else '',
                            'Productos': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                            'Observaciones': str(row[5]).strip() if len(row) > 5 and row[5] else ''
                        }
                        entregas.append(entrega)
            wb.close()
        except Exception as e:
            print(f"Error al cargar entregas de publicidad: {e}")
        return entregas
    
    # Tabs para organizar
    tab_productos, tab_reporte = st.tabs(["📦 Productos a Enviar", "📊 Reporte de Entregas por Cliente"])
    
    with tab_productos:
        # Productos a enviar (fuera del formulario para no reiniciarse)
        st.markdown("### 📦 Productos a Enviar")
        
        # Selector para agregar productos
        if inv_productos_pub:
            productos_lista_pub = [(f"{p['Codigo']} - {p['Nombre']}", p['Codigo']) for p in inv_productos_pub]
            productos_disponibles_pub = [opt[1] for opt in productos_lista_pub]
            productos_ya_seleccionados_pub = [prod['codigo'] for prod in st.session_state["pub_productos_detalles"]]
            productos_para_agregar_pub = [p for p in productos_disponibles_pub if p not in productos_ya_seleccionados_pub]
            
            col_add_pub1, col_add_pub2 = st.columns([3, 1])
            with col_add_pub1:
                valor_actual_pub = st.session_state.get("pub_select_new_product", "")
                if valor_actual_pub and valor_actual_pub not in productos_para_agregar_pub:
                    del st.session_state["pub_select_new_product"]
                producto_nuevo_pub = st.selectbox(
                    "Agregar Producto",
                    options=[""] + productos_para_agregar_pub,
                    format_func=lambda x: "" if x == "" else next((opt[0] for opt in productos_lista_pub if opt[1] == x), x),
                    key="pub_select_new_product",
                    index=0
                )
            with col_add_pub2:
                if st.button("➕ Agregar", key="pub_add_product", use_container_width=True):
                    if producto_nuevo_pub:
                        producto_info = next((p for p in inv_productos_pub if p['Codigo'] == producto_nuevo_pub), None)
                        if producto_info:
                            nuevo_producto_pub = {
                                'codigo': producto_info['Codigo'],
                                'nombre': producto_info['Nombre'],
                                'categoria': producto_info.get('Categoria', ''),
                                'cantidad': 1.0,
                                'precio_unit': float(producto_info.get('PrecioVenta', 0.0))
                            }
                            st.session_state["pub_productos_detalles"].append(nuevo_producto_pub)
                            st.rerun()
            
            # Mostrar stock del producto seleccionado antes de agregar
            if producto_nuevo_pub:
                stock_producto_seleccionado = get_stock_for_product_pub(producto_nuevo_pub)
                producto_nombre_seleccionado = next((p['Nombre'] for p in inv_productos_pub if p['Codigo'] == producto_nuevo_pub), "")
                if stock_producto_seleccionado >= 0:
                    st.info(f"📦 **Stock disponible:** {stock_producto_seleccionado:.0f} unidades - {producto_nombre_seleccionado}")
            
            # Mostrar tabla de productos
            if st.session_state["pub_productos_detalles"]:
                st.markdown("---")
                
                # Encabezados
                col_h_pub1, col_h_pub2, col_h_pub3, col_h_pub4, col_h_pub5, col_h_pub6 = st.columns([1, 3, 1, 1, 1, 1])
                with col_h_pub1:
                    st.markdown("**Código**")
                with col_h_pub2:
                    st.markdown("**Producto**")
                with col_h_pub3:
                    st.markdown("**Cantidad**")
                with col_h_pub4:
                    st.markdown("**Stock**")
                with col_h_pub5:
                    st.markdown("**Total**")
                with col_h_pub6:
                    st.markdown("**Acción**")
                
                # Filas
                productos_pub_a_eliminar = []
                total_general_pub = 0.0
                
                for idx, prod in enumerate(st.session_state["pub_productos_detalles"]):
                    col_pub1, col_pub2, col_pub3, col_pub4, col_pub5, col_pub6 = st.columns([1, 3, 1, 1, 1, 1])
                    
                    with col_pub1:
                        st.write(prod['codigo'])
                    with col_pub2:
                        st.write(prod['nombre'])
                    with col_pub3:
                        cantidad_pub = st.number_input(
                            "",
                            min_value=0.0,
                            step=1.0,
                            value=float(prod['cantidad']),
                            key=f"pub_qty_{idx}",
                            label_visibility="collapsed"
                        )
                        prod['cantidad'] = cantidad_pub
                    with col_pub4:
                        # Mostrar stock disponible
                        stock_disponible_pub = get_stock_for_product_pub(prod['codigo'])
                        # Usar color para indicar si hay stock disponible
                        if stock_disponible_pub >= cantidad_pub:
                            st.write(f"✅ {stock_disponible_pub:.0f}")
                        elif stock_disponible_pub > 0:
                            st.write(f"⚠️ {stock_disponible_pub:.0f}")
                        else:
                            st.write(f"❌ {stock_disponible_pub:.0f}")
                    with col_pub5:
                        # Mostrar costo total (cantidad × precio unitario)
                        # Si el producto no tiene precio_unit, intentar obtenerlo del inventario
                        if 'precio_unit' not in prod or prod.get('precio_unit', 0.0) == 0.0:
                            producto_info = next((p for p in inv_productos_pub if p['Codigo'] == prod['codigo']), None)
                            if producto_info:
                                prod['precio_unit'] = float(producto_info.get('PrecioVenta', 0.0))
                            else:
                                prod['precio_unit'] = 0.0
                        precio_unitario = prod.get('precio_unit', 0.0)
                        costo_total = cantidad_pub * precio_unitario
                        total_general_pub += costo_total
                        st.write(f"${costo_total:.2f}")
                    with col_pub6:
                        if st.button("🗑️", key=f"pub_del_{idx}", help="Eliminar"):
                            productos_pub_a_eliminar.append(idx)
                
                # Mostrar total general
                st.markdown("---")
                col_total1, col_total2 = st.columns([4, 1])
                with col_total1:
                    st.markdown("")
                with col_total2:
                    st.markdown(f"**Total: ${total_general_pub:.2f}**")
                
                # Eliminar productos marcados
                if productos_pub_a_eliminar:
                    indices_a_eliminar = sorted(set(productos_pub_a_eliminar), reverse=True)
                    for idx in indices_a_eliminar:
                        if 0 <= idx < len(st.session_state["pub_productos_detalles"]):
                            st.session_state["pub_productos_detalles"].pop(idx)
                    st.rerun()
        
        st.markdown("---")
        st.markdown("### 📋 Información de Entrega")
        
        with st.form("form_publicidad", clear_on_submit=True):
            st.markdown("#### 📋 Información del Cliente")
            
            # Selección de cliente
            if clientes:
                # Crear opciones con nombre fiscal y ciudad
                opciones_clientes = [""] + [f"{c['id']} - {c['nombre']} ({c['ciudad']})" for c in clientes]
                
                cliente_seleccionado = st.selectbox("Selecciona el Cliente *", opciones_clientes, key="pub_cliente")
                
                # Extraer información del cliente correctamente y validar
                cliente = None
                if cliente_seleccionado and len(cliente_seleccionado) > 0:
                    try:
                        # Extraer el ID (antes del " - ")
                        cliente_id = cliente_seleccionado.split(" - ")[0].strip()
                        
                        # Validar que el cliente existe en la lista completa de clientes
                        cliente_info = next((c for c in clientes if c['id'] == cliente_id), None)
                        
                        if cliente_info:
                            cliente = cliente_info['nombre']
                            
                            # Buscar monto de ventas del cliente
                            monto_ventas = 0
                            try:
                                if EXCEL_PATH.exists():
                                    wb_ventas = safe_load_workbook(EXCEL_PATH, tries=3, wait=0.1)
                                    cliente_norm = normalizar_nombre(cliente_info['nombre'])
                                    # Preferir VENTAS_VINCULADAS
                                    if "VENTAS_VINCULADAS" in wb_ventas.sheetnames:
                                        ws_v = wb_ventas["VENTAS_VINCULADAS"]
                                        for row in ws_v.iter_rows(min_row=2, values_only=True):
                                            if row and row[0]:
                                                if cliente_norm == normalizar_nombre(str(row[0]).strip()):
                                                    try:
                                                        monto_ventas += float(row[1]) if row[1] else 0
                                                    except:
                                                        pass
                                    # Fallback: VENTAS (Empresa, Venta_Bruta)
                                    elif "VENTAS" in wb_ventas.sheetnames:
                                        ws_v = wb_ventas["VENTAS"]
                                        for row in ws_v.iter_rows(min_row=2, values_only=True):
                                            if row and row[0]:
                                                if cliente_norm == normalizar_nombre(str(row[0]).strip()):
                                                    try:
                                                        monto_ventas += float(row[1]) if len(row) > 1 and row[1] else 0
                                                    except:
                                                        pass
                                    wb_ventas.close()
                            except:
                                pass
                            
                            # Mostrar monto de compra visiblemente al seleccionar
                            if monto_ventas > 0:
                                st.info(f"💰 **Venta Total del Cliente:** ${monto_ventas:,.2f}")
                            
                            # Guardar el agente del cliente en session_state para asignar automáticamente
                            agente_cliente = cliente_info.get('agente', '')
                            if agente_cliente and agente_cliente.strip() and agente_cliente != 'nan':
                                nuevo_agente = agente_cliente.strip()
                                # Actualizar session_state para que el selectbox lo use
                                st.session_state["pub_cliente_agente"] = nuevo_agente
                                # Actualizar el comercial automáticamente si existe en la lista
                                if nuevo_agente in agentes:
                                    st.session_state["pub_comercial"] = nuevo_agente
                            else:
                                # Limpiar si no hay agente
                                st.session_state["pub_cliente_agente"] = ""
                    except:
                        pass
        
            st.markdown("#### 👤 Información del Comercial/Vendedor")
            # Obtener el comercial automáticamente del cliente seleccionado
            comercial_automatico = st.session_state.get("pub_cliente_agente", "")
            
            # Opciones del selectbox
            opciones_comercial = [""] + agentes
            
            # Si hay un comercial automático del cliente, usarlo como valor por defecto
            valor_actual = st.session_state.get("pub_comercial", "")
            if comercial_automatico and comercial_automatico in agentes:
                # Si el valor actual no coincide con el automático, actualizarlo
                if valor_actual != comercial_automatico:
                    st.session_state["pub_comercial"] = comercial_automatico
                    valor_actual = comercial_automatico
            
            # Encontrar el índice del valor actual
            try:
                if valor_actual and valor_actual in opciones_comercial:
                    indice_comercial = opciones_comercial.index(valor_actual)
                else:
                    indice_comercial = 0
            except:
                indice_comercial = 0
            
            comercial_vendedor = st.selectbox(
                "Comercial/Vendedor *", 
                opciones_comercial, 
                key="pub_comercial",
                index=indice_comercial,
                format_func=lambda x: x if x else "Selecciona..."
            )
            
            # Mostrar mensaje si el comercial se asignó automáticamente
            if comercial_automatico and comercial_vendedor == comercial_automatico:
                st.success(f"✅ Comercial asignado automáticamente: **{comercial_automatico}**")
            
            st.markdown("#### 📅 Información de Entrega")
            fecha = st.date_input("Fecha de Entrega *", key="pub_fecha")
            
            observaciones = st.text_area("Observaciones", key="pub_obs", height=80, placeholder="Notas adicionales...")
            
            submit = st.form_submit_button("✅ Guardar Entrega de Publicidad", use_container_width=True)
        
        if submit:
            if not all([cliente, comercial_vendedor]):
                st.error("❌ Completa los campos obligatorios (Cliente y Comercial/Vendedor)")
            else:
                # Obtener productos
                if st.session_state.get("pub_productos_detalles"):
                    productos_text = "\n".join([
                        f"{int(prod['cantidad'])} {prod['codigo']} - {prod['nombre']}"
                        for prod in st.session_state["pub_productos_detalles"]
                        if prod['cantidad'] > 0
                    ])
                else:
                    productos_text = ""
                
                # Validar stock antes de guardar
                productos_validar_pub = []
                if st.session_state.get("pub_productos_detalles"):
                    productos_validar_pub = [prod for prod in st.session_state["pub_productos_detalles"] if prod.get('cantidad', 0) > 0]
                
                if productos_validar_pub:
                    # Función para validar stock
                    stock_suficiente_pub, productos_insuficientes_pub = validar_stock_suficiente(productos_validar_pub)
                    if not stock_suficiente_pub:
                        mensaje_error_pub = "❌ **Stock insuficiente para los siguientes productos:**\n\n"
                        for p in productos_insuficientes_pub:
                            mensaje_error_pub += f"- **{p['codigo']} - {p['nombre']}**: Solicitado: {int(p['solicitado'])}, Disponible: {int(p['disponible'])}\n"
                        st.error(mensaje_error_pub)
                    else:
                        # Guardar registro
                        next_id_val = next_id("PUBLICIDAD")
                        values = [next_id_val, cliente, comercial_vendedor, fecha.strftime("%Y-%m-%d"), productos_text, observaciones]
                        if append_row("PUBLICIDAD", values):
                            # Descontar stock
                            for prod in productos_validar_pub:
                                mov = {
                                    "Fecha": fecha.strftime("%Y-%m-%d"),
                                    "Tipo": "SALIDA",
                                    "Codigo": prod['codigo'],
                                    "Cantidad": float(prod['cantidad']),
                                    "CostoUnit": 0.0,
                                    "Proveedor": "",
                                    "Proceso": "Entrega Publicidad",
                                    "Nota": f"Entrega Publicidad - Cliente: {cliente}"
                                }
                                append_movement_global(mov)
                            
                            st.success("✅ **Entrega de Publicidad guardada exitosamente y stock descontado**")
                            # Limpiar productos
                            st.session_state["pub_productos_detalles"] = []
                            st.rerun()
                        else:
                            st.error("❌ No se pudo guardar la publicidad. Revisa los mensajes de error arriba.")
                else:
                    # No hay productos, solo guardar registro
                    next_id_val = next_id("PUBLICIDAD")
                    values = [next_id_val, cliente, comercial_vendedor, fecha.strftime("%Y-%m-%d"), productos_text, observaciones]
                    if append_row("PUBLICIDAD", values):
                        st.success("✅ **Entrega de Publicidad guardada exitosamente**")
                        st.session_state["pub_productos_detalles"] = []
                        st.rerun()
                    else:
                        st.error("❌ No se pudo guardar la publicidad. Revisa los mensajes de error arriba.")
    
    with tab_reporte:
        st.markdown("### 📊 Reporte de Entregas por Cliente")
        
        # Cargar entregas
        entregas = cargar_entregas_publicidad()
        
        if entregas:
            # Selector de cliente para filtrar
            clientes_con_entregas = sorted(list(set([e['Cliente'] for e in entregas if e['Cliente']])))
            cliente_filtro = st.selectbox(
                "Seleccionar Cliente",
                options=["Todos"] + clientes_con_entregas,
                key="pub_reporte_cliente"
            )
            
            # Filtrar entregas por cliente
            if cliente_filtro != "Todos":
                entregas_filtradas = [e for e in entregas if e['Cliente'] == cliente_filtro]
            else:
                entregas_filtradas = entregas
            
            # Mostrar resumen
            st.info(f"📋 **Total de entregas:** {len(entregas_filtradas)}")
            
            # Botón de descarga Excel
            if entregas_filtradas:
                # Preparar datos para Excel
                datos_excel = []
                for entrega in entregas_filtradas:
                    datos_excel.append({
                        'ID': entrega.get('ID', ''),
                        'Cliente': entrega.get('Cliente', ''),
                        'Comercial': entrega.get('Comercial', ''),
                        'Fecha': entrega.get('Fecha', ''),
                        'Productos': entrega.get('Productos', ''),
                        'Observaciones': entrega.get('Observaciones', '')
                    })
                
                df_reporte = pd.DataFrame(datos_excel)
                
                # Convertir a bytes para descarga
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df_reporte.to_excel(writer, index=False, sheet_name='Reporte Entregas')
                
                excel_bytes = excel_buffer.getvalue()
                
                st.download_button(
                    "⬇️ Descargar Reporte Excel",
                    excel_bytes,
                    file_name=f"reporte_entregas_publicidad_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_reporte_entregas"
                )
            
            st.markdown("---")
            
            # Agrupar productos por cliente
            productos_por_cliente = {}
            for entrega in entregas_filtradas:
                cliente = entrega['Cliente']
                if not cliente:
                    continue
                
                if cliente not in productos_por_cliente:
                    productos_por_cliente[cliente] = {}
                
                # Parsear productos de la entrega
                productos_text = entrega.get('Productos', '')
                if productos_text:
                    import re
                    for line in productos_text.split('\n'):
                        line = line.strip()
                        if line:
                            # Formato esperado: "cantidad CODIGO - NOMBRE"
                            match = re.match(r"^(\d+)\s+(.+)$", line)
                            if match:
                                cantidad = int(match.group(1))
                                producto_nombre = match.group(2).strip()
                                
                                # Extraer código si existe
                                codigo_match = re.match(r"^([A-Z0-9_]+)\s*-\s*", producto_nombre)
                                codigo = codigo_match.group(1) if codigo_match else producto_nombre
                                
                                if codigo in productos_por_cliente[cliente]:
                                    productos_por_cliente[cliente][codigo]['cantidad'] += cantidad
                                else:
                                    productos_por_cliente[cliente][codigo] = {
                                        'nombre': producto_nombre,
                                        'cantidad': cantidad
                                    }
            
            # Mostrar productos agrupados
            if productos_por_cliente:
                for cliente, productos in productos_por_cliente.items():
                    with st.expander(f"📦 **{cliente}**", expanded=False):
                        st.markdown(f"**Productos entregados:**")
                        total_productos_cliente = 0
                        for codigo, info in productos.items():
                            st.markdown(f"- **{info['cantidad']}** x {info['nombre']}")
                            total_productos_cliente += info['cantidad']
                        st.markdown(f"**Total productos:** {total_productos_cliente}")
                        
                        # Mostrar fechas de entrega para este cliente
                        entregas_cliente = [e for e in entregas_filtradas if e['Cliente'] == cliente]
                        if entregas_cliente:
                            st.markdown("**Entregas realizadas:**")
                            for entrega in entregas_cliente:
                                fecha_entrega = entrega.get('Fecha', 'N/A')
                                comercial = entrega.get('Comercial', 'N/A')
                                observaciones_entrega = entrega.get('Observaciones', '')
                                if observaciones_entrega:
                                    st.markdown(f"- 📅 {fecha_entrega} - 👤 {comercial}")
                                    st.markdown(f"  📝 **Observaciones:** {observaciones_entrega}")
                                else:
                                    st.markdown(f"- 📅 {fecha_entrega} - 👤 {comercial}")

# ===== SECCIÓN ENTREGA DE PERCHAS/EXHIBIDORES =====
elif selected_menu == "📦 ENTREGA DE PERCHAS/EXHIBIDORES":
    # Cargar agentes
    agentes = cargar_agentes()
    
    # Leer perchas
    perchas = leer_perchas()
    
    # Filtrar perchas por estado (normalizando para evitar problemas con espacios)
    perchas_registro = [p for p in perchas if str(p.get("Estado", "")).strip().upper() == "REGISTRO"]
    perchas_proceso = [p for p in perchas if str(p.get("Estado", "")).strip().upper() == "EN PROCESO"]
    perchas_entregadas = [p for p in perchas if str(p.get("Estado", "")).strip().upper() == "ENTREGADO"]
    
    # Tabs para organizar
    tab_registro, tab_proceso, tab_entregadas, tab_reportes = st.tabs(["📝 Registro", "⏳ En Proceso", "✅ Entregadas", "📊 Reportes"])
    
    with tab_registro:
        # Cargar productos del inventario
        def load_inv_products_perchas():
            """Cargar productos del inventario para perchas"""
            productos = []
            try:
                wb = safe_load_workbook(EXCEL_PATH)
                if "INVENTARIO_PRODUCTOS" in wb.sheetnames:
                    ws = wb["INVENTARIO_PRODUCTOS"]
                    headers = [cell.value for cell in ws[1]]
                    has_precio_venta = "PrecioVenta" in headers
                    is_old_format = "StockInicial" in headers
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            prod = {
                                "Codigo": str(row[0]).strip(),
                                "Nombre": str(row[1]).strip() if row[1] else "",
                                "Categoria": str(row[2]).strip() if row[2] else "",
                                "Unidad": str(row[3]).strip() if len(row) > 3 and row[3] else "pc",
                            }
                            if is_old_format and len(row) >= 9:
                                prod["PrecioVenta"] = float(row[3]) if row[3] else 0.0
                                prod["Unidad"] = str(row[4]).strip() if len(row) > 4 and row[4] else "pc"
                            elif has_precio_venta:
                                prod["PrecioVenta"] = float(row[5]) if len(row) > 5 and row[5] else 0.0
                            else:
                                prod["PrecioVenta"] = 0.0
                            productos.append(prod)
                wb.close()
            except Exception as e:
                print(f"Error al cargar productos: {e}")
            return productos
        
        def get_stock_for_product_perchas(codigo):
            """Obtener stock disponible de un producto"""
            try:
                stock_by_code = {}
                wb = safe_load_workbook(EXCEL_PATH)
                if "INVENTARIO_PRODUCTOS" in wb.sheetnames:
                    for row in wb["INVENTARIO_PRODUCTOS"].iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            code = str(row[0]).strip()
                            stock_by_code[code] = float(row[5]) if len(row) > 5 and row[5] else 0.0
                if "INVENTARIO_MOVIMIENTOS" in wb.sheetnames:
                    for row in wb["INVENTARIO_MOVIMIENTOS"].iter_rows(min_row=2, values_only=True):
                        if row and len(row) > 3 and row[1] and row[2] and row[3]:
                            t = str(row[1]).strip().upper()
                            code = str(row[2]).strip()
                            qty = float(row[3]) if row[3] else 0.0
                            if code not in stock_by_code:
                                stock_by_code[code] = 0.0
                            if t == "ENTRADA" or t == "AJUSTE+":
                                stock_by_code[code] += qty
                            elif t == "SALIDA" or t == "AJUSTE-":
                                stock_by_code[code] -= qty
                wb.close()
                return stock_by_code.get(codigo, 0.0)
            except Exception as e:
                print(f"Error obteniendo stock: {e}")
                return 0.0
        
        inv_productos_perchas = load_inv_products_perchas()
        
        # Inicializar lista de productos
        if "perchas_productos_detalles" not in st.session_state:
            st.session_state["perchas_productos_detalles"] = []
        
        # Productos a enviar (fuera del formulario para no reiniciarse)
        st.markdown("### 📦 Productos a Incluir")
        
        # Selector para agregar productos
        if inv_productos_perchas:
            productos_lista_perchas = [(f"{p['Codigo']} - {p['Nombre']}", p['Codigo']) for p in inv_productos_perchas]
            productos_disponibles_perchas = [opt[1] for opt in productos_lista_perchas]
            productos_ya_seleccionados_perchas = [prod['codigo'] for prod in st.session_state["perchas_productos_detalles"]]
            productos_para_agregar_perchas = [p for p in productos_disponibles_perchas if p not in productos_ya_seleccionados_perchas]
            
            col_add_perchas1, col_add_perchas2 = st.columns([3, 1])
            with col_add_perchas1:
                valor_actual_perchas = st.session_state.get("perchas_select_new_product", "")
                if valor_actual_perchas and valor_actual_perchas not in productos_para_agregar_perchas:
                    del st.session_state["perchas_select_new_product"]
                producto_nuevo_perchas = st.selectbox(
                    "Agregar Producto",
                    options=[""] + productos_para_agregar_perchas,
                    format_func=lambda x: "" if x == "" else next((opt[0] for opt in productos_lista_perchas if opt[1] == x), x),
                    key="perchas_select_new_product",
                    index=0
                )
            with col_add_perchas2:
                if st.button("➕ Agregar", key="perchas_add_product", use_container_width=True):
                    if producto_nuevo_perchas:
                        producto_info = next((p for p in inv_productos_perchas if p['Codigo'] == producto_nuevo_perchas), None)
                        if producto_info:
                            nuevo_producto_perchas = {
                                'codigo': producto_info['Codigo'],
                                'nombre': producto_info['Nombre'],
                                'categoria': producto_info.get('Categoria', ''),
                                'cantidad': 1.0,
                                'precio_unit': float(producto_info.get('PrecioVenta', 0.0))
                            }
                            st.session_state["perchas_productos_detalles"].append(nuevo_producto_perchas)
                            st.rerun()
            
            # Mostrar stock del producto seleccionado antes de agregar
            if producto_nuevo_perchas:
                stock_producto_seleccionado = get_stock_for_product_perchas(producto_nuevo_perchas)
                producto_nombre_seleccionado = next((p['Nombre'] for p in inv_productos_perchas if p['Codigo'] == producto_nuevo_perchas), "")
                if stock_producto_seleccionado >= 0:
                    st.info(f"📦 **Stock disponible:** {stock_producto_seleccionado:.0f} unidades - {producto_nombre_seleccionado}")
            
            # Mostrar tabla de productos
            if st.session_state["perchas_productos_detalles"]:
                st.markdown("---")
                
                # Encabezados
                col_h_perchas1, col_h_perchas2, col_h_perchas3, col_h_perchas4, col_h_perchas5, col_h_perchas6 = st.columns([1, 3, 1, 1, 1, 1])
                with col_h_perchas1:
                    st.markdown("**Código**")
                with col_h_perchas2:
                    st.markdown("**Producto**")
                with col_h_perchas3:
                    st.markdown("**Cantidad**")
                with col_h_perchas4:
                    st.markdown("**Stock**")
                with col_h_perchas5:
                    st.markdown("**Total**")
                with col_h_perchas6:
                    st.markdown("**Acción**")
                
                # Filas
                productos_perchas_a_eliminar = []
                total_general_perchas = 0.0
                
                for idx, prod in enumerate(st.session_state["perchas_productos_detalles"]):
                    col_perchas1, col_perchas2, col_perchas3, col_perchas4, col_perchas5, col_perchas6 = st.columns([1, 3, 1, 1, 1, 1])
                    
                    with col_perchas1:
                        st.write(prod['codigo'])
                    with col_perchas2:
                        st.write(prod['nombre'])
                    with col_perchas3:
                        cantidad_perchas = st.number_input(
                            "",
                            min_value=0.0,
                            step=1.0,
                            value=float(prod['cantidad']),
                            key=f"perchas_qty_{idx}",
                            label_visibility="collapsed"
                        )
                        prod['cantidad'] = cantidad_perchas
                    with col_perchas4:
                        # Mostrar stock disponible
                        stock_disponible_perchas = get_stock_for_product_perchas(prod['codigo'])
                        if stock_disponible_perchas >= cantidad_perchas:
                            st.write(f"✅ {stock_disponible_perchas:.0f}")
                        elif stock_disponible_perchas > 0:
                            st.write(f"⚠️ {stock_disponible_perchas:.0f}")
                        else:
                            st.write(f"❌ {stock_disponible_perchas:.0f}")
                    with col_perchas5:
                        # Mostrar costo total (cantidad × precio unitario)
                        if 'precio_unit' not in prod or prod.get('precio_unit', 0.0) == 0.0:
                            producto_info = next((p for p in inv_productos_perchas if p['Codigo'] == prod['codigo']), None)
                            if producto_info:
                                prod['precio_unit'] = float(producto_info.get('PrecioVenta', 0.0))
                            else:
                                prod['precio_unit'] = 0.0
                        precio_unitario_perchas = prod.get('precio_unit', 0.0)
                        costo_total_perchas = cantidad_perchas * precio_unitario_perchas
                        total_general_perchas += costo_total_perchas
                        st.write(f"${costo_total_perchas:.2f}")
                    with col_perchas6:
                        if st.button("🗑️", key=f"perchas_del_{idx}", help="Eliminar"):
                            productos_perchas_a_eliminar.append(idx)
                
                # Mostrar total general
                st.markdown("---")
                col_total_perchas1, col_total_perchas2 = st.columns([4, 1])
                with col_total_perchas1:
                    st.markdown("")
                with col_total_perchas2:
                    st.markdown(f"**Total: ${total_general_perchas:.2f}**")
                
                # Eliminar productos marcados
                if productos_perchas_a_eliminar:
                    indices_a_eliminar = sorted(set(productos_perchas_a_eliminar), reverse=True)
                    for idx in indices_a_eliminar:
                        if 0 <= idx < len(st.session_state["perchas_productos_detalles"]):
                            st.session_state["perchas_productos_detalles"].pop(idx)
                    st.rerun()
        
        st.markdown("---")
        
        # Formulario de registro de percha
        st.markdown("### 📦 REGISTRO DE PERCHA/EXHIBIDOR")
        
        with st.form("form_perchas", clear_on_submit=False):
            col1, col2 = st.columns(2)
            
            with col1:
                cliente = form_cliente_section(clientes, "p")
            
            with col2:
                # Campo Comercial/Vendedor
                st.markdown("#### 👤 Comercial/Vendedor")
                comercial_automatico_perchas = st.session_state.get("p_cliente_agente", "")
                opciones_comercial_perchas = [""] + agentes
                valor_actual_perchas = st.session_state.get("perchas_comercial", "")
                
                if comercial_automatico_perchas and comercial_automatico_perchas in agentes:
                    if valor_actual_perchas != comercial_automatico_perchas:
                        st.session_state["perchas_comercial"] = comercial_automatico_perchas
                        valor_actual_perchas = comercial_automatico_perchas
                
                try:
                    if valor_actual_perchas and valor_actual_perchas in opciones_comercial_perchas:
                        indice_comercial_perchas = opciones_comercial_perchas.index(valor_actual_perchas)
                    else:
                        indice_comercial_perchas = 0
                except:
                    indice_comercial_perchas = 0
                
                comercial_vendedor = st.selectbox(
                    "Comercial/Vendedor",
                    opciones_comercial_perchas,
                    key="perchas_comercial",
                    index=indice_comercial_perchas,
                    format_func=lambda x: x if x else "Selecciona..."
                )
                
                if comercial_automatico_perchas and comercial_vendedor == comercial_automatico_perchas:
                    st.success(f"✅ Comercial asignado automáticamente: **{comercial_automatico_perchas}**")
            
            col3, col4 = st.columns(2)
            fecha = col3.date_input("Fecha *", key="p_fecha")
            estado = "REGISTRO"  # Estado inicial fijo
            
            cantidad_compra = st.number_input("Cantidad de Compra", min_value=0, value=0, step=1, key="p_cantidad_compra")
            
            observaciones = st.text_area("Observaciones", key="p_obs", height=100)
            
            submit = st.form_submit_button("✅ Guardar Percha", use_container_width=True)
            
            if submit:
                if not cliente:
                    st.error("❌ Selecciona un cliente")
                else:
                    # Obtener productos
                    if st.session_state.get("perchas_productos_detalles"):
                        productos_text = "\n".join([
                            f"{int(prod['cantidad'])} {prod['codigo']} - {prod['nombre']}"
                            for prod in st.session_state["perchas_productos_detalles"]
                            if prod['cantidad'] > 0
                        ])
                    else:
                        productos_text = ""
                    
                    # Validar stock antes de guardar
                    productos_validar_perchas = []
                    if st.session_state.get("perchas_productos_detalles"):
                        productos_validar_perchas = [prod for prod in st.session_state["perchas_productos_detalles"] if prod.get('cantidad', 0) > 0]
                    
                    if productos_validar_perchas:
                        # Validar stock antes de guardar
                        stock_suficiente_perchas, productos_insuficientes_perchas = validar_stock_suficiente(productos_validar_perchas)
                        
                        # Guardar registro siempre, pero determinar el estado según el stock
                        next_id_val = next_id("PERCHAS")
                        estado_final = "REGISTRO"  # Por defecto, quedarse en REGISTRO si no hay stock
                        
                        if stock_suficiente_perchas:
                            # Si hay stock suficiente, se moverá a EN PROCESO y se descontará el stock
                            estado_final = "REGISTRO"  # Se guarda como REGISTRO, luego se actualiza a EN PROCESO
                        else:
                            # Si no hay stock suficiente, mostrar advertencia pero guardar igual
                            mensaje_advertencia_perchas = "⚠️ **Stock insuficiente para los siguientes productos:**\n\n"
                            for p in productos_insuficientes_perchas:
                                mensaje_advertencia_perchas += f"- **{p['codigo']} - {p['nombre']}**: Solicitado: {int(p['solicitado'])}, Disponible: {int(p['disponible'])}\n"
                            mensaje_advertencia_perchas += "\n📋 **La percha se guardará en estado 'REGISTRO' para verificación de stock.**"
                            st.warning(mensaje_advertencia_perchas)
                        
                        values = [next_id_val, cliente, comercial_vendedor, fecha.strftime("%Y-%m-%d"), estado_final, cantidad_compra, productos_text, observaciones, "", "", ""]
                        if append_row("PERCHAS", values):
                            if stock_suficiente_perchas:
                                # Descontar stock solo si hay stock suficiente
                                for prod in productos_validar_perchas:
                                    mov = {
                                        "Fecha": fecha.strftime("%Y-%m-%d"),
                                        "Tipo": "SALIDA",
                                        "Codigo": prod['codigo'],
                                        "Cantidad": float(prod['cantidad']),
                                        "CostoUnit": 0.0,
                                        "Proveedor": "",
                                        "Proceso": "Entrega Perchas",
                                        "Nota": f"Entrega Perchas - Cliente: {cliente}"
                                    }
                                    append_movement_global(mov)
                                
                                # Actualizar estado a "EN PROCESO" automáticamente
                                if actualizar_estado_percha(next_id_val, "EN PROCESO"):
                                    st.success("✅ **Percha guardada exitosamente, stock descontado y movida a 'EN PROCESO'**")
                                else:
                                    st.success("✅ **Percha guardada exitosamente con estado 'REGISTRO' y stock descontado**")
                                    st.warning("⚠️ No se pudo actualizar el estado a 'EN PROCESO' automáticamente")
                            else:
                                # Sin stock suficiente, quedarse en REGISTRO
                                st.info("📋 **Percha guardada en estado 'REGISTRO' pendiente de verificación de stock. Una vez que haya stock disponible, puedes moverla a 'EN PROCESO'.**")
                            
                            leer_perchas.clear()
                            st.session_state["perchas_productos_detalles"] = []
                            st.rerun()
                        else:
                            st.error("❌ No se pudo guardar la percha. Revisa los mensajes de error arriba.")
                    else:
                        # No hay productos, solo guardar registro
                        next_id_val = next_id("PERCHAS")
                        estado_final = "REGISTRO"
                        values = [next_id_val, cliente, comercial_vendedor, fecha.strftime("%Y-%m-%d"), estado_final, cantidad_compra, productos_text, observaciones, "", "", ""]
                        if append_row("PERCHAS", values):
                            # Actualizar estado a "EN PROCESO" automáticamente
                            if actualizar_estado_percha(next_id_val, "EN PROCESO"):
                                st.success("✅ **Percha guardada exitosamente y movida a 'EN PROCESO'**")
                            else:
                                st.success("✅ **Percha guardada exitosamente con estado 'REGISTRO'**")
                                st.warning("⚠️ No se pudo actualizar el estado a 'EN PROCESO' automáticamente")
                            leer_perchas.clear()
                            st.session_state["perchas_productos_detalles"] = []
                            st.rerun()
                        else:
                            st.error("❌ No se pudo guardar la percha. Revisa los mensajes de error arriba.")
        
        st.markdown("---")
        st.markdown("### 📋 Perchas en Registro")
        
        if perchas_registro:
            st.info(f"📋 {len(perchas_registro)} percha(s) en registro")
            for percha in perchas_registro:
                with st.expander(f"📦 Percha #{percha['ID']} - {percha['Cliente']}", expanded=False):
                    col1, col2 = st.columns(2)
                    col1.write(f"**Cliente:** {percha['Cliente']}")
                    col1.write(f"**Fecha:** {percha['Fecha']}")
                    if percha.get('Comercial/Vendedor'):
                        col1.write(f"**Comercial/Vendedor:** {percha['Comercial/Vendedor']}")
                    col2.write(f"**Estado:** {percha['Estado']}")
                    if percha.get('Cantidad de Compra'):
                        col2.write(f"**Cantidad de Compra:** {percha['Cantidad de Compra']}")
                    if percha.get('Número de Guía'):
                        col2.write(f"**Número de Guía:** {percha['Número de Guía']}")
                    if percha.get('Medio de Envío'):
                        col2.write(f"**Medio de Envío:** {percha['Medio de Envío']}")
                    if percha.get('Productos'):
                        st.write(f"**Productos:** {percha['Productos']}")
                    if percha.get('Observaciones'):
                        st.write(f"**Observaciones:** {percha['Observaciones']}")
                    
                    # Verificar stock si hay productos
                    if percha.get('Productos'):
                        productos_percha_text = percha['Productos']
                        productos_percha_lines = productos_percha_text.split('\n')
                        productos_para_validar = []
                        
                        for line in productos_percha_lines:
                            line = line.strip()
                            if not line:
                                continue
                            # Formato: "2 ES10001 - EXHIBIDOR DE BATERIAS"
                            import re
                            m = re.match(r"^(\d+)\s+(.+)$", line)
                            if m:
                                try:
                                    cantidad = float(m.group(1))
                                    codigo_nombre = m.group(2).strip()
                                    codigo_match = re.match(r"^([A-Z0-9_]+)\s*-\s*", codigo_nombre)
                                    if codigo_match:
                                        codigo = codigo_match.group(1).strip()
                                        nombre = codigo_nombre.replace(codigo + " -", "").strip()
                                        productos_para_validar.append({
                                            'codigo': codigo,
                                            'nombre': nombre,
                                            'cantidad': cantidad
                                        })
                                except:
                                    pass
                        
                        if productos_para_validar:
                            stock_suficiente_registro, productos_insuficientes_registro = validar_stock_suficiente(productos_para_validar)
                            
                            if stock_suficiente_registro:
                                st.success("✅ **Stock disponible - Listo para procesar**")
                                if st.button(f"▶️ Mover a 'EN PROCESO' y Descontar Stock", key=f"mover_proceso_{percha['ID']}", use_container_width=True):
                                    # Descontar stock
                                    for prod in productos_para_validar:
                                        mov = {
                                            "Fecha": percha.get('Fecha', datetime.now().strftime("%Y-%m-%d")),
                                            "Tipo": "SALIDA",
                                            "Codigo": prod['codigo'],
                                            "Cantidad": float(prod['cantidad']),
                                            "CostoUnit": 0.0,
                                            "Proveedor": "",
                                            "Proceso": "Entrega Perchas",
                                            "Nota": f"Entrega Perchas - Cliente: {percha['Cliente']}"
                                        }
                                        append_movement_global(mov)
                                    
                                    # Actualizar estado a "EN PROCESO"
                                    if actualizar_estado_percha(percha['ID'], "EN PROCESO"):
                                        st.success(f"✅ **Percha #{percha['ID']} movida a 'EN PROCESO' y stock descontado**")
                                        leer_perchas.clear()
                                        st.rerun()
                                    else:
                                        st.error("❌ No se pudo actualizar el estado de la percha")
                            else:
                                mensaje_stock_registro = "⚠️ **Stock insuficiente:**\n\n"
                                for p in productos_insuficientes_registro:
                                    mensaje_stock_registro += f"- **{p['codigo']} - {p['nombre']}**: Solicitado: {int(p['solicitado'])}, Disponible: {int(p['disponible'])}\n"
                                st.warning(mensaje_stock_registro)
                                st.info("📋 **Esperando disponibilidad de stock para procesar**")
                    else:
                        # No hay productos, puede moverse directamente a EN PROCESO
                        if st.button(f"▶️ Mover a 'EN PROCESO'", key=f"mover_proceso_sin_prod_{percha['ID']}", use_container_width=True):
                            if actualizar_estado_percha(percha['ID'], "EN PROCESO"):
                                st.success(f"✅ **Percha #{percha['ID']} movida a 'EN PROCESO'**")
                                leer_perchas.clear()
                                st.rerun()
                            else:
                                st.error("❌ No se pudo actualizar el estado de la percha")
        else:
            st.info("No hay perchas en registro")
    
    with tab_proceso:
        st.markdown("### ⏳ Perchas en Proceso")
        
        if perchas_proceso:
            st.info(f"⏳ {len(perchas_proceso)} percha(s) en proceso")
            for percha in perchas_proceso:
                with st.expander(f"⏳ Percha #{percha['ID']} - {percha['Cliente']}", expanded=False):
                    col1, col2 = st.columns(2)
                    col1.write(f"**Cliente:** {percha['Cliente']}")
                    col1.write(f"**Fecha:** {percha['Fecha']}")
                    if percha.get('Comercial/Vendedor'):
                        col1.write(f"**Comercial/Vendedor:** {percha['Comercial/Vendedor']}")
                    col2.write(f"**Estado:** {percha['Estado']}")
                    if percha.get('Cantidad de Compra'):
                        col2.write(f"**Cantidad de Compra:** {percha['Cantidad de Compra']}")
                    if percha.get('Número de Guía'):
                        col2.write(f"**Número de Guía:** {percha['Número de Guía']}")
                    if percha.get('Medio de Envío'):
                        col2.write(f"**Medio de Envío:** {percha['Medio de Envío']}")
                    if percha.get('Productos'):
                        st.write(f"**Productos:** {percha['Productos']}")
                    if percha.get('Observaciones'):
                        st.write(f"**Observaciones:** {percha['Observaciones']}")
                    
                    st.divider()
                    
                    # Formulario para marcar como entregado
                    with st.form(f"form_entregar_percha_{percha['ID']}", clear_on_submit=True):
                        st.markdown("#### 📦 Marcar como Entregado")
                        col_form1, col_form2 = st.columns(2)
                        numero_guia = col_form1.text_input("Número de Guía", key=f"guia_{percha['ID']}", placeholder="Ej: 1234567890")
                        medio_envio = col_form2.selectbox(
                            "Medio de Envío",
                            options=["", "Servientrega", "Servientrega Express", "Interrapidisimo", "DHL", "FedEx", "Envío Propio", "Otro"],
                            key=f"medio_{percha['ID']}"
                        )
                        
                        if medio_envio == "Otro":
                            medio_envio_otro = st.text_input("Especificar medio de envío", key=f"medio_otro_{percha['ID']}")
                            if medio_envio_otro:
                                medio_envio = medio_envio_otro
                        
                        submit_entregar = st.form_submit_button("✅ Marcar como Entregado", use_container_width=True)
                        
                        if submit_entregar:
                            if not numero_guia:
                                st.error("❌ Por favor ingresa el número de guía")
                            elif not medio_envio:
                                st.error("❌ Por favor selecciona el medio de envío")
                            else:
                                if actualizar_percha_entregada(percha['ID'], numero_guia, medio_envio):
                                    st.success(f"✅ **Percha #{percha['ID']} marcada como entregada exitosamente**")
                                    st.rerun()
                                else:
                                    st.error("❌ No se pudo actualizar la percha. Revisa los mensajes de error arriba.")
        else:
            st.info("No hay perchas en proceso")
    
    with tab_entregadas:
        st.markdown("### ✅ Perchas Entregadas")
        
        if perchas_entregadas:
            st.info(f"✅ {len(perchas_entregadas)} percha(s) entregada(s)")
            for percha in perchas_entregadas:
                with st.expander(f"✅ Percha #{percha['ID']} - {percha['Cliente']}", expanded=False):
                    col1, col2 = st.columns(2)
                    col1.write(f"**Cliente:** {percha['Cliente']}")
                    col1.write(f"**Fecha:** {percha['Fecha']}")
                    if percha.get('Comercial/Vendedor'):
                        col1.write(f"**Comercial/Vendedor:** {percha['Comercial/Vendedor']}")
                    col2.write(f"**Estado:** {percha['Estado']}")
                    if percha.get('Cantidad de Compra'):
                        col2.write(f"**Cantidad de Compra:** {percha['Cantidad de Compra']}")
                    if percha.get('Número de Guía'):
                        col2.write(f"**Número de Guía:** {percha['Número de Guía']}")
                    if percha.get('Medio de Envío'):
                        col2.write(f"**Medio de Envío:** {percha['Medio de Envío']}")
                    if percha.get('Productos'):
                        st.write(f"**Productos:** {percha['Productos']}")
                    if percha.get('Observaciones'):
                        st.write(f"**Observaciones:** {percha['Observaciones']}")
                    
                    st.markdown("---")
                    st.markdown("#### 📸 Fotos de la Instalación")
                    
                    # Mostrar fotos existentes si hay
                    fotos_existentes = []
                    if percha.get('Fotos'):
                        fotos_existentes = [f.strip() for f in percha['Fotos'].split('|') if f.strip()]
                    
                    if fotos_existentes:
                        st.markdown("**Fotos subidas:**")
                        # Crear columnas según la cantidad de fotos (máximo 3 por fila)
                        num_fotos = len(fotos_existentes)
                        num_columnas = min(3, num_fotos)
                        if num_columnas > 0:
                            cols_fotos = st.columns(num_columnas)
                            for idx, ruta_foto in enumerate(fotos_existentes):
                                col_idx = idx % num_columnas
                                with cols_fotos[col_idx]:
                                    try:
                                        if es_url(ruta_foto):
                                            st.image(ruta_foto, caption=f"Foto {idx + 1}", use_container_width=True)
                                        elif Path(ruta_foto).exists():
                                            st.image(str(ruta_foto), caption=f"Foto {idx + 1}", use_container_width=True)
                                        else:
                                            st.warning("No se encontró la foto almacenada")
                                    except Exception as e:
                                        st.error(f"Error al cargar foto: {e}")
                                    if (idx + 1) % num_columnas == 0 and (idx + 1) < num_fotos:
                                        cols_fotos = st.columns(num_columnas)
                    
                    # Formulario para subir nuevas fotos
                    with st.form(f"form_fotos_percha_{percha['ID']}", clear_on_submit=True):
                        fotos_uploaded = st.file_uploader(
                            "Subir fotos de la instalación",
                            type=['png', 'jpg', 'jpeg', 'gif', 'webp'],
                            accept_multiple_files=True,
                            key=f"fotos_percha_{percha['ID']}"
                        )
                        
                        submit_fotos = st.form_submit_button("📤 Subir Fotos", use_container_width=True)
                        
                        if submit_fotos:
                            if fotos_uploaded:
                                # Guardar fotos
                                rutas_fotos = guardar_fotos_percha(percha['ID'], fotos_uploaded)
                                
                                if rutas_fotos:
                                    # Actualizar Excel
                                    if actualizar_fotos_percha(percha['ID'], rutas_fotos):
                                        st.success(f"✅ {len(rutas_fotos)} foto(s) subida(s) exitosamente")
                                        leer_perchas.clear()
                                        st.rerun()
                                    else:
                                        st.error("❌ Error al guardar las fotos en el registro")
                                else:
                                    st.error("❌ Error al guardar las fotos en el sistema")
                            else:
                                st.warning("⚠️ Por favor selecciona al menos una foto")
        else:
            st.info("No hay perchas entregadas")
    
    with tab_reportes:
        st.markdown("### 📊 Reportes de Perchas/Exhibidores")
        
        if perchas:
            # Filtros de búsqueda
            st.markdown("#### 🔍 Filtros de Búsqueda")
            col_filtro1, col_filtro2, col_filtro3 = st.columns(3)
            
            with col_filtro1:
                # Filtro por cliente
                clientes_disponibles = sorted(list(set([p['Cliente'] for p in perchas if p.get('Cliente')])))
                cliente_filtro = st.selectbox(
                    "Filtrar por Cliente",
                    options=["Todos"] + clientes_disponibles,
                    key="perchas_reporte_cliente"
                )
            
            with col_filtro2:
                # Filtro por vendedor/comercial
                vendedores_disponibles = sorted(list(set([p.get('Comercial/Vendedor', '') for p in perchas if p.get('Comercial/Vendedor')])))
                vendedor_filtro = st.selectbox(
                    "Filtrar por Vendedor/Comercial",
                    options=["Todos"] + vendedores_disponibles,
                    key="perchas_reporte_vendedor"
                )
            
            with col_filtro3:
                # Filtro por estado
                estados_disponibles = sorted(list(set([p['Estado'] for p in perchas if p.get('Estado')])))
                estado_filtro = st.selectbox(
                    "Filtrar por Estado",
                    options=["Todos"] + estados_disponibles,
                    key="perchas_reporte_estado"
                )
            
            # Filtros de fecha
            col_fecha1, col_fecha2 = st.columns(2)
            with col_fecha1:
                fecha_desde = st.date_input(
                    "Fecha Desde",
                    value=None,
                    key="perchas_reporte_fecha_desde"
                )
            with col_fecha2:
                fecha_hasta = st.date_input(
                    "Fecha Hasta",
                    value=None,
                    key="perchas_reporte_fecha_hasta"
                )
            
            # Aplicar filtros
            perchas_filtradas = perchas.copy()
            
            # Filtrar por cliente
            if cliente_filtro != "Todos":
                perchas_filtradas = [p for p in perchas_filtradas if p.get('Cliente') == cliente_filtro]
            
            # Filtrar por vendedor/comercial
            if vendedor_filtro != "Todos":
                perchas_filtradas = [p for p in perchas_filtradas if p.get('Comercial/Vendedor') == vendedor_filtro]
            
            # Filtrar por estado
            if estado_filtro != "Todos":
                perchas_filtradas = [p for p in perchas_filtradas if p.get('Estado') == estado_filtro]
            
            # Filtrar por fechas
            if fecha_desde:
                fecha_desde_str = fecha_desde.strftime("%Y-%m-%d")
                perchas_filtradas = [p for p in perchas_filtradas if p.get('Fecha') and p.get('Fecha') >= fecha_desde_str]
            
            if fecha_hasta:
                fecha_hasta_str = fecha_hasta.strftime("%Y-%m-%d")
                perchas_filtradas = [p for p in perchas_filtradas if p.get('Fecha') and p.get('Fecha') <= fecha_hasta_str]
            
            # Botón para limpiar filtros
            if st.button("🔄 Limpiar Filtros", key="limpiar_filtros_perchas", use_container_width=False):
                # Limpiar todos los filtros del session_state
                if "perchas_reporte_cliente" in st.session_state:
                    del st.session_state["perchas_reporte_cliente"]
                if "perchas_reporte_vendedor" in st.session_state:
                    del st.session_state["perchas_reporte_vendedor"]
                if "perchas_reporte_estado" in st.session_state:
                    del st.session_state["perchas_reporte_estado"]
                if "perchas_reporte_fecha_desde" in st.session_state:
                    del st.session_state["perchas_reporte_fecha_desde"]
                if "perchas_reporte_fecha_hasta" in st.session_state:
                    del st.session_state["perchas_reporte_fecha_hasta"]
                st.rerun()
            
            st.markdown("---")
            
            # Mostrar resumen
            if perchas_filtradas:
                st.info(f"📋 **Total de perchas encontradas:** {len(perchas_filtradas)}")
            else:
                st.warning("⚠️ No se encontraron perchas con los filtros seleccionados. Intenta ajustar los filtros.")
            
            # Resumen por cliente
            if perchas_filtradas:
                st.markdown("---")
                st.markdown("### 📊 Resumen por Cliente")
                
                # Agrupar perchas por cliente
                from collections import defaultdict
                perchas_por_cliente = defaultdict(list)
                
                for percha in perchas_filtradas:
                    cliente_nombre = percha.get('Cliente', 'Sin Cliente')
                    fecha_envio = percha.get('Fecha', 'Sin fecha')
                    estado = percha.get('Estado', '')
                    numero_guia = percha.get('Número de Guía', '')
                    medio_envio = percha.get('Medio de Envío', '')
                    
                    perchas_por_cliente[cliente_nombre].append({
                        'ID': percha.get('ID', ''),
                        'Fecha': fecha_envio,
                        'Estado': estado,
                        'Número de Guía': numero_guia,
                        'Medio de Envío': medio_envio,
                        'Cantidad de Compra': percha.get('Cantidad de Compra', '')
                    })
                
                # Mostrar resumen para cada cliente
                for cliente_nombre, lista_perchas in sorted(perchas_por_cliente.items()):
                    cantidad_perchas = len(lista_perchas)
                    
                    # Obtener fechas de envío (fechas únicas)
                    fechas_envio = sorted(set([p['Fecha'] for p in lista_perchas if p['Fecha']]))
                    
                    # Contar perchas entregadas vs en proceso
                    entregadas = len([p for p in lista_perchas if p['Estado'] == 'ENTREGADO'])
                    en_proceso = len([p for p in lista_perchas if p['Estado'] == 'EN PROCESO'])
                    registro = len([p for p in lista_perchas if p['Estado'] == 'REGISTRO'])
                    
                    # Mensaje informativo
                    with st.expander(f"📦 **{cliente_nombre}** - {cantidad_perchas} percha(s)", expanded=False):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown(f"**Total de Perchas:** {cantidad_perchas}")
                            if entregadas > 0:
                                st.markdown(f"✅ **Entregadas:** {entregadas}")
                            if en_proceso > 0:
                                st.markdown(f"⏳ **En Proceso:** {en_proceso}")
                            if registro > 0:
                                st.markdown(f"📋 **En Registro:** {registro}")
                        
                        with col2:
                            if fechas_envio:
                                st.markdown(f"**Fechas de Envío:**")
                                for fecha_envio in fechas_envio:
                                    st.markdown(f"- {fecha_envio}")
                            
                            # Mostrar información de guías si hay perchas entregadas
                            if entregadas > 0:
                                st.markdown(f"**Perchas Entregadas:**")
                                for p in lista_perchas:
                                    if p['Estado'] == 'ENTREGADO':
                                        guia_info = f"Guía: {p['Número de Guía']}" if p['Número de Guía'] else "Sin guía"
                                        medio_info = f" - {p['Medio de Envío']}" if p['Medio de Envío'] else ""
                                        st.caption(f"Percha #{p['ID']}: {guia_info}{medio_info}")
                
                st.markdown("---")
            
            # Mostrar tabla de perchas
            if perchas_filtradas:
                datos_reporte = []
                for percha in perchas_filtradas:
                    # Obtener productos parseados
                    productos_percha = percha.get('Productos', '')
                    productos_lista = []
                    
                    if productos_percha:
                        productos_lines = productos_percha.split('\n')
                        import re
                        for line in productos_lines:
                            line = line.strip()
                            if not line:
                                continue
                            # Formato: "2 ES10001 - EXHIBIDOR DE BATERIAS"
                            m = re.match(r"^(\d+)\s+(.+)$", line)
                            if m:
                                try:
                                    cantidad = int(m.group(1))
                                    codigo_nombre = m.group(2).strip()
                                    codigo_match = re.match(r"^([A-Z0-9_]+)\s*-\s*", codigo_nombre)
                                    if codigo_match:
                                        codigo = codigo_match.group(1).strip()
                                        nombre = codigo_nombre.replace(codigo + " -", "").strip()
                                        productos_lista.append({
                                            'cantidad': cantidad,
                                            'codigo': codigo,
                                            'nombre': nombre
                                        })
                                except:
                                    pass
                    
                    # Si hay productos, crear una fila por cada producto
                    if productos_lista:
                        for producto in productos_lista:
                            datos_reporte.append({
                                'ID': percha.get('ID', ''),
                                'Cliente': percha.get('Cliente', ''),
                                'Comercial/Vendedor': percha.get('Comercial/Vendedor', ''),
                                'Fecha': percha.get('Fecha', ''),
                                'Estado': percha.get('Estado', ''),
                                'Cantidad de Compra': percha.get('Cantidad de Compra', ''),
                                'Cantidad Percha': producto['cantidad'],
                                'Código Producto': producto['codigo'],
                                'Nombre Producto': producto['nombre'],
                                'Observaciones': percha.get('Observaciones', ''),
                                'Número de Guía': percha.get('Número de Guía', ''),
                                'Medio de Envío': percha.get('Medio de Envío', '')
                            })
                    else:
                        # Si no hay productos, mostrar solo la información de la percha
                        datos_reporte.append({
                            'ID': percha.get('ID', ''),
                            'Cliente': percha.get('Cliente', ''),
                            'Comercial/Vendedor': percha.get('Comercial/Vendedor', ''),
                            'Fecha': percha.get('Fecha', ''),
                            'Estado': percha.get('Estado', ''),
                            'Cantidad de Compra': percha.get('Cantidad de Compra', ''),
                            'Cantidad Percha': '',
                            'Código Producto': '',
                            'Nombre Producto': '',
                            'Observaciones': percha.get('Observaciones', ''),
                            'Número de Guía': percha.get('Número de Guía', ''),
                            'Medio de Envío': percha.get('Medio de Envío', '')
                        })
                
                df_reporte = pd.DataFrame(datos_reporte)
                st.dataframe(df_reporte.sort_values(by=['Fecha', 'ID'], ascending=[False, True]), use_container_width=True, hide_index=True)
                
                # Descarga en formato Excel
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    # Hoja de resumen por cliente
                    resumen_data = []
                    for cliente_nombre, lista_perchas in sorted(perchas_por_cliente.items()):
                        cantidad_perchas = len(lista_perchas)
                        fechas_envio = sorted(set([p['Fecha'] for p in lista_perchas if p['Fecha']]))
                        entregadas = len([p for p in lista_perchas if p['Estado'] == 'ENTREGADO'])
                        en_proceso = len([p for p in lista_perchas if p['Estado'] == 'EN PROCESO'])
                        registro = len([p for p in lista_perchas if p['Estado'] == 'REGISTRO'])
                        
                        fechas_str = ", ".join(fechas_envio) if fechas_envio else "Sin fecha"
                        
                        resumen_data.append({
                            'Cliente': cliente_nombre,
                            'Total Perchas': cantidad_perchas,
                            'Entregadas': entregadas,
                            'En Proceso': en_proceso,
                            'En Registro': registro,
                            'Fechas de Envío': fechas_str
                        })
                    
                    df_resumen = pd.DataFrame(resumen_data)
                    df_resumen.to_excel(writer, index=False, sheet_name='Resumen por Cliente')
                    
                    # Hoja de detalle
                    df_reporte.to_excel(writer, index=False, sheet_name='Detalle Perchas')
                
                excel_bytes = excel_buffer.getvalue()
                
                st.download_button(
                    "⬇️ Descargar Reporte Excel",
                    excel_bytes,
                    file_name=f"reporte_perchas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_reporte_perchas"
                )
            else:
                st.info("No hay perchas para mostrar con el filtro seleccionado")
        else:
            st.info("No hay perchas registradas aún")

# ===== SECCIÓN ENTREGA A COMERCIALES =====
elif selected_menu == "💼 ENTREGA A COMERCIALES":
    st.markdown("### 💼 REGISTRO DE ENTREGA A COMERCIALES")
    
    # Cargar productos del inventario
    def load_inv_products_comerciales():
        productos = []
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if "INVENTARIO_PRODUCTOS" in wb.sheetnames:
                ws = wb["INVENTARIO_PRODUCTOS"]
                headers = [cell.value for cell in ws[1]]
                has_precio_venta = "PrecioVenta" in headers
                is_old_format = "StockInicial" in headers
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        prod = {
                            "Codigo": str(row[0]).strip(),
                            "Nombre": str(row[1]).strip() if row[1] else "",
                            "Categoria": str(row[2]).strip() if row[2] else "",
                            "Unidad": str(row[3]).strip() if len(row) > 3 and row[3] else "pc",
                        }
                        if is_old_format and len(row) >= 9:
                            prod["PrecioVenta"] = float(row[3]) if row[3] else 0.0
                            prod["Unidad"] = str(row[4]).strip() if len(row) > 4 and row[4] else "pc"
                        elif has_precio_venta:
                            prod["PrecioVenta"] = float(row[5]) if len(row) > 5 and row[5] else 0.0
                        else:
                            prod["PrecioVenta"] = 0.0
                        productos.append(prod)
                wb.close()
        except Exception as e:
            print(f"Error al cargar productos: {e}")
        return productos
    
    inv_productos_comerciales = load_inv_products_comerciales()
    
    # Inicializar lista de productos
    if "comerciales_productos_detalles" not in st.session_state:
        st.session_state["comerciales_productos_detalles"] = []
    
    # Función para obtener stock de productos
    def get_stock_for_product_comerciales(codigo):
        """Obtener stock disponible de un producto"""
        try:
            stock_by_code = {}
            wb = safe_load_workbook(EXCEL_PATH)
            if "INVENTARIO_PRODUCTOS" in wb.sheetnames:
                for row in wb["INVENTARIO_PRODUCTOS"].iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        code = str(row[0]).strip()
                        stock_by_code[code] = float(row[5]) if len(row) > 5 and row[5] else 0.0
            if "INVENTARIO_MOVIMIENTOS" in wb.sheetnames:
                for row in wb["INVENTARIO_MOVIMIENTOS"].iter_rows(min_row=2, values_only=True):
                    if row and len(row) > 3 and row[1] and row[2] and row[3]:
                        t = str(row[1]).strip().upper()
                        code = str(row[2]).strip()
                        qty = float(row[3]) if row[3] else 0.0
                        if code not in stock_by_code:
                            stock_by_code[code] = 0.0
                        if t == "ENTRADA" or t == "AJUSTE+":
                            stock_by_code[code] += qty
                        elif t == "SALIDA" or t == "AJUSTE-":
                            stock_by_code[code] -= qty
            wb.close()
            return stock_by_code.get(codigo, 0.0)
        except Exception as e:
            print(f"Error obteniendo stock: {e}")
            return 0.0
    
    # Funciones para gestionar inventario de vendedores
    def get_stock_vendedor(vendedor, codigo_producto):
        """Obtener stock disponible de un producto para un vendedor específico"""
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if "INVENTARIO_VENDEDORES" not in wb.sheetnames:
                wb.close()
                return 0.0
            
            ws = wb["INVENTARIO_VENDEDORES"]
            stock_total = 0.0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 3:
                    vendedor_row = str(row[0]).strip() if row[0] else ""
                    codigo_row = str(row[1]).strip() if row[1] else ""
                    cantidad = float(row[2]) if row[2] else 0.0
                    
                    if vendedor_row == vendedor and codigo_row == codigo_producto:
                        stock_total += cantidad
            
            wb.close()
            return stock_total
        except Exception as e:
            print(f"Error obteniendo stock del vendedor: {e}")
            return 0.0
    
    def agregar_stock_vendedor(vendedor, codigo_producto, cantidad):
        """Agregar stock a un vendedor"""
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            
            # Crear hoja si no existe
            if "INVENTARIO_VENDEDORES" not in wb.sheetnames:
                ws = wb.create_sheet("INVENTARIO_VENDEDORES")
                ws.append(["Vendedor", "Codigo_Producto", "Cantidad"])
            else:
                ws = wb["INVENTARIO_VENDEDORES"]
            
            # Buscar si ya existe un registro para este vendedor y producto
            encontrado = False
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value and row[1].value:
                    vendedor_row = str(row[0].value).strip()
                    codigo_row = str(row[1].value).strip()
                    
                    if vendedor_row == vendedor and codigo_row == codigo_producto:
                        # Actualizar cantidad existente
                        cantidad_actual = float(row[2].value) if row[2].value else 0.0
                        ws.cell(row=row_idx, column=3).value = cantidad_actual + cantidad
                        encontrado = True
                        break
            
            # Si no existe, agregar nuevo registro
            if not encontrado:
                ws.append([vendedor, codigo_producto, cantidad])
            
            if safe_save_workbook(wb, EXCEL_PATH):
                wb.close()
                print(f"✅ Stock agregado al vendedor {vendedor}: {cantidad} unidades de {codigo_producto}")
                return True
            
            wb.close()
            return False
        except Exception as e:
            print(f"Error agregando stock al vendedor: {e}")
            return False
    
    def descontar_stock_vendedor(vendedor, codigo_producto, cantidad):
        """Descontar stock de un vendedor"""
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            
            if "INVENTARIO_VENDEDORES" not in wb.sheetnames:
                wb.close()
                print(f"⚠️ No existe inventario para el vendedor {vendedor}")
                return False
            
            ws = wb["INVENTARIO_VENDEDORES"]
            
            # Buscar registro del vendedor y producto
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value and row[1].value:
                    vendedor_row = str(row[0].value).strip()
                    codigo_row = str(row[1].value).strip()
                    
                    if vendedor_row == vendedor and codigo_row == codigo_producto:
                        cantidad_actual = float(row[2].value) if row[2].value else 0.0
                        
                        if cantidad_actual >= cantidad:
                            nueva_cantidad = cantidad_actual - cantidad
                            ws.cell(row=row_idx, column=3).value = nueva_cantidad
                            
                            # Si la cantidad queda en 0 o menos, eliminar la fila
                            if nueva_cantidad <= 0:
                                ws.delete_rows(row_idx)
                            
                            if safe_save_workbook(wb, EXCEL_PATH):
                                wb.close()
                                print(f"✅ Stock descontado del vendedor {vendedor}: {cantidad} unidades de {codigo_producto}")
                                return True
                        else:
                            wb.close()
                            print(f"⚠️ Stock insuficiente para el vendedor {vendedor}: tiene {cantidad_actual}, necesita {cantidad}")
                            return False
            
            wb.close()
            print(f"⚠️ No se encontró el producto {codigo_producto} en el inventario del vendedor {vendedor}")
            return False
        except Exception as e:
            print(f"Error descontando stock del vendedor: {e}")
            return False
    
    def validar_stock_vendedor_suficiente(vendedor, productos):
        """Validar que el vendedor tiene stock suficiente para los productos"""
        productos_insuficientes = []
        
        for prod in productos:
            codigo = prod.get('codigo', '')
            cantidad_solicitada = float(prod.get('cantidad', 0))
            stock_disponible = get_stock_vendedor(vendedor, codigo)
            
            if cantidad_solicitada > stock_disponible:
                productos_insuficientes.append({
                    'codigo': codigo,
                    'nombre': prod.get('nombre', ''),
                    'solicitado': cantidad_solicitada,
                    'disponible': stock_disponible
                })
        
        return len(productos_insuficientes) == 0, productos_insuficientes
    
    # Función para calcular cantidades disponibles de una entrega (sin caché para siempre tener datos actualizados)
    def calcular_cantidades_disponibles(entrega_id_original):
        """Calcula las cantidades disponibles de una entrega original restando las entregas a clientes ya realizadas"""
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if "COMERCIALES" not in wb.sheetnames:
                wb.close()
                return {}
            
            ws = wb["COMERCIALES"]
            headers = [cell.value for cell in ws[1]]
            
            # Buscar la entrega original
            entrega_original = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] and str(row[0]).strip() == str(entrega_id_original):
                    productos_str = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                    if productos_str:
                        entrega_original = productos_str
                    break
            
            if not entrega_original:
                wb.close()
                return {}
            
            # Parsear productos originales
            import re
            productos_originales = {}
            for prod_line in entrega_original.split('\n'):
                if prod_line.strip():
                    match = re.match(r"^(\d+)\s+(.+)$", prod_line.strip())
                    if match:
                        cantidad = int(match.group(1))
                        codigo_nombre = match.group(2).strip()
                        codigo_match = re.match(r"^([A-Z0-9_]+)\s*-\s*", codigo_nombre)
                        codigo = codigo_match.group(1) if codigo_match else ""
                        # Si no hay código con guión, intentar extraer de otra forma
                        if not codigo:
                            # Buscar código al inicio (letras y números hasta el primer espacio)
                            codigo_match2 = re.match(r"^([A-Z0-9_]+)", codigo_nombre)
                            codigo = codigo_match2.group(1) if codigo_match2 else ""
                        if codigo:
                            productos_originales[codigo] = cantidad
                            print(f"🔍 Debug calcular_cantidades: Producto original - Código: {codigo}, Cantidad: {cantidad}, Texto completo: '{codigo_nombre}'")
            
            # Buscar todas las entregas a clientes relacionadas con esta entrega original
            # Buscamos entregas que tengan el ID de entrega original en el campo "ID_Entrega_Original"
            productos_entregados = {}
            
            # Normalizar el ID original para comparación (convertir a string y limpiar)
            # Asegurar que sea string, pero también mantener el valor numérico para comparación flexible
            id_original_normalizado = str(entrega_id_original).strip()
            # También crear versión numérica si es posible
            try:
                id_original_num = float(entrega_id_original)
            except:
                id_original_num = None
            
            # Buscar columna ID_Entrega_Original
            id_entrega_original_col_idx = None
            productos_entregados_col_idx = None
            tipo_entrega_col_idx = None
            
            for idx, header in enumerate(headers):
                if header == "ID_Entrega_Original":
                    id_entrega_original_col_idx = idx  # Índice basado en 0 para acceder a row
                elif header == "Productos Entregados":
                    productos_entregados_col_idx = idx
                elif header == "Tipo Entrega":
                    tipo_entrega_col_idx = idx
            
            print(f"🔍 Debug calcular_cantidades: ===== INICIO CÁLCULO CANTIDADES DISPONIBLES =====")
            print(f"🔍 Debug calcular_cantidades: Buscando entregas relacionadas con ID original: {id_original_normalizado} (raw: {entrega_id_original}, type: {type(entrega_id_original)})")
            print(f"🔍 Debug calcular_cantidades: ID numérico: {id_original_num}")
            print(f"🔍 Debug calcular_cantidades: Columna ID_Entrega_Original está en índice: {id_entrega_original_col_idx}")
            print(f"🔍 Debug calcular_cantidades: Columna Productos Entregados está en índice: {productos_entregados_col_idx}")
            print(f"🔍 Debug calcular_cantidades: Columna Tipo Entrega está en índice: {tipo_entrega_col_idx}")
            print(f"🔍 Debug calcular_cantidades: Total filas en hoja: {ws.max_row}")
            print(f"🔍 Debug calcular_cantidades: Headers encontrados: {headers}")
            
            entregas_encontradas = 0
            total_filas_procesadas = 0
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue
                
                total_filas_procesadas += 1
                
                # NO contar la entrega original misma (debe tener el mismo ID)
                id_fila_actual = str(row[0]).strip() if row[0] else ''
                if id_fila_actual == id_original_normalizado:
                    print(f"🔍 Debug calcular_cantidades: Fila {row_idx} es la entrega original (ID={id_fila_actual}), saltando...")
                    continue
                
                # Verificar tipo de entrega primero (debe ser "ENTREGA COMERCIAL A CLIENTE")
                if tipo_entrega_col_idx is not None and len(row) > tipo_entrega_col_idx:
                    tipo_entrega_row = str(row[tipo_entrega_col_idx]).strip() if row[tipo_entrega_col_idx] else ''
                else:
                    tipo_entrega_row = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                
                if tipo_entrega_row != "ENTREGA COMERCIAL A CLIENTE":
                    continue
                
                # Verificar si esta entrega está relacionada con la entrega original
                esta_relacionada = False
                
                # Verificar en columna ID_Entrega_Original si existe
                if id_entrega_original_col_idx is not None:
                    if len(row) > id_entrega_original_col_idx:
                        id_ref_raw = row[id_entrega_original_col_idx]
                        if id_ref_raw is not None:
                            # Normalizar el ID de referencia (puede ser número o string)
                            id_ref = str(id_ref_raw).strip()
                            
                            # Comparar normalizados (también intentar comparar como números si ambos son numéricos)
                            match_encontrado = False
                            
                            # Primero intentar comparación exacta de strings
                            if id_ref == id_original_normalizado:
                                esta_relacionada = True
                                match_encontrado = True
                                print(f"🔍 Debug calcular_cantidades: ✅ Match exacto string - Entrega relacionada encontrada en fila {row_idx} (ID={id_fila_actual}, ID_Original='{id_ref}', Original='{id_original_normalizado}', Tipo={tipo_entrega_row})")
                            else:
                                # Intentar comparar como números si ambos son numéricos
                                try:
                                    id_ref_num = float(id_ref_raw)  # Usar el valor raw directamente
                                    if id_original_num is not None and id_ref_num == id_original_num:
                                        esta_relacionada = True
                                        match_encontrado = True
                                        print(f"🔍 Debug calcular_cantidades: ✅ Match numérico - Entrega relacionada encontrada en fila {row_idx} (ID={id_fila_actual}, ID_Original={id_ref_raw}->{id_ref_num}, Original={entrega_id_original}->{id_original_num}, Tipo={tipo_entrega_row})")
                                    elif id_original_num is not None:
                                        print(f"🔍 Debug calcular_cantidades: ❌ No match numérico - Fila {row_idx} (ID={id_fila_actual}, ID_Original={id_ref_raw}->{id_ref_num}, Original={entrega_id_original}->{id_original_num})")
                                except Exception as e:
                                    print(f"🔍 Debug calcular_cantidades: ❌ No match - Fila {row_idx} (ID={id_fila_actual}, ID_Original='{id_ref}' (raw: {id_ref_raw}, type: {type(id_ref_raw)}), Original='{id_original_normalizado}' (raw: {entrega_id_original}, type: {type(entrega_id_original)}), Error: {e})")
                            
                            if not match_encontrado and id_ref_raw != "" and id_ref_raw is not None:
                                # Debug adicional: mostrar todos los valores para diagnóstico
                                print(f"🔍 Debug calcular_cantidades: ⚠️ Comparación fallida - Fila {row_idx}:")
                                print(f"   - ID fila actual: {id_fila_actual} (type: {type(row[0])})")
                                print(f"   - ID_Entrega_Original raw: {id_ref_raw} (type: {type(id_ref_raw)})")
                                print(f"   - ID_Entrega_Original str: '{id_ref}'")
                                print(f"   - ID original buscado raw: {entrega_id_original} (type: {type(entrega_id_original)})")
                                print(f"   - ID original buscado str: '{id_original_normalizado}'")
                                print(f"   - Tipo entrega: {tipo_entrega_row}")
                                print(f"   - Total columnas en fila: {len(row)}")
                        else:
                            print(f"🔍 Debug calcular_cantidades: ⚠️ Fila {row_idx} tiene ID_Entrega_Original vacío o None")
                    else:
                        print(f"🔍 Debug calcular_cantidades: ⚠️ Fila {row_idx} no tiene suficientes columnas (tiene {len(row)}, necesita {id_entrega_original_col_idx + 1})")
                else:
                    print(f"🔍 Debug calcular_cantidades: ⚠️ Columna ID_Entrega_Original no existe en headers")
                
                if esta_relacionada:
                    entregas_encontradas += 1
                    # Leer productos entregados usando el índice correcto basado en headers
                    if productos_entregados_col_idx is not None and len(row) > productos_entregados_col_idx:
                        productos_entregados_str = str(row[productos_entregados_col_idx]).strip() if row[productos_entregados_col_idx] else ''
                    else:
                        # Fallback al índice fijo si no se encuentra la columna
                        productos_entregados_str = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                    print(f"🔍 Debug calcular_cantidades: Procesando entrega fila {row_idx}, Productos Entregados: '{productos_entregados_str}'")
                    
                    if productos_entregados_str:
                        for prod_line in productos_entregados_str.split('\n'):
                            if prod_line.strip():
                                match = re.match(r"^(\d+)\s+(.+)$", prod_line.strip())
                                if match:
                                    cantidad_entregada = int(match.group(1))
                                    codigo_nombre = match.group(2).strip()
                                    codigo_match = re.match(r"^([A-Z0-9_]+)\s*-\s*", codigo_nombre)
                                    codigo = codigo_match.group(1) if codigo_match else ""
                                    # Si no hay código con guión, intentar extraer de otra forma
                                    if not codigo:
                                        codigo_match2 = re.match(r"^([A-Z0-9_]+)", codigo_nombre)
                                        codigo = codigo_match2.group(1) if codigo_match2 else ""
                                    if codigo and codigo in productos_originales:
                                        cantidad_anterior = productos_entregados.get(codigo, 0)
                                        productos_entregados[codigo] = cantidad_anterior + cantidad_entregada
                                        print(f"🔍 Debug calcular_cantidades: ✅ Producto entregado - {codigo}: +{cantidad_entregada} (Anterior: {cantidad_anterior}, Total acumulado: {productos_entregados[codigo]})")
                                    elif codigo:
                                        print(f"🔍 Debug calcular_cantidades: ⚠️ Código '{codigo}' no está en productos originales {list(productos_originales.keys())}")
            
            print(f"🔍 Debug calcular_cantidades: Total filas procesadas: {total_filas_procesadas}")
            print(f"🔍 Debug calcular_cantidades: Total entregas relacionadas encontradas: {entregas_encontradas}")
            print(f"🔍 Debug calcular_cantidades: Productos entregados acumulados: {productos_entregados}")
            
            # Si no se encontraron entregas relacionadas, mostrar advertencia
            if entregas_encontradas == 0:
                print(f"⚠️ Debug calcular_cantidades: NO SE ENCONTRARON ENTREGAS RELACIONADAS para ID {entrega_id_original}")
                print(f"   Esto puede significar que:")
                print(f"   1. El ID_Entrega_Original no se está guardando correctamente")
                print(f"   2. El ID no coincide (buscando: {id_original_normalizado}, tipo: {type(entrega_id_original)})")
                print(f"   3. La columna ID_Entrega_Original está en índice: {id_entrega_original_col_idx}")
            
            # Calcular disponibles
            productos_disponibles = {}
            for codigo, cantidad_original in productos_originales.items():
                cantidad_entregada = productos_entregados.get(codigo, 0)
                productos_disponibles[codigo] = cantidad_original - cantidad_entregada
                print(f"🔍 Debug calcular_cantidades: Resultado final - Código: {codigo}, Original: {cantidad_original}, Entregado: {cantidad_entregada}, Disponible: {productos_disponibles[codigo]}")
            
            print(f"🔍 Debug calcular_cantidades: ===== FIN CÁLCULO CANTIDADES DISPONIBLES =====")
            
            wb.close()
            return productos_disponibles
        except Exception as e:
            print(f"Error calculando cantidades disponibles: {e}")
            return {}
    
    # Función para validar stock suficiente
    def validar_stock_suficiente_comerciales(productos_validar):
        """Validar si hay stock suficiente para los productos"""
        productos_insuficientes = []
        stock_suficiente = True
        
        for prod in productos_validar:
            codigo = prod.get('codigo', '')
            cantidad_solicitada = float(prod.get('cantidad', 0))
            stock_disponible = get_stock_for_product_comerciales(codigo)
            
            if cantidad_solicitada > stock_disponible:
                stock_suficiente = False
                productos_insuficientes.append({
                    'codigo': codigo,
                    'nombre': prod.get('nombre', ''),
                    'solicitado': cantidad_solicitada,
                    'disponible': stock_disponible
                })
        
        return stock_suficiente, productos_insuficientes
    
    # Tabs para organizar
    tab_entrega_comercial, tab_entrega_cliente, tab_seguimiento = st.tabs([
        "📦 Entrega a Comercial", 
        "👤 Entrega Comercial a Cliente", 
        "📊 Seguimiento"
    ])
    
    # ===== TAB 1: ENTREGA A COMERCIAL =====
    with tab_entrega_comercial:
        st.markdown("### 📦 Entrega de Productos a Comercial/Vendedor")
        st.info("💡 Registra los productos que se le entregan al comercial. Estos se descontarán del inventario.")
        
        # Selector de productos
        if inv_productos_comerciales:
            productos_lista_com = [(f"{p['Codigo']} - {p['Nombre']}", p['Codigo']) for p in inv_productos_comerciales]
            productos_disponibles_com = [opt[1] for opt in productos_lista_com]
            productos_ya_seleccionados_com = [prod['codigo'] for prod in st.session_state["comerciales_productos_detalles"]]
            productos_para_agregar_com = [p for p in productos_disponibles_com if p not in productos_ya_seleccionados_com]
            
            col_add_com1, col_add_com2 = st.columns([3, 1])
            with col_add_com1:
                valor_actual_com = st.session_state.get("com_select_new_product", "")
                if valor_actual_com and valor_actual_com not in productos_para_agregar_com:
                    del st.session_state["com_select_new_product"]
                producto_nuevo_com = st.selectbox(
                    "Agregar Producto",
                    options=[""] + productos_para_agregar_com,
                    format_func=lambda x: "" if x == "" else next((opt[0] for opt in productos_lista_com if opt[1] == x), x),
                    key="com_select_new_product",
                    index=0
                )
            with col_add_com2:
                if st.button("➕ Agregar", key="com_add_product", use_container_width=True):
                    if producto_nuevo_com:
                        producto_info = next((p for p in inv_productos_comerciales if p['Codigo'] == producto_nuevo_com), None)
                        if producto_info:
                            nuevo_producto_com = {
                                'codigo': producto_info['Codigo'],
                                'nombre': producto_info['Nombre'],
                                'categoria': producto_info.get('Categoria', ''),
                                'cantidad': 1.0,
                                'precio_unit': float(producto_info.get('PrecioVenta', 0.0))
                            }
                            st.session_state["comerciales_productos_detalles"].append(nuevo_producto_com)
                            st.rerun()
            
            # Mostrar stock del producto seleccionado
            if producto_nuevo_com:
                stock_producto_seleccionado = get_stock_for_product_comerciales(producto_nuevo_com)
                producto_nombre_seleccionado = next((p['Nombre'] for p in inv_productos_comerciales if p['Codigo'] == producto_nuevo_com), "")
                if stock_producto_seleccionado >= 0:
                    st.info(f"📦 **Stock disponible:** {stock_producto_seleccionado:.0f} unidades - {producto_nombre_seleccionado}")
            
            # Mostrar tabla de productos
            if st.session_state["comerciales_productos_detalles"]:
                st.markdown("---")
                
                # Encabezados
                col_h_com1, col_h_com2, col_h_com3, col_h_com4, col_h_com5 = st.columns([1, 3, 1, 1, 1])
                with col_h_com1:
                    st.markdown("**Código**")
                with col_h_com2:
                    st.markdown("**Producto**")
                with col_h_com3:
                    st.markdown("**Cantidad**")
                with col_h_com4:
                    st.markdown("**Stock**")
                with col_h_com5:
                    st.markdown("**Acción**")
                
                # Filas
                productos_com_a_eliminar = []
                
                for idx, prod in enumerate(st.session_state["comerciales_productos_detalles"]):
                    col_com1, col_com2, col_com3, col_com4, col_com5 = st.columns([1, 3, 1, 1, 1])
                    
                    with col_com1:
                        st.write(prod['codigo'])
                    with col_com2:
                        st.write(prod['nombre'])
                    with col_com3:
                        cantidad_com = st.number_input(
                            "",
                            min_value=0.0,
                            step=1.0,
                            value=float(prod['cantidad']),
                            key=f"com_qty_{idx}",
                            label_visibility="collapsed"
                        )
                        prod['cantidad'] = cantidad_com
                    with col_com4:
                        stock_disponible_com = get_stock_for_product_comerciales(prod['codigo'])
                        if stock_disponible_com >= cantidad_com:
                            st.write(f"✅ {stock_disponible_com:.0f}")
                        elif stock_disponible_com > 0:
                            st.write(f"⚠️ {stock_disponible_com:.0f}")
                        else:
                            st.write(f"❌ {stock_disponible_com:.0f}")
                    with col_com5:
                        if st.button("🗑️", key=f"com_del_{idx}", help="Eliminar"):
                            productos_com_a_eliminar.append(idx)
                
                # Eliminar productos marcados
                if productos_com_a_eliminar:
                    indices_a_eliminar = sorted(set(productos_com_a_eliminar), reverse=True)
                    for idx in indices_a_eliminar:
                        if 0 <= idx < len(st.session_state["comerciales_productos_detalles"]):
                            st.session_state["comerciales_productos_detalles"].pop(idx)
                    st.rerun()
        
        st.markdown("---")
        st.markdown("### 📋 Información de Entrega")
        
        with st.form("form_entrega_comercial", clear_on_submit=True):
            # Selección de comercial/vendedor
            st.markdown("#### 👤 Comercial/Vendedor")
            if agentes:
                comercial_vendedor = st.selectbox("Comercial/Vendedor *", [""] + agentes, key="com_comercial")
            else:
                st.warning("⚠️ No hay comerciales/vendedores disponibles")
                comercial_vendedor = ""
            
            col_fecha_com1, col_fecha_com2 = st.columns(2)
            fecha = col_fecha_com1.date_input("Fecha *", key="com_fecha")
            
            observaciones = st.text_area("Observaciones", key="com_obs", height=100)
            
            submit = st.form_submit_button("✅ Guardar Entrega a Comercial", use_container_width=True)
            
            if submit:
                if not comercial_vendedor:
                    st.error("❌ Selecciona un comercial/vendedor")
                elif not st.session_state["comerciales_productos_detalles"]:
                    st.error("❌ Agrega al menos un producto")
                else:
                    # Obtener productos
                    productos_text = "\n".join([
                        f"{int(prod['cantidad'])} {prod['codigo']} - {prod['nombre']}"
                        for prod in st.session_state["comerciales_productos_detalles"]
                        if prod['cantidad'] > 0
                    ])
                    
                    # Validar stock antes de guardar
                    productos_validar_com = [prod for prod in st.session_state["comerciales_productos_detalles"] if prod.get('cantidad', 0) > 0]
                    
                    stock_suficiente_com, productos_insuficientes_com = validar_stock_suficiente_comerciales(productos_validar_com)
                    
                    if not stock_suficiente_com:
                        mensaje_error_com = "❌ **Stock insuficiente para los siguientes productos:**\n\n"
                        for p in productos_insuficientes_com:
                            mensaje_error_com += f"- **{p['codigo']} - {p['nombre']}**: Solicitado: {int(p['solicitado'])}, Disponible: {int(p['disponible'])}\n"
                        st.error(mensaje_error_com)
                    else:
                        # Guardar registro
                        next_id_val = next_id("COMERCIALES")
                        values = [
                            next_id_val, 
                            comercial_vendedor, 
                            "ENTREGA A COMERCIAL", 
                            fecha.strftime("%Y-%m-%d"), 
                            productos_text, 
                            "",  # Cliente Destino vacío
                            "ENTREGADO", 
                            observaciones,
                            "",  # Productos Entregados vacío
                            ""   # Fotos vacío
                        ]
                        
                        if append_row("COMERCIALES", values):
                            # Descontar stock del inventario general
                            for prod in productos_validar_com:
                                mov = {
                                    "Fecha": fecha.strftime("%Y-%m-%d"),
                                    "Tipo": "SALIDA",
                                    "Codigo": prod['codigo'],
                                    "Cantidad": float(prod['cantidad']),
                                    "CostoUnit": 0.0,
                                    "Proveedor": "",
                                    "Proceso": "Entrega a Comercial",
                                    "Nota": f"Entrega a Comercial: {comercial_vendedor}"
                                }
                                append_movement_global(mov)
                                
                                # Agregar stock al inventario del vendedor
                                agregar_stock_vendedor(comercial_vendedor, prod['codigo'], float(prod['cantidad']))
                            
                            st.success("✅ **Entrega a comercial guardada exitosamente. Stock descontado del inventario general y agregado al inventario del vendedor.**")
                            # Limpiar productos
                            st.session_state["comerciales_productos_detalles"] = []
                            leer_entregas_comerciales.clear()
                            st.rerun()
                        else:
                            st.error("❌ No se pudo guardar la entrega. Revisa los mensajes de error arriba.")
    
    # ===== TAB 2: ENTREGA COMERCIAL A CLIENTE =====
    with tab_entrega_cliente:
        st.markdown("### 👤 Registro de Entrega del Comercial a Cliente")
        st.info("💡 Selecciona una entrega a comercial y registra qué productos se entregaron al cliente.")
        
        # Cargar entregas a comerciales
        entregas_comerciales_tab2 = leer_entregas_comerciales()
        
        # Filtrar solo las entregas a comerciales (sin cliente destino aún)
        entregas_disponibles = [
            e for e in entregas_comerciales_tab2 
            if e.get('Tipo Entrega') == 'ENTREGA A COMERCIAL' 
            and not e.get('Cliente Destino')
        ]
        
        if not entregas_disponibles:
            st.warning("⚠️ No hay entregas a comerciales disponibles. Primero registra una entrega en el tab 'Entrega a Comercial'.")
        else:
            # Selector de entrega a comercial
            opciones_entregas = [
                f"ID #{e['ID']} - {e.get('Comercial/Vendedor', 'N/A')} - {e.get('Fecha', 'N/A')}"
                for e in entregas_disponibles
            ]
            
            entrega_seleccionada_idx = st.selectbox(
                "Selecciona la Entrega a Comercial *",
                options=range(len(opciones_entregas)),
                format_func=lambda x: opciones_entregas[x] if x < len(opciones_entregas) else "",
                key="entrega_comercial_seleccionada"
            )
            
            if entrega_seleccionada_idx is not None and entrega_seleccionada_idx < len(entregas_disponibles):
                entrega_seleccionada = entregas_disponibles[entrega_seleccionada_idx]
                entrega_id_original = entrega_seleccionada['ID']
                
                # Limpiar productos si cambió la entrega seleccionada
                if "entrega_comercial_anterior" not in st.session_state:
                    st.session_state["entrega_comercial_anterior"] = entrega_id_original
                
                if st.session_state["entrega_comercial_anterior"] != entrega_id_original:
                    st.session_state["comercial_cliente_productos"] = []
                    st.session_state["entrega_comercial_anterior"] = entrega_id_original
                
                st.markdown("---")
                st.markdown(f"#### 📦 Entrega Seleccionada: ID #{entrega_id_original}")
                
                col_info1, col_info2 = st.columns(2)
                with col_info1:
                    st.write(f"**Comercial/Vendedor:** {entrega_seleccionada.get('Comercial/Vendedor', 'N/A')}")
                    st.write(f"**Fecha:** {entrega_seleccionada.get('Fecha', 'N/A')}")
                with col_info2:
                    st.write(f"**Estado:** {entrega_seleccionada.get('Estado', 'N/A')}")
                    if entrega_seleccionada.get('Observaciones'):
                        st.write(f"**Observaciones:** {entrega_seleccionada.get('Observaciones')}")
                
                # Mostrar productos originales
                productos_originales_str = entrega_seleccionada.get('Productos', '')
                if productos_originales_str:
                    st.markdown("**Productos de la Entrega Original:**")
                    for prod_line in productos_originales_str.split('\n'):
                        if prod_line.strip():
                            st.write(f"- {prod_line.strip()}")
                
                # Calcular productos disponibles (siempre recalcular para tener datos actualizados)
                # Limpiar caché antes de calcular para asegurar datos actualizados
                leer_entregas_comerciales.clear()
                
                productos_disponibles_dict = calcular_cantidades_disponibles(entrega_id_original)
                
                # Parsear productos originales para mostrar información completa
                productos_info = {}
                for prod_line in productos_originales_str.split('\n'):
                    if prod_line.strip():
                        match = re.match(r"^(\d+)\s+(.+)$", prod_line.strip())
                        if match:
                            cantidad = int(match.group(1))
                            codigo_nombre = match.group(2).strip()
                            codigo_match = re.match(r"^([A-Z0-9_]+)\s*-\s*", codigo_nombre)
                            codigo = codigo_match.group(1) if codigo_match else ""
                            nombre = codigo_nombre.replace(f"{codigo} - ", "").strip() if codigo_match else codigo_nombre
                            if codigo:
                                disponible_calculado = productos_disponibles_dict.get(codigo, cantidad)
                                # Asegurar que disponible sea un entero y no negativo
                                disponible_final = max(0, int(disponible_calculado)) if disponible_calculado else cantidad
                                
                                productos_info[codigo] = {
                                    'cantidad_original': cantidad,
                                    'nombre': nombre,
                                    'disponible': disponible_final
                                }
                
                # Mostrar información de cantidades disponibles para debug
                if productos_info:
                    st.markdown("**📊 Cantidades Disponibles:**")
                    for codigo, info in productos_info.items():
                        cantidad_entregada_total = info['cantidad_original'] - info['disponible']
                        st.caption(f"• **{codigo} - {info['nombre']}**: {info['disponible']} disponibles (de {info['cantidad_original']} originales, {cantidad_entregada_total} ya entregados)")
                    
                    # Mostrar advertencia si no se encontraron entregas relacionadas
                    todas_originales = all(info['disponible'] == info['cantidad_original'] for info in productos_info.values())
                    if todas_originales and productos_disponibles_dict:
                        st.warning("⚠️ **Nota:** No se encontraron entregas a clientes registradas para esta entrega. Si ya registraste entregas, verifica que el ID_Entrega_Original se haya guardado correctamente en Excel.")
                        
                        # Botón para intentar corregir entregas sin ID_Entrega_Original
                        if st.button("🔧 Intentar Corregir Entregas Sin ID", key=f"corregir_entregas_{entrega_id_original}", help="Busca entregas a clientes del mismo vendedor y fecha y les asigna el ID_Entrega_Original correcto"):
                            with st.spinner("Buscando y corrigiendo entregas..."):
                                try:
                                    wb = safe_load_workbook(EXCEL_PATH)
                                    if "COMERCIALES" in wb.sheetnames:
                                        ws = wb["COMERCIALES"]
                                        headers = [cell.value for cell in ws[1]]
                                        
                                        # Buscar índice de columnas
                                        id_col_idx = None
                                        vendedor_col_idx = None
                                        tipo_col_idx = None
                                        fecha_col_idx = None
                                        id_original_col_idx = None
                                        
                                        for idx, header in enumerate(headers):
                                            if header == "ID":
                                                id_col_idx = idx
                                            elif header == "Comercial/Vendedor":
                                                vendedor_col_idx = idx
                                            elif header == "Tipo Entrega":
                                                tipo_col_idx = idx
                                            elif header == "Fecha":
                                                fecha_col_idx = idx
                                            elif header == "ID_Entrega_Original":
                                                id_original_col_idx = idx
                                        
                                        # Obtener información de la entrega original
                                        vendedor_original = entrega_seleccionada.get('Comercial/Vendedor', '')
                                        fecha_original = entrega_seleccionada.get('Fecha', '')
                                        
                                        correcciones = 0
                                        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                                            if len(row) > max(id_col_idx or 0, tipo_col_idx or 0, vendedor_col_idx or 0):
                                                id_fila = str(row[id_col_idx].value).strip() if id_col_idx is not None and row[id_col_idx].value else ""
                                                tipo_fila = str(row[tipo_col_idx].value).strip() if tipo_col_idx is not None and row[tipo_col_idx].value else ""
                                                vendedor_fila = str(row[vendedor_col_idx].value).strip() if vendedor_col_idx is not None and row[vendedor_col_idx].value else ""
                                                
                                                # Si es una entrega a cliente del mismo vendedor y no tiene ID_Entrega_Original
                                                if tipo_fila == "ENTREGA COMERCIAL A CLIENTE" and vendedor_fila == vendedor_original:
                                                    if id_original_col_idx is not None and len(row) > id_original_col_idx:
                                                        id_original_actual = row[id_original_col_idx].value
                                                        if not id_original_actual or str(id_original_actual).strip() == "":
                                                            # Asignar el ID_Entrega_Original
                                                            ws.cell(row=row_idx, column=id_original_col_idx + 1).value = str(entrega_id_original)
                                                            correcciones += 1
                                                            print(f"✅ Corregida fila {row_idx}: Asignado ID_Entrega_Original={entrega_id_original}")
                                        
                                        if correcciones > 0:
                                            if safe_save_workbook(wb, EXCEL_PATH):
                                                st.success(f"✅ Se corrigieron {correcciones} entrega(s). Recarga la página para ver los cambios.")
                                                wb.close()
                                                st.rerun()
                                            else:
                                                st.error("❌ Error al guardar las correcciones.")
                                        else:
                                            st.info("ℹ️ No se encontraron entregas para corregir.")
                                        
                                        wb.close()
                                except Exception as e:
                                    st.error(f"❌ Error al corregir entregas: {e}")
                                    print(f"Error corrigiendo entregas: {e}")
                
                st.markdown("---")
                
                # Inicializar productos para entrega al cliente
                if "comercial_cliente_productos" not in st.session_state:
                    st.session_state["comercial_cliente_productos"] = []
                
                # Selector de productos para entregar
                if productos_info:
                    st.markdown("#### 📋 Productos a Entregar al Cliente")
                    
                    productos_disponibles_lista = [
                        (codigo, info) 
                        for codigo, info in productos_info.items() 
                        if info['disponible'] > 0
                    ]
                    
                    if not productos_disponibles_lista:
                        st.warning("⚠️ No hay productos disponibles para entregar. Todos los productos de esta entrega ya fueron entregados a clientes.")
                    else:
                        # Agregar producto
                        col_add1, col_add2 = st.columns([3, 1])
                        with col_add1:
                            productos_ya_seleccionados = [p['codigo'] for p in st.session_state["comercial_cliente_productos"]]
                            productos_para_agregar = [
                                (codigo, info) 
                                for codigo, info in productos_disponibles_lista 
                                if codigo not in productos_ya_seleccionados
                            ]
                            
                            if productos_para_agregar:
                                producto_seleccionado = st.selectbox(
                                    "Agregar Producto",
                                    options=[""] + [codigo for codigo, _ in productos_para_agregar],
                                    format_func=lambda x: "" if x == "" else f"{x} - {productos_info[x]['nombre']} (Disponible: {productos_info[x]['disponible']})",
                                    key="comercial_cliente_select_producto"
                                )
                            else:
                                producto_seleccionado = ""
                                st.info("✅ Todos los productos disponibles ya están agregados.")
                        
                        with col_add2:
                            if st.button("➕ Agregar", key="comercial_cliente_add", use_container_width=True):
                                if producto_seleccionado and producto_seleccionado in productos_info:
                                    info = productos_info[producto_seleccionado]
                                    
                                    # Calcular cuánto ya está seleccionado de este producto en la lista actual
                                    cantidad_ya_seleccionada = sum(
                                        p.get('cantidad', 0) 
                                        for p in st.session_state["comercial_cliente_productos"] 
                                        if p['codigo'] == producto_seleccionado
                                    )
                                    
                                    # El disponible real es: disponible total - lo ya seleccionado
                                    disponible_real = max(0, int(info['disponible']) - cantidad_ya_seleccionada)
                                    
                                    if disponible_real > 0:
                                        nuevo_producto = {
                                            'codigo': producto_seleccionado,
                                            'nombre': info['nombre'],
                                            'cantidad_max': int(info['disponible']),  # Máximo total disponible
                                            'cantidad': disponible_real  # Iniciar con lo que realmente está disponible
                                        }
                                        st.session_state["comercial_cliente_productos"].append(nuevo_producto)
                                        st.rerun()
                                    else:
                                        st.warning(f"⚠️ No hay más cantidad disponible de este producto. Ya se seleccionaron {cantidad_ya_seleccionada} de {info['disponible']} disponibles.")
                        
                        # Mostrar productos seleccionados
                        if st.session_state["comercial_cliente_productos"]:
                            st.markdown("---")
                            col_h1, col_h2, col_h3, col_h4 = st.columns([2, 3, 2, 1])
                            with col_h1:
                                st.markdown("**Código**")
                            with col_h2:
                                st.markdown("**Producto**")
                            with col_h3:
                                st.markdown("**Cantidad**")
                            with col_h4:
                                st.markdown("**Acción**")
                            
                            productos_a_eliminar = []
                            for idx, prod in enumerate(st.session_state["comercial_cliente_productos"]):
                                # Actualizar cantidad_max con el valor más reciente de productos_info
                                codigo_prod = prod['codigo']
                                if codigo_prod in productos_info:
                                    # Obtener el disponible actualizado desde productos_info
                                    cantidad_disponible_total = productos_info[codigo_prod]['disponible']
                                    
                                    # Calcular cuánto ya está seleccionado en otras filas del mismo producto
                                    cantidad_ya_seleccionada_otras_filas = sum(
                                        p.get('cantidad', 0) 
                                        for i, p in enumerate(st.session_state["comercial_cliente_productos"]) 
                                        if p['codigo'] == codigo_prod and i != idx
                                    )
                                    
                                    # El máximo para esta fila es: disponible total - lo ya seleccionado en otras filas
                                    cantidad_max_real = max(0, cantidad_disponible_total - cantidad_ya_seleccionada_otras_filas)
                                    prod['cantidad_max'] = cantidad_max_real
                                else:
                                    cantidad_max_real = prod.get('cantidad_max', 0)
                                
                                col_p1, col_p2, col_p3, col_p4 = st.columns([2, 3, 2, 1])
                                
                                with col_p1:
                                    st.write(prod['codigo'])
                                with col_p2:
                                    st.write(prod['nombre'])
                                    st.caption(f"Máx: {cantidad_max_real}")
                                with col_p3:
                                    cantidad_entregar = st.number_input(
                                        "",
                                        min_value=0,
                                        max_value=int(cantidad_max_real),
                                        value=int(prod['cantidad']),
                                        step=1,
                                        key=f"comercial_cliente_qty_{idx}",
                                        label_visibility="collapsed"
                                    )
                                    prod['cantidad'] = cantidad_entregar
                                with col_p4:
                                    if st.button("🗑️", key=f"comercial_cliente_del_{idx}", help="Eliminar"):
                                        productos_a_eliminar.append(idx)
                            
                            # Eliminar productos marcados
                            if productos_a_eliminar:
                                indices_a_eliminar = sorted(set(productos_a_eliminar), reverse=True)
                                for idx in indices_a_eliminar:
                                    if 0 <= idx < len(st.session_state["comercial_cliente_productos"]):
                                        st.session_state["comercial_cliente_productos"].pop(idx)
                                st.rerun()
                            
                            st.markdown("---")
                            
                            # Formulario para completar la entrega
                            with st.form("form_entrega_comercial_cliente", clear_on_submit=True):
                                st.markdown("#### 📋 Información de Entrega al Cliente")
                                
                                # Selección de cliente
                                cliente_destino = form_cliente_section(clientes, "comercial_cliente", con_buscador=True)
                                
                                col_fecha1, col_fecha2 = st.columns(2)
                                fecha_entrega = col_fecha1.date_input("Fecha de Entrega *", key="comercial_cliente_fecha")
                                
                                observaciones_entrega = st.text_area("Observaciones", key="comercial_cliente_obs", height=100)
                                
                                # Subir fotos
                                fotos_entrega = st.file_uploader(
                                    "📸 Fotos de la Entrega (opcional)",
                                    type=['png', 'jpg', 'jpeg'],
                                    accept_multiple_files=True,
                                    key="comercial_cliente_fotos"
                                )
                                
                                submit_entrega = st.form_submit_button("✅ Guardar Entrega al Cliente", use_container_width=True)
                                
                                if submit_entrega:
                                    # Validaciones
                                    if not cliente_destino:
                                        st.error("❌ Selecciona un cliente destino")
                                    elif not st.session_state["comercial_cliente_productos"]:
                                        st.error("❌ Agrega al menos un producto para entregar")
                                    else:
                                        # Filtrar productos con cantidad > 0
                                        productos_entregar = [
                                            p for p in st.session_state["comercial_cliente_productos"] 
                                            if p.get('cantidad', 0) > 0
                                        ]
                                        
                                        if not productos_entregar:
                                            st.error("❌ Debes especificar al menos una cantidad mayor a 0 para algún producto")
                                        else:
                                            # Validar que las cantidades no excedan lo disponible
                                            productos_excedidos = []
                                            for prod in productos_entregar:
                                                disponible = productos_info[prod['codigo']]['disponible']
                                                if prod['cantidad'] > disponible:
                                                    productos_excedidos.append({
                                                        'codigo': prod['codigo'],
                                                        'nombre': prod['nombre'],
                                                        'solicitado': prod['cantidad'],
                                                        'disponible': disponible
                                                    })
                                            
                                            if productos_excedidos:
                                                mensaje_error = "❌ **Las siguientes cantidades exceden lo disponible:**\n\n"
                                                for p in productos_excedidos:
                                                    mensaje_error += f"- **{p['codigo']} - {p['nombre']}**: Solicitado: {p['solicitado']}, Disponible: {p['disponible']}\n"
                                                st.error(mensaje_error)
                                            else:
                                                # Validar stock del vendedor antes de guardar
                                                vendedor_nombre = entrega_seleccionada.get('Comercial/Vendedor', '')
                                                stock_vendedor_ok, productos_insuficientes_vendedor = validar_stock_vendedor_suficiente(vendedor_nombre, productos_entregar)
                                                
                                                if not stock_vendedor_ok:
                                                    mensaje_error_vendedor = "❌ **El vendedor no tiene stock suficiente para los siguientes productos:**\n\n"
                                                    for p in productos_insuficientes_vendedor:
                                                        mensaje_error_vendedor += f"- **{p['codigo']} - {p['nombre']}**: Solicitado: {int(p['solicitado'])}, Disponible en inventario del vendedor: {int(p['disponible'])}\n"
                                                    st.error(mensaje_error_vendedor)
                                                else:
                                                    # Preparar texto de productos entregados
                                                    productos_entregados_text = "\n".join([
                                                        f"{int(prod['cantidad'])} {prod['codigo']} - {prod['nombre']}"
                                                        for prod in productos_entregar
                                                    ])
                                                    
                                                    # Guardar fotos si hay
                                                    rutas_fotos = ""
                                                    if fotos_entrega:
                                                        rutas_fotos_lista = guardar_fotos_comercial(entrega_id_original, fotos_entrega)
                                                        rutas_fotos = "|".join(rutas_fotos_lista)
                                                    
                                                    # Guardar registro
                                                    next_id_val = next_id("COMERCIALES")
                                                    
                                                    # Extraer nombre del cliente del selectbox
                                                    cliente_nombre = ""
                                                    if cliente_destino:
                                                        match = re.match(r"^(\d+)\s+-\s+(.+?)\s+\(", cliente_destino)
                                                        if match:
                                                            cliente_nombre = match.group(2).strip()
                                                        else:
                                                            cliente_nombre = cliente_destino.split(" - ")[-1].split(" (")[0] if " - " in cliente_destino else cliente_destino
                                                    
                                                    # Crear values según el orden de HEADERS["COMERCIALES"]
                                                    # HEADERS: ["ID", "Comercial/Vendedor", "Tipo Entrega", "Fecha", "Productos", "Cliente Destino", "Estado", "Observaciones", "Productos Entregados", "Fotos", "ID_Entrega_Original"]
                                                    values = [
                                                        next_id_val,  # ID
                                                        entrega_seleccionada.get('Comercial/Vendedor', ''),  # Comercial/Vendedor
                                                        "ENTREGA COMERCIAL A CLIENTE",  # Tipo Entrega
                                                        fecha_entrega.strftime("%Y-%m-%d"),  # Fecha
                                                        entrega_seleccionada.get('Productos', ''),  # Productos (originales)
                                                        cliente_nombre,  # Cliente Destino
                                                        "ENTREGADO",  # Estado
                                                        observaciones_entrega,  # Observaciones
                                                        productos_entregados_text,  # Productos Entregados
                                                        rutas_fotos,  # Fotos
                                                        str(entrega_id_original)  # ID_Entrega_Original (posición 10, índice 10)
                                                    ]
                                                    
                                                    # Debug: mostrar valores antes de guardar
                                                    print(f"🔍 Debug guardar: Guardando entrega con ID_Entrega_Original={str(entrega_id_original)}")
                                                    print(f"🔍 Debug guardar: Tipo de ID: {type(entrega_id_original)}")
                                                    print(f"🔍 Debug guardar: Valores a guardar (antes de append_row): {values}")
                                                    print(f"🔍 Debug guardar: Longitud de values: {len(values)}")
                                                    
                                                    # Verificar que el ID_Entrega_Original esté en la posición correcta (índice 10)
                                                    if len(values) > 10:
                                                        print(f"🔍 Debug guardar: ID_Entrega_Original en values[10]: '{values[10]}' (tipo: {type(values[10])})")
                                                    else:
                                                        print(f"⚠️ Debug guardar: values tiene solo {len(values)} elementos, necesita al menos 11 para ID_Entrega_Original")
                                                    
                                                    if append_row("COMERCIALES", values):
                                                        # Verificar que se guardó correctamente
                                                        print(f"✅ Debug guardar: Entrega guardada. Verificando que ID_Entrega_Original se guardó correctamente...")
                                                        try:
                                                            wb_check = safe_load_workbook(EXCEL_PATH)
                                                            if "COMERCIALES" in wb_check.sheetnames:
                                                                ws_check = wb_check["COMERCIALES"]
                                                                headers_check = [cell.value for cell in ws_check[1]]
                                                                id_col_idx = None
                                                                for idx, h in enumerate(headers_check):
                                                                    if h == "ID_Entrega_Original":
                                                                        id_col_idx = idx
                                                                        break
                                                                
                                                                if id_col_idx is not None:
                                                                    # Leer la última fila
                                                                    last_row = list(ws_check.iter_rows(min_row=ws_check.max_row, values_only=True))[0]
                                                                    if len(last_row) > id_col_idx:
                                                                        id_guardado = last_row[id_col_idx]
                                                                        print(f"🔍 Debug guardar: ID_Entrega_Original guardado en Excel: '{id_guardado}' (tipo: {type(id_guardado)})")
                                                                        if str(id_guardado).strip() != str(entrega_id_original).strip():
                                                                            print(f"⚠️ Debug guardar: ERROR - ID guardado '{id_guardado}' no coincide con ID esperado '{entrega_id_original}'")
                                                                        else:
                                                                            print(f"✅ Debug guardar: ID_Entrega_Original se guardó correctamente")
                                                                    else:
                                                                        print(f"⚠️ Debug guardar: La última fila no tiene suficientes columnas")
                                                                else:
                                                                    print(f"⚠️ Debug guardar: No se encontró la columna ID_Entrega_Original en headers")
                                                            wb_check.close()
                                                        except Exception as e:
                                                            print(f"⚠️ Debug guardar: Error verificando guardado: {e}")
                                                        # Descontar stock del inventario del vendedor
                                                        for prod in productos_entregar:
                                                            descontar_stock_vendedor(vendedor_nombre, prod['codigo'], float(prod['cantidad']))
                                                        
                                                        # Limpiar productos y cachés ANTES de mostrar el mensaje
                                                        st.session_state["comercial_cliente_productos"] = []
                                                        leer_entregas_comerciales.clear()
                                                        
                                                        # Esperar un momento para que el archivo se guarde completamente
                                                        import time
                                                        time.sleep(1.0)  # Aumentado a 1 segundo para asegurar que el archivo se guarde
                                                        
                                                        # Forzar recálculo de cantidades disponibles limpiando el estado
                                                        if "entrega_comercial_anterior" in st.session_state:
                                                            del st.session_state["entrega_comercial_anterior"]
                                                        
                                                        # Calcular nuevas cantidades disponibles para mostrar (forzar recálculo)
                                                        nuevas_cantidades = calcular_cantidades_disponibles(entrega_id_original)
                                                        
                                                        # Parsear productos originales para obtener nombres
                                                        productos_info_mensaje = {}
                                                        productos_originales_str = entrega_seleccionada.get('Productos', '')
                                                        for prod_line in productos_originales_str.split('\n'):
                                                            if prod_line.strip():
                                                                match = re.match(r"^(\d+)\s+(.+)$", prod_line.strip())
                                                                if match:
                                                                    cantidad = int(match.group(1))
                                                                    codigo_nombre = match.group(2).strip()
                                                                    codigo_match = re.match(r"^([A-Z0-9_]+)\s*-\s*", codigo_nombre)
                                                                    codigo = codigo_match.group(1) if codigo_match else ""
                                                                    if not codigo:
                                                                        codigo_match2 = re.match(r"^([A-Z0-9_]+)", codigo_nombre)
                                                                        codigo = codigo_match2.group(1) if codigo_match2 else ""
                                                                    nombre = codigo_nombre.replace(f"{codigo} - ", "").strip() if codigo_match else codigo_nombre
                                                                    if codigo:
                                                                        productos_info_mensaje[codigo] = {
                                                                            'nombre': nombre,
                                                                            'cantidad_original': cantidad
                                                                        }
                                                        
                                                        # Crear mensaje con información de cantidades actualizadas
                                                        mensaje_exito = "✅ **Entrega al cliente registrada exitosamente**\n\n"
                                                        mensaje_exito += "📦 **Cantidades disponibles actualizadas:**\n"
                                                        mensaje_exito += f"📉 **Stock del vendedor descontado**\n"
                                                        for codigo, cantidad_disponible in nuevas_cantidades.items():
                                                            prod_info = productos_info_mensaje.get(codigo, {})
                                                            nombre_prod = prod_info.get('nombre', codigo)
                                                            cantidad_original = prod_info.get('cantidad_original', 0)
                                                            mensaje_exito += f"- **{codigo} - {nombre_prod}**: {cantidad_disponible} de {cantidad_original} disponibles\n"
                                                        
                                                        st.success(mensaje_exito)
                                                        st.rerun()
                                                    else:
                                                        st.error("❌ No se pudo guardar la entrega. Revisa los mensajes de error arriba.")
                else:
                    st.warning("⚠️ No se pudieron cargar los productos de esta entrega.")
    
    # ===== TAB 3: SEGUIMIENTO =====
    with tab_seguimiento:
        st.markdown("### 📊 Seguimiento de Entregas a Comerciales")
        
        # Cargar todas las entregas
        todas_entregas_seguimiento = leer_entregas_comerciales()
        
        # Separar entregas originales y entregas a clientes
        entregas_originales_seg = [e for e in todas_entregas_seguimiento if e.get('Tipo Entrega') == 'ENTREGA A COMERCIAL']
        entregas_a_clientes_seg = [e for e in todas_entregas_seguimiento if e.get('Tipo Entrega') == 'ENTREGA COMERCIAL A CLIENTE']
        
        # Agrupar entregas a clientes por vendedor y luego por cliente
        entregas_por_vendedor = {}
        
        for entrega_cliente in entregas_a_clientes_seg:
            vendedor = entrega_cliente.get('Comercial/Vendedor', 'Sin Vendedor')
            cliente = entrega_cliente.get('Cliente Destino', 'Sin Cliente')
            id_original = entrega_cliente.get('ID_Entrega_Original', '')
            
            if vendedor not in entregas_por_vendedor:
                entregas_por_vendedor[vendedor] = {}
            
            if cliente not in entregas_por_vendedor[vendedor]:
                entregas_por_vendedor[vendedor][cliente] = []
            
            entregas_por_vendedor[vendedor][cliente].append(entrega_cliente)
        
        # Métricas generales
        col_met1, col_met2, col_met3 = st.columns(3)
        col_met1.metric("Total Entregas a Comerciales", len(entregas_originales_seg))
        col_met2.metric("Total Entregas a Clientes", len(entregas_a_clientes_seg))
        col_met3.metric("Vendedores Activos", len(entregas_por_vendedor))
        
        st.markdown("---")
        
        # Filtros y descarga en Excel
        st.markdown("#### 🔍 Filtros y Exportación")
        col_filtro1, col_filtro2, col_filtro3, col_filtro4 = st.columns(4)
        
        # Obtener listas únicas para filtros
        vendedores_lista = sorted(list(set([e.get('Comercial/Vendedor', '') for e in entregas_a_clientes_seg if e.get('Comercial/Vendedor')])))
        clientes_lista = sorted(list(set([e.get('Cliente Destino', '') for e in entregas_a_clientes_seg if e.get('Cliente Destino')])))
        fechas_lista = sorted(list(set([e.get('Fecha', '') for e in entregas_a_clientes_seg if e.get('Fecha')])))
        
        # Filtros
        filtro_vendedor = col_filtro1.selectbox("Filtrar por Vendedor", [""] + vendedores_lista, key="filtro_vendedor_seg")
        filtro_cliente = col_filtro2.selectbox("Filtrar por Cliente", [""] + clientes_lista, key="filtro_cliente_seg")
        
        # Filtro de fecha (rango)
        if fechas_lista:
            try:
                fechas_datetime = [datetime.strptime(f, "%Y-%m-%d") for f in fechas_lista if f]
                fecha_min = min(fechas_datetime).date() if fechas_datetime else datetime.now().date()
                fecha_max = max(fechas_datetime).date() if fechas_datetime else datetime.now().date()
            except:
                fecha_min = datetime.now().date()
                fecha_max = datetime.now().date()
        else:
            fecha_min = datetime.now().date()
            fecha_max = datetime.now().date()
        
        rango_fechas = col_filtro3.date_input("Rango de Fechas", value=(fecha_min, fecha_max), key="filtro_fecha_seg")
        
        # Aplicar filtros
        entregas_filtradas = entregas_a_clientes_seg.copy()
        
        if filtro_vendedor:
            entregas_filtradas = [e for e in entregas_filtradas if e.get('Comercial/Vendedor', '') == filtro_vendedor]
        
        if filtro_cliente:
            entregas_filtradas = [e for e in entregas_filtradas if e.get('Cliente Destino', '') == filtro_cliente]
        
        if isinstance(rango_fechas, tuple) and len(rango_fechas) == 2:
            fecha_inicio, fecha_fin = rango_fechas
            entregas_filtradas = [
                e for e in entregas_filtradas 
                if e.get('Fecha', '') and 
                fecha_inicio <= datetime.strptime(e.get('Fecha', ''), "%Y-%m-%d").date() <= fecha_fin
            ]
        
        # Botón de descarga
        if entregas_filtradas:
            col_filtro4.markdown("<br>", unsafe_allow_html=True)  # Espaciado
            if st.button("📥 Descargar Excel", key="descargar_excel_seg", use_container_width=True):
                try:
                    # Crear DataFrame con los datos filtrados
                    datos_exportar = []
                    for entrega in entregas_filtradas:
                        # Parsear productos entregados
                        productos_entregados_str = entrega.get('Productos Entregados', '')
                        productos_lista = productos_entregados_str.split('\n') if productos_entregados_str else []
                        
                        if productos_lista:
                            for prod_line in productos_lista:
                                if prod_line.strip():
                                    # Parsear el formato: "cantidad codigo - nombre"
                                    cantidad = ""
                                    codigo = ""
                                    nombre = ""
                                    
                                    # Intentar parsear el formato estándar
                                    match = re.match(r"^(\d+)\s+([A-Z0-9_]+)\s*-\s*(.+)$", prod_line.strip())
                                    if match:
                                        cantidad = match.group(1)
                                        codigo = match.group(2)
                                        nombre = match.group(3).strip()
                                    else:
                                        # Si no coincide el formato, intentar otras variaciones
                                        partes = prod_line.strip().split(' ', 2)
                                        if len(partes) >= 2:
                                            cantidad = partes[0]
                                            resto = ' '.join(partes[1:])
                                            # Buscar código (letras y números hasta el guión)
                                            codigo_match = re.match(r"^([A-Z0-9_]+)\s*-\s*(.+)$", resto)
                                            if codigo_match:
                                                codigo = codigo_match.group(1)
                                                nombre = codigo_match.group(2).strip()
                                            else:
                                                # Si no hay guión, tomar todo como nombre
                                                nombre = resto
                                    
                                    datos_exportar.append({
                                        'ID Entrega': entrega.get('ID', ''),
                                        'Vendedor': entrega.get('Comercial/Vendedor', ''),
                                        'Cliente': entrega.get('Cliente Destino', ''),
                                        'Fecha': entrega.get('Fecha', ''),
                                        'Estado': entrega.get('Estado', ''),
                                        'Cantidad': cantidad,
                                        'Código Producto': codigo,
                                        'Nombre Producto': nombre,
                                        'Observaciones': entrega.get('Observaciones', ''),
                                        'ID Entrega Original': entrega.get('ID_Entrega_Original', '')
                                    })
                        else:
                            # Si no hay productos, agregar fila con datos generales
                            datos_exportar.append({
                                'ID Entrega': entrega.get('ID', ''),
                                'Vendedor': entrega.get('Comercial/Vendedor', ''),
                                'Cliente': entrega.get('Cliente Destino', ''),
                                'Fecha': entrega.get('Fecha', ''),
                                'Estado': entrega.get('Estado', ''),
                                'Cantidad': '',
                                'Código Producto': '',
                                'Nombre Producto': '',
                                'Observaciones': entrega.get('Observaciones', ''),
                                'ID Entrega Original': entrega.get('ID_Entrega_Original', '')
                            })
                    
                    if datos_exportar:
                        df_exportar = pd.DataFrame(datos_exportar)
                        
                        # Generar nombre de archivo con filtros aplicados
                        nombre_archivo_parts = ["Reporte_Entregas"]
                        if filtro_vendedor:
                            nombre_archivo_parts.append(f"Vendedor_{filtro_vendedor[:20]}")
                        if filtro_cliente:
                            nombre_archivo_parts.append(f"Cliente_{filtro_cliente[:20]}")
                        if isinstance(rango_fechas, tuple) and len(rango_fechas) == 2:
                            fecha_inicio, fecha_fin = rango_fechas
                            nombre_archivo_parts.append(f"{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}")
                        
                        nombre_archivo = "_".join(nombre_archivo_parts) + ".xlsx"
                        
                        # Crear Excel en memoria
                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df_exportar.to_excel(writer, sheet_name='Entregas', index=False)
                        
                        output.seek(0)
                        
                        st.download_button(
                            label="⬇️ Descargar Archivo Excel",
                            data=output.getvalue(),
                            file_name=nombre_archivo,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_excel_seg"
                        )
                        
                        st.success(f"✅ Archivo Excel generado con {len(datos_exportar)} registro(s)")
                    else:
                        st.warning("⚠️ No hay datos para exportar con los filtros seleccionados")
                except Exception as e:
                    st.error(f"❌ Error al generar Excel: {e}")
                    print(f"Error generando Excel: {e}")
        
        st.markdown(f"**Mostrando {len(entregas_filtradas)} de {len(entregas_a_clientes_seg)} entregas**")
        st.markdown("---")
        
        # Aplicar filtros también a la visualización
        entregas_por_vendedor_filtradas = {}
        for entrega_cliente in entregas_filtradas:
            vendedor = entrega_cliente.get('Comercial/Vendedor', 'Sin Vendedor')
            cliente = entrega_cliente.get('Cliente Destino', 'Sin Cliente')
            
            if vendedor not in entregas_por_vendedor_filtradas:
                entregas_por_vendedor_filtradas[vendedor] = {}
            
            if cliente not in entregas_por_vendedor_filtradas[vendedor]:
                entregas_por_vendedor_filtradas[vendedor][cliente] = []
            
            entregas_por_vendedor_filtradas[vendedor][cliente].append(entrega_cliente)
        
        if entregas_por_vendedor_filtradas:
            st.markdown("#### 👥 Entregas por Vendedor")
            
            for vendedor in sorted(entregas_por_vendedor_filtradas.keys()):
                entregas_vendedor = entregas_por_vendedor_filtradas[vendedor]
                total_clientes = len(entregas_vendedor)
                total_entregas = sum(len(entregas) for entregas in entregas_vendedor.values())
                
                with st.expander(f"📦 **{vendedor}** - {total_clientes} cliente(s) - {total_entregas} entrega(s)", expanded=False):
                    # Buscar la entrega original de este vendedor para mostrar productos totales
                    entrega_original_vendedor = None
                    for entrega_orig in entregas_originales_seg:
                        if entrega_orig.get('Comercial/Vendedor') == vendedor:
                            # Buscar la entrega original más reciente o la que tenga más productos
                            if not entrega_original_vendedor:
                                entrega_original_vendedor = entrega_orig
                            else:
                                # Comparar por fecha o cantidad de productos
                                productos_orig = entrega_orig.get('Productos', '')
                                productos_actual = entrega_original_vendedor.get('Productos', '')
                                if len(productos_orig) > len(productos_actual):
                                    entrega_original_vendedor = entrega_orig
                    
                    if entrega_original_vendedor:
                        st.markdown(f"**📦 Entrega Original:** ID #{entrega_original_vendedor['ID']} - Fecha: {entrega_original_vendedor.get('Fecha', 'N/A')}")
                        productos_orig_str = entrega_original_vendedor.get('Productos', '')
                        if productos_orig_str:
                            st.markdown("**Productos Totales:**")
                            for prod_line in productos_orig_str.split('\n'):
                                if prod_line.strip():
                                    st.write(f"- {prod_line.strip()}")
                        st.markdown("---")
                    
                    # Mostrar entregas por cliente
                    for cliente in sorted(entregas_vendedor.keys()):
                        entregas_cliente = entregas_vendedor[cliente]
                        
                        st.markdown(f"**👤 Cliente: {cliente}** ({len(entregas_cliente)} entrega(s))")
                        
                        for idx, entrega in enumerate(entregas_cliente, 1):
                            col_cli1, col_cli2 = st.columns([2, 1])
                            
                            with col_cli1:
                                st.write(f"**Entrega #{idx}:** ID #{entrega['ID']} - Fecha: {entrega.get('Fecha', 'N/A')}")
                                productos_entregados_str = entrega.get('Productos Entregados', '')
                                if productos_entregados_str:
                                    st.write("**Productos Entregados:**")
                                    for prod_line in productos_entregados_str.split('\n'):
                                        if prod_line.strip():
                                            st.write(f"  • {prod_line.strip()}")
                            
                            with col_cli2:
                                st.write(f"**Estado:** {entrega.get('Estado', 'N/A')}")
                                if entrega.get('Observaciones'):
                                    st.caption(f"Obs: {entrega.get('Observaciones')[:50]}...")
                            
                            # Mostrar fotos si hay
                            fotos_str = entrega.get('Fotos', '')
                            if fotos_str:
                                fotos_lista = [f.strip() for f in fotos_str.split('|') if f.strip()]
                                if fotos_lista:
                                    st.markdown("**📸 Fotos:**")
                                    num_cols_fotos = min(3, len(fotos_lista))
                                    cols_fotos = st.columns(num_cols_fotos)
                                    for idx_foto, ruta_foto in enumerate(fotos_lista[:num_cols_fotos]):
                                        with cols_fotos[idx_foto % num_cols_fotos]:
                                            try:
                                                if es_url(ruta_foto):
                                                    st.image(ruta_foto, use_container_width=True)
                                                elif Path(ruta_foto).exists():
                                                    st.image(str(ruta_foto), use_container_width=True)
                                                else:
                                                    st.caption("No se encontró la foto")
                                            except Exception:
                                                st.caption("Error al cargar foto")
                            
                            st.markdown("---")
        else:
            st.info("📋 No hay entregas a clientes registradas aún.")

# ===== SECCIÓN INVENTARIO =====
elif selected_menu == "📊 INVENTARIO":
    st.markdown("### 📊 INVENTARIO")
    
    # Utilidades de inventario avanzado (productos / movimientos / proveedores)
    INV_SHEET_PRODUCTS = "INVENTARIO_PRODUCTOS"
    INV_SHEET_MOVS = "INVENTARIO_MOVIMIENTOS"
    INV_SHEET_PROVS = "PROVEEDORES"
    INV_SHEET_COSTS = "COSTOS_PRODUCTO"
    INV_SHEET_LINKS = "PRODUCTO_PROVEEDOR"
    INV_SHEET_RECIPES = "RECETAS"
    INV_SHEET_PROD_ORDERS = "ORDENES_PROD"

    def ensure_inventory_sheets():
        try:
            ensure_workbook(EXCEL_PATH)
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_PRODUCTS not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_PRODUCTS)
                ws.append(["Codigo", "Nombre", "Categoria", "Unidad", "TipoUso", "PrecioVenta", "Activo", "Observaciones"])
            if INV_SHEET_MOVS not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_MOVS)
                ws.append(["Fecha", "Tipo", "Codigo", "Cantidad", "CostoUnit", "Proveedor", "Proceso", "Nota"])
            if INV_SHEET_PROVS not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_PROVS)
                ws.append(["ProveedorID", "Nombre", "Tipo", "Contacto", "Telefono", "Email", "Material", "Costo", "Observaciones"])
            if INV_SHEET_LINKS not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_LINKS)
                ws.append(["Codigo", "ProveedorID", "Rol", "CostoUnit", "LeadTimeDias", "Observaciones"])
            if INV_SHEET_COSTS not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_COSTS)
                ws.append(["Fecha", "Codigo", "CantidadProducida", "CostoTotal", "CostoUnitario", "DetalleJSON"]) 
            if INV_SHEET_RECIPES not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_RECIPES)
                ws.append(["ProductoTerminado", "InsumoCodigo", "CantidadPorUnidad", "Observaciones"])
            if INV_SHEET_PROD_ORDERS not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_PROD_ORDERS)
                ws.append(["Fecha", "Producto", "Cantidad", "Estado", "Nota"])
            safe_save_workbook(wb, EXCEL_PATH)
            try:
                wb.close()
            except:
                pass
        except Exception as e:
            st.error(f"Error preparando hojas de inventario: {e}")

    def load_recipes(producto_terminado=None):
        recetas = []
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_RECIPES in wb.sheetnames:
                ws = wb[INV_SHEET_RECIPES]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0] and row[1]:
                        rec = {
                            "ProductoTerminado": str(row[0]).strip(),
                            "InsumoCodigo": str(row[1]).strip(),
                            "CantidadPorUnidad": float(row[2]) if row[2] else 0.0,
                            "Observaciones": str(row[3]).strip() if len(row) > 3 and row[3] else "",
                        }
                        if not producto_terminado or rec["ProductoTerminado"] == producto_terminado:
                            recetas.append(rec)
            wb.close()
        except Exception:
            pass
        return recetas

    def upsert_recipe(rec):
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_RECIPES not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_RECIPES)
                ws.append(["ProductoTerminado", "InsumoCodigo", "CantidadPorUnidad", "Observaciones"])
            ws = wb[INV_SHEET_RECIPES]
            updated = False
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if str(row[0].value).strip() == rec["ProductoTerminado"] and str(row[1].value).strip() == rec["InsumoCodigo"]:
                    ws.cell(row=idx, column=3).value = rec["CantidadPorUnidad"]
                    ws.cell(row=idx, column=4).value = rec["Observaciones"]
                    updated = True
                    break
            if not updated:
                ws.append([rec["ProductoTerminado"], rec["InsumoCodigo"], rec["CantidadPorUnidad"], rec["Observaciones"]])
            if safe_save_workbook(wb, EXCEL_PATH):
                wb.close()
                return True
            wb.close()
        except Exception as e:
            st.error(f"Error guardando receta: {e}")
        return False

    def compute_consumption(producto_terminado, cantidad_producir):
        consumo = []
        for r in load_recipes(producto_terminado):
            total = float(r.get("CantidadPorUnidad", 0.0)) * float(cantidad_producir)
            if total > 0:
                consumo.append({"InsumoCodigo": r["InsumoCodigo"], "Cantidad": total})
        return consumo

    def append_production_order(po):
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_PROD_ORDERS not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_PROD_ORDERS)
                ws.append(["Fecha", "Producto", "Cantidad", "Estado", "Nota"])
            ws = wb[INV_SHEET_PROD_ORDERS]
            ws.append([po["Fecha"], po["Producto"], po["Cantidad"], po["Estado"], po["Nota"]])
            if safe_save_workbook(wb, EXCEL_PATH):
                wb.close()
                return True
            wb.close()
        except Exception as e:
            st.error(f"Error guardando orden de producción: {e}")
        return False

    def execute_production(producto_terminado, cantidad, nota=""):
        # Validación básica de receta
        receta = load_recipes(producto_terminado)
        if not receta:
            st.error("No hay receta para el producto terminado seleccionado.")
            return False
        # Validar stock de insumos
        stock_map_local = calc_stock_map()
        faltantes = []
        for c in compute_consumption(producto_terminado, cantidad):
            disp = stock_map_local.get(c["InsumoCodigo"], 0.0)
            if disp < c["Cantidad"]:
                faltantes.append({"Codigo": c["InsumoCodigo"], "Necesita": c["Cantidad"], "Disponible": disp})
        if faltantes:
            msg = "\n".join([f"- {f['Codigo']}: necesita {f['Necesita']}, disponible {f['Disponible']}" for f in faltantes])
            st.error("Stock insuficiente para producir:\n" + msg)
            return False
        # Registrar SALIDAS de insumos
        for c in compute_consumption(producto_terminado, cantidad):
            mov = {
                "Fecha": datetime.now().strftime("%Y-%m-%d"),
                "Tipo": "SALIDA",
                "Codigo": c["InsumoCodigo"],
                "Cantidad": c["Cantidad"],
                "CostoUnit": 0.0,
                "Proveedor": "",
                "Proceso": "Producción",
                "Nota": nota or f"Consumo para {producto_terminado} x {cantidad}"
            }
            if not append_movement(mov):
                return False
        # Registrar ENTRADA de producto terminado
        mov_fin = {
            "Fecha": datetime.now().strftime("%Y-%m-%d"),
            "Tipo": "ENTRADA",
            "Codigo": producto_terminado,
            "Cantidad": cantidad,
            "CostoUnit": 0.0,
            "Proveedor": "",
            "Proceso": "Producción",
            "Nota": nota or "Producción"
        }
        if not append_movement(mov_fin):
            return False
        # Registrar orden
        po = {
            "Fecha": datetime.now().strftime("%Y-%m-%d"),
            "Producto": producto_terminado,
            "Cantidad": cantidad,
            "Estado": "COMPLETADA",
            "Nota": nota or ""
        }
        append_production_order(po)
        return True

    def load_products():
        productos = []
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_PRODUCTS in wb.sheetnames:
                ws = wb[INV_SHEET_PRODUCTS]
                # Verificar estructura leyendo los headers
                headers = [cell.value for cell in ws[1]]
                has_precio_venta = "PrecioVenta" in headers
                is_old_format = "StockInicial" in headers
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        if is_old_format and len(row) >= 9:
                            # Estructura antigua: Codigo, Nombre, Categoria, PrecioVenta, Unidad, StockInicial, StockMinimo, Activo, Observaciones
                            productos.append({
                                "Codigo": str(row[0]).strip(),
                                "Nombre": str(row[1]).strip() if row[1] else "",
                                "Categoria": str(row[2]).strip() if row[2] else "",
                                "Unidad": str(row[4]).strip() if row[4] else "pc",
                                "TipoUso": "Alquiler",  # Valor por defecto para migración
                                "PrecioVenta": float(row[3]) if row[3] else 0.0,
                                "Activo": str(row[7]).strip() if row[7] else "SI",
                                "Observaciones": str(row[8]).strip() if len(row) > 8 and row[8] else "",
                            })
                        elif has_precio_venta:
                            # Estructura nueva: Codigo, Nombre, Categoria, Unidad, TipoUso, PrecioVenta, Activo, Observaciones
                            productos.append({
                                "Codigo": str(row[0]).strip(),
                                "Nombre": str(row[1]).strip() if row[1] else "",
                                "Categoria": str(row[2]).strip() if row[2] else "",
                                "Unidad": str(row[3]).strip() if row[3] else "pc",
                                "TipoUso": str(row[4]).strip() if len(row) > 4 and row[4] else "Alquiler",
                                "PrecioVenta": float(row[5]) if len(row) > 5 and row[5] else 0.0,
                                "Activo": str(row[6]).strip() if len(row) > 6 and row[6] else "SI",
                                "Observaciones": str(row[7]).strip() if len(row) > 7 and row[7] else "",
                            })
                        else:
                            # Estructura sin PrecioVenta: Codigo, Nombre, Categoria, Unidad, TipoUso, Activo, Observaciones
                            productos.append({
                                "Codigo": str(row[0]).strip(),
                                "Nombre": str(row[1]).strip() if row[1] else "",
                                "Categoria": str(row[2]).strip() if row[2] else "",
                                "Unidad": str(row[3]).strip() if row[3] else "pc",
                                "TipoUso": str(row[4]).strip() if len(row) > 4 and row[4] else "Alquiler",
                                "PrecioVenta": 0.0,  # Valor por defecto
                                "Activo": str(row[5]).strip() if len(row) > 5 and row[5] else "SI",
                                "Observaciones": str(row[6]).strip() if len(row) > 6 and row[6] else "",
                            })
            wb.close()
        except Exception:
            pass
        return productos

    def upsert_product(prod):
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_PRODUCTS not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_PRODUCTS)
                ws.append(["Codigo", "Nombre", "Categoria", "Unidad", "TipoUso", "PrecioVenta", "Activo", "Observaciones"])
            ws = wb[INV_SHEET_PRODUCTS]
            
            # Verificar si hay que migrar la estructura antigua
            headers = [cell.value for cell in ws[1]]
            if "StockInicial" in headers:
                # Migrar estructura antigua con StockInicial
                new_ws = wb.create_sheet(INV_SHEET_PRODUCTS + "_new")
                new_ws.append(["Codigo", "Nombre", "Categoria", "Unidad", "TipoUso", "PrecioVenta", "Activo", "Observaciones"])
                dropping = wb[INV_SHEET_PRODUCTS]
                for row in dropping.iter_rows(min_row=2):
                    if row[0].value:
                        new_ws.append([
                            row[0].value,  # Codigo
                            row[1].value if len(row) > 1 else "",  # Nombre
                            row[2].value if len(row) > 2 else "",  # Categoria
                            row[4].value if len(row) > 4 else "pc",  # Unidad (saltando PrecioVenta en posición 3)
                            "Alquiler",  # TipoUso por defecto
                            row[3].value if len(row) > 3 else 0.0,  # PrecioVenta de la posición original 3
                            row[7].value if len(row) > 7 else "SI",  # Activo
                            row[8].value if len(row) > 8 else ""  # Observaciones
                        ])
                wb.remove(dropping)
                new_ws.title = INV_SHEET_PRODUCTS
                ws = wb[INV_SHEET_PRODUCTS]
            elif "PrecioVenta" not in headers:
                # Migrar estructura sin PrecioVenta
                new_ws = wb.create_sheet(INV_SHEET_PRODUCTS + "_new")
                new_ws.append(["Codigo", "Nombre", "Categoria", "Unidad", "TipoUso", "PrecioVenta", "Activo", "Observaciones"])
                dropping = wb[INV_SHEET_PRODUCTS]
                for row in dropping.iter_rows(min_row=2):
                    if row[0].value:
                        new_ws.append([
                            row[0].value,  # Codigo
                            row[1].value if len(row) > 1 else "",  # Nombre
                            row[2].value if len(row) > 2 else "",  # Categoria
                            row[3].value if len(row) > 3 else "pc",  # Unidad
                            row[4].value if len(row) > 4 else "Alquiler",  # TipoUso
                            0.0,  # PrecioVenta por defecto
                            row[5].value if len(row) > 5 else "SI",  # Activo
                            row[6].value if len(row) > 6 else ""  # Observaciones
                        ])
                wb.remove(dropping)
                new_ws.title = INV_SHEET_PRODUCTS
                ws = wb[INV_SHEET_PRODUCTS]
            
            updated = False
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if str(row[0].value).strip() == prod["Codigo"]:
                    ws.cell(row=idx, column=2).value = prod["Nombre"]
                    ws.cell(row=idx, column=3).value = prod["Categoria"]
                    ws.cell(row=idx, column=4).value = prod["Unidad"]
                    ws.cell(row=idx, column=5).value = prod["TipoUso"]
                    ws.cell(row=idx, column=6).value = prod.get("PrecioVenta", 0.0)
                    ws.cell(row=idx, column=7).value = prod["Activo"]
                    ws.cell(row=idx, column=8).value = prod["Observaciones"]
                    updated = True
                    break
            if not updated:
                ws.append([
                    prod["Codigo"], prod["Nombre"], prod["Categoria"], prod["Unidad"],
                    prod["TipoUso"], prod.get("PrecioVenta", 0.0), prod["Activo"], prod["Observaciones"]
                ])
            if safe_save_workbook(wb, EXCEL_PATH):
                wb.close()
                return True
            wb.close()
        except Exception as e:
            st.error(f"Error guardando producto: {e}")
        return False

    def load_providers():
        provs = []
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_PROVS in wb.sheetnames:
                ws = wb[INV_SHEET_PROVS]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        provs.append({
                            "ProveedorID": str(row[0]).strip(),
                            "Nombre": str(row[1]).strip() if row[1] else "",
                            "Tipo": str(row[2]).strip() if row[2] else "",
                            "Contacto": str(row[3]).strip() if row[3] else "",
                            "Telefono": str(row[4]).strip() if row[4] else "",
                            "Email": str(row[5]).strip() if row[5] else "",
                            "Material": str(row[6]).strip() if len(row) > 6 and row[6] else "",
                            "Costo": float(row[7]) if len(row) > 7 and row[7] else 0.0,
                            "Observaciones": str(row[8]).strip() if len(row) > 8 and row[8] else "",
                        })
            wb.close()
        except Exception:
            pass
        return provs

    def load_links():
        links = []
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_LINKS in wb.sheetnames:
                ws = wb[INV_SHEET_LINKS]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        links.append({
                            "Codigo": str(row[0]).strip(),
                            "ProveedorID": str(row[1]).strip() if row[1] else "",
                            "Rol": str(row[2]).strip() if row[2] else "",
                            "CostoUnit": float(row[3]) if row[3] else 0.0,
                            "LeadTimeDias": float(row[4]) if row[4] else 0.0,
                            "Observaciones": str(row[5]).strip() if len(row) > 5 and row[5] else "",
                        })
            wb.close()
        except Exception:
            pass
        return links

    def upsert_link(link):
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_LINKS not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_LINKS)
                ws.append(["Codigo", "ProveedorID", "Rol", "CostoUnit", "LeadTimeDias", "Observaciones"])
            ws = wb[INV_SHEET_LINKS]
            updated = False
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if str(row[0].value).strip() == link["Codigo"] and str(row[1].value).strip() == link["ProveedorID"] and str(row[2].value).strip() == link["Rol"]:
                    ws.cell(row=idx, column=4).value = link["CostoUnit"]
                    ws.cell(row=idx, column=5).value = link["LeadTimeDias"]
                    ws.cell(row=idx, column=6).value = link["Observaciones"]
                    updated = True
                    break
            if not updated:
                ws.append([link["Codigo"], link["ProveedorID"], link["Rol"], link["CostoUnit"], link["LeadTimeDias"], link["Observaciones"]])
            if safe_save_workbook(wb, EXCEL_PATH):
                wb.close()
                return True
            wb.close()
        except Exception as e:
            st.error(f"Error guardando vínculo: {e}")
        return False

    def upsert_provider(p):
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_PROVS not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_PROVS)
                ws.append(["ProveedorID", "Nombre", "Tipo", "Contacto", "Telefono", "Email", "Material", "Costo", "Observaciones"])
            ws = wb[INV_SHEET_PROVS]
            updated = False
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if str(row[0].value).strip() == p["ProveedorID"]:
                    ws.cell(row=idx, column=2).value = p["Nombre"]
                    ws.cell(row=idx, column=3).value = p["Tipo"]
                    ws.cell(row=idx, column=4).value = p["Contacto"]
                    ws.cell(row=idx, column=5).value = p["Telefono"]
                    ws.cell(row=idx, column=6).value = p["Email"]
                    ws.cell(row=idx, column=7).value = p.get("Material", "")
                    ws.cell(row=idx, column=8).value = p.get("Costo", 0.0)
                    ws.cell(row=idx, column=9).value = p.get("Observaciones", "")
                    updated = True
                    break
            if not updated:
                ws.append([p["ProveedorID"], p["Nombre"], p["Tipo"], p["Contacto"], p["Telefono"], p["Email"], p.get("Material", ""), p.get("Costo", 0.0), p.get("Observaciones", "")])
            if safe_save_workbook(wb, EXCEL_PATH):
                wb.close()
                return True
            wb.close()
        except Exception as e:
            st.error(f"Error guardando proveedor: {e}")
        return False

    def load_manual_costs():
        costs = []
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_COSTS in wb.sheetnames:
                ws = wb[INV_SHEET_COSTS]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[1]:
                        costs.append({
                            "Fecha": str(row[0]) if row[0] else "",
                            "Codigo": str(row[1]).strip(),
                            "CantidadProducida": float(row[2]) if row[2] else 0.0,
                            "CostoTotal": float(row[3]) if row[3] else 0.0,
                            "CostoUnitario": float(row[4]) if row[4] else 0.0,
                            "DetalleJSON": str(row[5]) if len(row) > 5 and row[5] else "",
                        })
            wb.close()
        except Exception:
            pass
        return costs

    def append_manual_cost(cost_row):
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_COSTS not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_COSTS)
                ws.append(["Fecha", "Codigo", "CantidadProducida", "CostoTotal", "CostoUnitario", "DetalleJSON"]) 
            ws = wb[INV_SHEET_COSTS]
            ws.append([
                cost_row.get("Fecha", ""),
                cost_row.get("Codigo", ""),
                cost_row.get("CantidadProducida", 0.0),
                cost_row.get("CostoTotal", 0.0),
                cost_row.get("CostoUnitario", 0.0),
                cost_row.get("DetalleJSON", ""),
            ])
            if safe_save_workbook(wb, EXCEL_PATH):
                wb.close()
                return True
            wb.close()
        except Exception as e:
            st.error(f"Error guardando costo de producto: {e}")
        return False

    def get_unit_cost_by_code():
        """Devuelve un mapa de código -> último costo unitario"""
        cost_map = {}
        try:
            costs = load_manual_costs()
            for c in costs:
                cod = c.get("Codigo", "")
                if cod:
                    cost_map[cod] = c.get("CostoUnitario", 0.0)
        except Exception:
            pass
        return cost_map

    def calc_stock_map():
        stock_by_code = {}
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            # base: stock inicial de productos
            if INV_SHEET_PRODUCTS in wb.sheetnames:
                for row in wb[INV_SHEET_PRODUCTS].iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        code = str(row[0]).strip()
                        stock_by_code[code] = float(row[5]) if row[5] else 0.0
            # movimientos
            if INV_SHEET_MOVS in wb.sheetnames:
                for row in wb[INV_SHEET_MOVS].iter_rows(min_row=2, values_only=True):
                    if row and row[1] and row[2] and row[3]:
                        t = str(row[1]).strip().upper()
                        code = str(row[2]).strip()
                        qty = float(row[3]) if row[3] else 0.0
                        if code not in stock_by_code:
                            stock_by_code[code] = 0.0
                        if t == "ENTRADA" or t == "AJUSTE+":
                            stock_by_code[code] += qty
                        elif t == "SALIDA" or t == "AJUSTE-":
                            stock_by_code[code] -= qty
            wb.close()
        except Exception:
            pass
        return stock_by_code

    def append_movement(mov):
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if INV_SHEET_MOVS not in wb.sheetnames:
                ws = wb.create_sheet(INV_SHEET_MOVS)
                ws.append(["Fecha", "Tipo", "Codigo", "Cantidad", "CostoUnit", "Proveedor", "Proceso", "Nota"])
            ws = wb[INV_SHEET_MOVS]
            ws.append([mov["Fecha"], mov["Tipo"], mov["Codigo"], mov["Cantidad"], mov["CostoUnit"], mov["Proveedor"], mov["Proceso"], mov["Nota"]])
            if safe_save_workbook(wb, EXCEL_PATH):
                wb.close()
                return True
            wb.close()
        except Exception as e:
            st.error(f"Error guardando movimiento: {e}")
        return False

    # Asegurar hojas
    ensure_inventory_sheets()

    tab_prod, tab_prov, tab_movs, tab_reportes = st.tabs(["📇 Productos", "🏭 Proveedores", "🚚 Movimientos", "📊 Reportes"])

    with tab_prod:
        st.markdown("#### 📇 Crear productos")
        with st.form("form_inv_product", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            code = c1.text_input("Codigo *", key="inv_p_code")
            name = c2.text_input("Nombre *", key="inv_p_name")
            category = c3.text_input("Categoria", key="inv_p_cat")
            c4, c5, c6 = st.columns(3)
            unit = c4.selectbox("Unidad", ["pc", "par", "set"], index=0, key="inv_p_unit")
            tipo_uso = c5.selectbox("Tipo de Uso", ["Alquiler", "Obsequios"], index=0, key="inv_p_tipo_uso")
            precio_venta = c6.number_input("Precio de Venta", min_value=0.0, step=0.01, key="inv_p_precio", value=0.0, format="%.2f")
            c7 = st.columns(1)[0]
            activo = c7.selectbox("Activo", ["SI", "NO"], index=0, key="inv_p_activo")
            obs = st.text_area("Observaciones", key="inv_p_obs")
            
            submit_p = st.form_submit_button("💾 Guardar Producto", use_container_width=True)
            if submit_p:
                if not code or not name:
                    st.error("Completa Codigo y Nombre")
                else:
                    prod = {"Codigo": code.strip(), "Nombre": name.strip(), "Categoria": category.strip(), "Unidad": unit, "TipoUso": tipo_uso, "PrecioVenta": precio_venta, "Activo": activo, "Observaciones": obs}
                    if upsert_product(prod):
                        st.success("Producto guardado")
                        st.rerun()

    with tab_prov:
        st.markdown("#### 🏭 Guardar Proveedores")
        with st.form("form_inv_prov", clear_on_submit=True):
            p1, p2, p3 = st.columns(3)
            pid = p1.text_input("Razón social *", key="inv_ruc")
            pname = p2.text_input("Nombre *", key="inv_prov_name")
            ptype = p3.text_input("Tipo", key="inv_prov_tipo")
            p4, p5, p6 = st.columns(3)
            pcontact = p4.text_input("Contacto", key="inv_prov_contact")
            ptel = p5.text_input("Telefono", key="inv_prov_tel")
            pemail = p6.text_input("Email", key="inv_prov_email")
            pmaterial = st.text_input("Material o Servicio", key="inv_prov_material")
            pcosto = st.number_input("Costo del Material o Mano de Obra", min_value=0.0, step=0.01, key="inv_prov_costo", value=0.0)
            pobs = st.text_area("Observaciones", key="inv_prov_obs")
            submit_pr = st.form_submit_button("💾 Guardar Proveedor", use_container_width=True)
            if submit_pr:
                if not pid or not pname:
                    st.error("Razón social y Nombre son obligatorios")
                else:
                    prov = {"ProveedorID": pid.strip(), "Nombre": pname.strip(), "Tipo": ptype, "Contacto": pcontact, "Telefono": ptel, "Email": pemail, "Material": pmaterial.strip(), "Costo": round(pcosto, 2), "Observaciones": pobs}
                    if upsert_provider(prov):
                        st.success("Proveedor guardado")
                        st.rerun()

        st.markdown("#### 📋 Proveedores Guardados")
        proveedores_list = load_providers()
        if proveedores_list:
            df = pd.DataFrame(proveedores_list)
            st.dataframe(df, use_container_width=True, hide_index=True)
        else:
            st.info("No hay proveedores registrados")

        st.markdown("#### 🧮 Crear costo de producto")
        productos = load_products()
        provs = load_providers()
        cods = [p.get("Codigo", "") for p in productos]
        prov_ids = [p.get("ProveedorID", "") for p in provs]
        prov_id_to_name = {p.get("ProveedorID", ""): p.get("Nombre", "") for p in provs}

        with st.form("form_costeo_producto", clear_on_submit=True):
            c1, c2 = st.columns([2, 1])
            codigo_sel = c1.selectbox("Producto (Código)", cods, key="cost_prod_codigo")
            qty_prod = c2.number_input("Cantidad producida", min_value=1.0, step=1.0, value=1.0, key="cost_qty")

            st.markdown("##### Partidas de costo por proveedor")
            num_items = st.number_input("Nº de partidas", min_value=1, max_value=10, step=1, value=4, key="cost_items_num")

            partidas = []
            total = 0.0
            for i in range(int(num_items)):
                r1, r2, r3 = st.columns([2, 3, 2])
                prov_i = r1.selectbox(f"Proveedor {i+1}", prov_ids, key=f"cost_p_{i}")
                desc_i = r2.text_input(f"Concepto {i+1}", key=f"cost_desc_{i}")
                cost_i = r3.number_input("COSTO TOTAL", min_value=0.0, step=0.01, key=f"cost_val_{i}")
                partidas.append({
                    "ProveedorID": prov_i,
                    "ProveedorNombre": prov_id_to_name.get(prov_i, prov_i),
                    "Concepto": desc_i,
                    "Costo": cost_i,
                })
                total += float(cost_i)

            costo_unit = total / qty_prod if qty_prod > 0 else 0.0
            st.info(f"Costo total: ${total:,.2f} | Costo unitario: ${costo_unit:,.2f}")

            submit_cost = st.form_submit_button("💾 Guardar costo de producto", use_container_width=True)
            if submit_cost:
                detalle = [{"ProveedorID": x["ProveedorID"], "Concepto": x["Concepto"], "Costo": round(x["Costo"], 2)} for x in partidas if x["Costo"] > 0]
                total = round(total, 2)
                costo_unit = round(costo_unit, 2)
                row = {
                    "Fecha": datetime.now().strftime("%Y-%m-%d"),
                    "Codigo": codigo_sel,
                    "CantidadProducida": qty_prod,
                    "CostoTotal": total,
                    "CostoUnitario": costo_unit,
                    "DetalleJSON": json.dumps(detalle, ensure_ascii=False),
                }
                if append_manual_cost(row):
                    st.success("Costo guardado")
                    st.rerun()

    with tab_movs:
        st.markdown("#### 🚚 Registro de Movimientos")
        productos = load_products()
        codes = [p["Codigo"] for p in productos]
        stock_map = calc_stock_map()
        with st.form("form_inv_mov", clear_on_submit=True):
            m1, m2, m3 = st.columns(3)
            tipo = m1.selectbox("Tipo *", ["ENTRADA", "SALIDA", "AJUSTE+", "AJUSTE-"], key="inv_m_tipo")
            codigo = m2.selectbox("Codigo *", codes, key="inv_m_codigo")
            cantidad = m3.number_input("Cantidad *", min_value=0.0, step=1.0, key="inv_m_cant")
            m4, m5 = st.columns(2)
            costo = m4.number_input("Costo Unitario", min_value=0.0, step=0.01, key="inv_m_costo")
            proceso = m5.selectbox("Proceso", ["Compra", "Bordado", "Impresión", "Otro"], key="inv_m_proc")
            nota = st.text_area("Nota", key="inv_m_nota")
            submit_m = st.form_submit_button("💾 Guardar Movimiento", use_container_width=True)
            if submit_m:
                if not codigo or cantidad <= 0:
                    st.error("Selecciona producto y cantidad válida")
                else:
                    stock_actual = stock_map.get(codigo, 0.0)
                    delta = cantidad if tipo in ["ENTRADA", "AJUSTE+"] else -cantidad
                    if stock_actual + delta < 0:
                        st.error(f"Stock insuficiente. Actual: {stock_actual}")
                    else:
                        mov = {
                            "Fecha": datetime.now().strftime("%Y-%m-%d"),
                            "Tipo": tipo,
                            "Codigo": codigo,
                            "Cantidad": cantidad,
                            "CostoUnit": round(costo, 2),
                            "Proveedor": "",
                            "Proceso": proceso,
                            "Nota": nota
                        }
                        if append_movement(mov):
                            st.success("Movimiento guardado")
                            st.rerun()

        st.markdown("#### 📈 Stock (después de movimientos)")
        productos = load_products()
        stock_map = calc_stock_map()
        if productos:
            df = pd.DataFrame([{"Codigo": p["Codigo"], "Nombre": p["Nombre"], "Stock": stock_map.get(p["Codigo"], 0.0)} for p in productos])
            st.dataframe(df, use_container_width=True, hide_index=True)
        else:
            st.info("No hay productos")

    with tab_reportes:
        st.markdown("#### 📦 Stock Actual con Costos")
        productos = load_products()
        stock_map = calc_stock_map()
        cost_map = get_unit_cost_by_code()
        
        # Filtros
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            buscador = st.text_input("Buscar por Código o Nombre", key="inv_r_search")
        with col_f2:
            filtro_tipo_uso = st.selectbox("Filtrar por Tipo de Uso", ["Todos", "Alquiler", "Obsequios"], key="inv_r_tipo_uso")
        
        if productos:
            df_stock = pd.DataFrame([{**p, "StockActual": stock_map.get(p["Codigo"], 0.0), "CostoUnitario": cost_map.get(p["Codigo"], 0.0)} for p in productos])
            
            # Aplicar filtro de búsqueda
            if buscador:
                q = str(buscador).strip().lower()
                df_stock = df_stock[df_stock.apply(lambda r: q in str(r["Codigo"]).lower() or q in str(r["Nombre"]).lower(), axis=1)]
            
            # Aplicar filtro de Tipo de Uso
            if filtro_tipo_uso != "Todos":
                df_stock = df_stock[df_stock["TipoUso"].fillna("Alquiler") == filtro_tipo_uso]
            
            st.dataframe(df_stock, use_container_width=True, hide_index=True)
            # Botón de descarga
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_stock.to_excel(writer, sheet_name='Stock Actual', index=False)
            excel_stock = output.getvalue()
            st.download_button(
                label="📥 Descargar Stock Actual como Excel",
                data=excel_stock,
                file_name=f"stock_actual_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("No hay productos registrados")

        st.markdown("#### 🏭 Búsqueda de Proveedores")
        proveedores_list = load_providers()
        buscar_prov = st.text_input("Buscar por Razón Social", key="inv_r_search_prov")
        if proveedores_list:
            df_prov = pd.DataFrame(proveedores_list)
            if buscar_prov:
                q = str(buscar_prov).strip().lower()
                df_prov = df_prov[df_prov.apply(lambda r: q in str(r.get("ProveedorID", "")).lower() or q in str(r.get("Nombre", "")).lower(), axis=1)]
            st.dataframe(df_prov, use_container_width=True, hide_index=True)
            # Botón de descarga
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_prov.to_excel(writer, sheet_name='Proveedores', index=False)
            excel_prov = output.getvalue()
            st.download_button(
                label="📥 Descargar Proveedores como Excel",
                data=excel_prov,
                file_name=f"proveedores_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("No hay proveedores registrados")
    

# ===== SECCIÓN REPORTES GENERALES =====
elif selected_menu == "📊 REPORTES GENERALES":
    st.markdown("### 📊 REPORTES GENERALES")
    
    # Función para cargar entregas de publicidad (si no está definida globalmente)
    def cargar_entregas_publicidad_reporte():
        """Cargar todas las entregas de publicidad desde Excel"""
        entregas = []
        try:
            wb = safe_load_workbook(EXCEL_PATH)
            if "PUBLICIDAD" in wb.sheetnames:
                ws = wb["PUBLICIDAD"]
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0] and row[1]:  # ID y Cliente deben existir
                        entrega = {
                            'ID': row[0],
                            'Cliente': str(row[1]).strip() if row[1] else '',
                            'Comercial': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                            'Fecha': str(row[3]).strip() if len(row) > 3 and row[3] else '',
                            'Productos': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                            'Observaciones': str(row[5]).strip() if len(row) > 5 and row[5] else ''
                        }
                        entregas.append(entrega)
            wb.close()
        except Exception as e:
            print(f"Error al cargar entregas de publicidad: {e}")
        return entregas
    
    # Cargar todos los datos
    entregas_publicidad = cargar_entregas_publicidad_reporte()
    perchas = leer_perchas()
    eventos = leer_eventos()
    
    # Cargar entregas comerciales a clientes
    todas_entregas_comerciales = leer_entregas_comerciales()
    entregas_comerciales_a_clientes = [e for e in todas_entregas_comerciales if e.get('Tipo Entrega') == 'ENTREGA COMERCIAL A CLIENTE']
    
    # Filtros
    st.markdown("#### 🔍 Filtros de Búsqueda")
    col_filtro1, col_filtro2, col_filtro3 = st.columns(3)
    
    with col_filtro1:
        # Filtro por cliente
        todos_clientes = set()
        todos_clientes.update([e['Cliente'] for e in entregas_publicidad if e.get('Cliente')])
        todos_clientes.update([p['Cliente'] for p in perchas if p.get('Cliente')])
        todos_clientes.update([e['Cliente'] for e in eventos if e.get('Cliente')])
        todos_clientes.update([e.get('Cliente Destino', '') for e in entregas_comerciales_a_clientes if e.get('Cliente Destino')])
        clientes_disponibles = sorted(list(todos_clientes))
        cliente_filtro = st.selectbox(
            "Filtrar por Cliente",
            options=["Todos"] + clientes_disponibles,
            key="reportes_generales_cliente"
        )
    
    with col_filtro2:
        # Filtro por comercial/vendedor
        todos_comerciales = set()
        todos_comerciales.update([e.get('Comercial', '') for e in entregas_publicidad if e.get('Comercial')])
        todos_comerciales.update([p.get('Comercial/Vendedor', '') for p in perchas if p.get('Comercial/Vendedor')])
        todos_comerciales.update([e.get('Comercial/Agente', '') for e in eventos if e.get('Comercial/Agente')])
        todos_comerciales.update([e.get('Comercial/Vendedor', '') for e in entregas_comerciales_a_clientes if e.get('Comercial/Vendedor')])
        comerciales_disponibles = sorted([c for c in todos_comerciales if c])
        comercial_filtro = st.selectbox(
            "Filtrar por Comercial/Vendedor",
            options=["Todos"] + comerciales_disponibles,
            key="reportes_generales_comercial"
        )
    
    with col_filtro3:
        # Filtro por tipo de entrega
        tipo_entrega_filtro = st.selectbox(
            "Filtrar por Tipo",
            options=["Todos", "Publicidad", "Perchas", "Eventos", "Entregas Comerciales"],
            key="reportes_generales_tipo"
        )
    
    # Filtros de fecha
    col_fecha1, col_fecha2 = st.columns(2)
    with col_fecha1:
        fecha_desde = st.date_input(
            "Fecha Desde",
            value=None,
            key="reportes_generales_fecha_desde"
        )
    with col_fecha2:
        fecha_hasta = st.date_input(
            "Fecha Hasta",
            value=None,
            key="reportes_generales_fecha_hasta"
        )
    
    # Aplicar filtros
    entregas_filtradas = entregas_publicidad.copy()
    perchas_filtradas = perchas.copy()
    eventos_filtrados = eventos.copy()
    entregas_comerciales_filtradas = entregas_comerciales_a_clientes.copy()
    
    # Filtrar por cliente
    if cliente_filtro != "Todos":
        entregas_filtradas = [e for e in entregas_filtradas if e.get('Cliente') == cliente_filtro]
        perchas_filtradas = [p for p in perchas_filtradas if p.get('Cliente') == cliente_filtro]
        eventos_filtrados = [e for e in eventos_filtrados if e.get('Cliente') == cliente_filtro]
        entregas_comerciales_filtradas = [e for e in entregas_comerciales_filtradas if e.get('Cliente Destino', '') == cliente_filtro]
    
    # Filtrar por comercial
    if comercial_filtro != "Todos":
        entregas_filtradas = [e for e in entregas_filtradas if e.get('Comercial') == comercial_filtro]
        perchas_filtradas = [p for p in perchas_filtradas if p.get('Comercial/Vendedor') == comercial_filtro]
        eventos_filtrados = [e for e in eventos_filtrados if e.get('Comercial/Agente') == comercial_filtro]
        entregas_comerciales_filtradas = [e for e in entregas_comerciales_filtradas if e.get('Comercial/Vendedor', '') == comercial_filtro]
    
    # Filtrar por tipo
    if tipo_entrega_filtro != "Todos":
        if tipo_entrega_filtro == "Publicidad":
            perchas_filtradas = []
            eventos_filtrados = []
            entregas_comerciales_filtradas = []
        elif tipo_entrega_filtro == "Perchas":
            entregas_filtradas = []
            eventos_filtrados = []
            entregas_comerciales_filtradas = []
        elif tipo_entrega_filtro == "Eventos":
            entregas_filtradas = []
            perchas_filtradas = []
            entregas_comerciales_filtradas = []
        elif tipo_entrega_filtro == "Entregas Comerciales":
            entregas_filtradas = []
            perchas_filtradas = []
            eventos_filtrados = []
    
    # Filtrar por fechas
    if fecha_desde:
        fecha_desde_str = fecha_desde.strftime("%Y-%m-%d")
        entregas_filtradas = [e for e in entregas_filtradas if e.get('Fecha') and e.get('Fecha') >= fecha_desde_str]
        perchas_filtradas = [p for p in perchas_filtradas if p.get('Fecha') and p.get('Fecha') >= fecha_desde_str]
        eventos_filtrados = [e for e in eventos_filtrados if e.get('Fecha') and e.get('Fecha') >= fecha_desde_str]
        entregas_comerciales_filtradas = [e for e in entregas_comerciales_filtradas if e.get('Fecha') and e.get('Fecha') >= fecha_desde_str]
    
    if fecha_hasta:
        fecha_hasta_str = fecha_hasta.strftime("%Y-%m-%d")
        entregas_filtradas = [e for e in entregas_filtradas if e.get('Fecha') and e.get('Fecha') <= fecha_hasta_str]
        perchas_filtradas = [p for p in perchas_filtradas if p.get('Fecha') and p.get('Fecha') <= fecha_hasta_str]
        eventos_filtrados = [e for e in eventos_filtrados if e.get('Fecha') and e.get('Fecha') <= fecha_hasta_str]
        entregas_comerciales_filtradas = [e for e in entregas_comerciales_filtradas if e.get('Fecha') and e.get('Fecha') <= fecha_hasta_str]
    
    # Botón para limpiar filtros
    if st.button("🔄 Limpiar Filtros", key="limpiar_filtros_reportes", use_container_width=False):
        if "reportes_generales_cliente" in st.session_state:
            del st.session_state["reportes_generales_cliente"]
        if "reportes_generales_comercial" in st.session_state:
            del st.session_state["reportes_generales_comercial"]
        if "reportes_generales_tipo" in st.session_state:
            del st.session_state["reportes_generales_tipo"]
        if "reportes_generales_fecha_desde" in st.session_state:
            del st.session_state["reportes_generales_fecha_desde"]
        if "reportes_generales_fecha_hasta" in st.session_state:
            del st.session_state["reportes_generales_fecha_hasta"]
        st.rerun()
    
    st.markdown("---")
    
    # Resumen total
    total_entregas = len(entregas_filtradas) + len(perchas_filtradas) + len(eventos_filtrados) + len(entregas_comerciales_filtradas)
    st.info(f"📋 **Total de entregas encontradas:** {total_entregas} (Publicidad: {len(entregas_filtradas)}, Perchas: {len(perchas_filtradas)}, Eventos: {len(eventos_filtrados)}, Entregas Comerciales: {len(entregas_comerciales_filtradas)})")
    
    # Filtrar solo los finalizados para la descarga de Excel
    # Publicidad: todas están finalizadas (no tienen estados)
    entregas_finalizadas_pub = entregas_filtradas.copy()
    # Perchas: solo las que están en estado "ENTREGADO"
    perchas_finalizadas = [p for p in perchas_filtradas if str(p.get("Estado", "")).strip().upper() == "ENTREGADO"]
    # Eventos: solo los que están en estado "REALIZADO"
    eventos_finalizados = [e for e in eventos_filtrados if str(e.get("Estado", "")).strip().upper() == "REALIZADO"]
    # Entregas comerciales: solo las que están en estado "ENTREGADO"
    entregas_comerciales_finalizadas = [e for e in entregas_comerciales_filtradas if str(e.get("Estado", "")).strip().upper() == "ENTREGADO"]
    
    total_finalizados = len(entregas_finalizadas_pub) + len(perchas_finalizadas) + len(eventos_finalizados) + len(entregas_comerciales_finalizadas)
    
    # Botón de descarga de Excel (solo finalizados)
    if total_finalizados > 0:
        # Preparar datos para Excel (solo finalizados)
        excel_buffer = BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Hoja de Resumen
            resumen_data = {
                'Métrica': ['Total Entregas Finalizadas', 'Publicidad', 'Perchas Entregadas', 'Eventos Realizados', 'Entregas Comerciales'],
                'Cantidad': [total_finalizados, len(entregas_finalizadas_pub), len(perchas_finalizadas), len(eventos_finalizados), len(entregas_comerciales_finalizadas)]
            }
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, index=False, sheet_name='Resumen')
            
            # Hoja de Publicidad (todas finalizadas)
            if entregas_finalizadas_pub:
                datos_publicidad = []
                for entrega in entregas_finalizadas_pub:
                    datos_publicidad.append({
                        'ID': entrega.get('ID', ''),
                        'Cliente': entrega.get('Cliente', ''),
                        'Comercial/Vendedor': entrega.get('Comercial', ''),
                        'Fecha': entrega.get('Fecha', ''),
                        'Productos': entrega.get('Productos', ''),
                        'Observaciones': entrega.get('Observaciones', '')
                    })
                df_publicidad = pd.DataFrame(datos_publicidad)
                df_publicidad.to_excel(writer, index=False, sheet_name='Publicidad')
            
            # Hoja de Perchas (solo ENTREGADAS)
            if perchas_finalizadas:
                datos_perchas = []
                for percha in perchas_finalizadas:
                    datos_perchas.append({
                        'ID': percha.get('ID', ''),
                        'Cliente': percha.get('Cliente', ''),
                        'Comercial/Vendedor': percha.get('Comercial/Vendedor', ''),
                        'Fecha': percha.get('Fecha', ''),
                        'Estado': percha.get('Estado', ''),
                        'Cantidad de Compra': percha.get('Cantidad de Compra', ''),
                        'Productos': percha.get('Productos', ''),
                        'Observaciones': percha.get('Observaciones', ''),
                        'Número de Guía': percha.get('Número de Guía', ''),
                        'Medio de Envío': percha.get('Medio de Envío', '')
                    })
                df_perchas = pd.DataFrame(datos_perchas)
                df_perchas.to_excel(writer, index=False, sheet_name='Perchas')
            
            # Hoja de Eventos (solo REALIZADOS)
            if eventos_finalizados:
                datos_eventos = []
                for evento in eventos_finalizados:
                    datos_eventos.append({
                        'ID': evento.get('ID', ''),
                        'Cliente': evento.get('Cliente', ''),
                        'Tipo': evento.get('Tipo', ''),
                        'Fecha': evento.get('Fecha', ''),
                        'Estado': evento.get('Estado', ''),
                        'Descripción': evento.get('Descripcion', ''),
                        'Productos': evento.get('Productos', ''),
                        'Comercial/Agente': evento.get('Comercial/Agente', ''),
                        'Número de Guía': evento.get('Número de Guía', ''),
                        'Observaciones Envío': evento.get('Observaciones Envío', ''),
                        'Productos Retornados': evento.get('Productos Retornados', ''),
                        'Fecha Retorno': evento.get('Fecha Retorno', ''),
                        'Observaciones Retorno': evento.get('Observaciones Retorno', '')
                    })
                df_eventos = pd.DataFrame(datos_eventos)
                df_eventos.to_excel(writer, index=False, sheet_name='Eventos')
            
            # Hoja de Entregas Comerciales (solo ENTREGADAS)
            if entregas_comerciales_finalizadas:
                datos_comerciales = []
                for entrega in entregas_comerciales_finalizadas:
                    # Parsear productos entregados para separar cantidad, código y nombre
                    productos_entregados_str = entrega.get('Productos Entregados', '')
                    productos_lista = productos_entregados_str.split('\n') if productos_entregados_str else []
                    
                    if productos_lista:
                        for prod_line in productos_lista:
                            if prod_line.strip():
                                # Parsear el formato: "cantidad codigo - nombre"
                                cantidad = ""
                                codigo = ""
                                nombre = ""
                                
                                match = re.match(r"^(\d+)\s+([A-Z0-9_]+)\s*-\s*(.+)$", prod_line.strip())
                                if match:
                                    cantidad = match.group(1)
                                    codigo = match.group(2)
                                    nombre = match.group(3).strip()
                                else:
                                    # Si no coincide el formato, intentar otras variaciones
                                    partes = prod_line.strip().split(' ', 2)
                                    if len(partes) >= 2:
                                        cantidad = partes[0]
                                        resto = ' '.join(partes[1:])
                                        codigo_match = re.match(r"^([A-Z0-9_]+)\s*-\s*(.+)$", resto)
                                        if codigo_match:
                                            codigo = codigo_match.group(1)
                                            nombre = codigo_match.group(2).strip()
                                        else:
                                            nombre = resto
                                
                                datos_comerciales.append({
                                    'ID Entrega': entrega.get('ID', ''),
                                    'Vendedor': entrega.get('Comercial/Vendedor', ''),
                                    'Cliente': entrega.get('Cliente Destino', ''),
                                    'Fecha': entrega.get('Fecha', ''),
                                    'Estado': entrega.get('Estado', ''),
                                    'Cantidad': cantidad,
                                    'Código Producto': codigo,
                                    'Nombre Producto': nombre,
                                    'Observaciones': entrega.get('Observaciones', ''),
                                    'ID Entrega Original': entrega.get('ID_Entrega_Original', '')
                                })
                    else:
                        # Si no hay productos, agregar fila con datos generales
                        datos_comerciales.append({
                            'ID Entrega': entrega.get('ID', ''),
                            'Vendedor': entrega.get('Comercial/Vendedor', ''),
                            'Cliente': entrega.get('Cliente Destino', ''),
                            'Fecha': entrega.get('Fecha', ''),
                            'Estado': entrega.get('Estado', ''),
                            'Cantidad': '',
                            'Código Producto': '',
                            'Nombre Producto': '',
                            'Observaciones': entrega.get('Observaciones', ''),
                            'ID Entrega Original': entrega.get('ID_Entrega_Original', '')
                        })
                
                df_comerciales = pd.DataFrame(datos_comerciales)
                df_comerciales.to_excel(writer, index=False, sheet_name='Entregas Comerciales')
        
        excel_bytes = excel_buffer.getvalue()
        
        col_desc1, col_desc2 = st.columns([1, 3])
        with col_desc1:
            st.download_button(
                "⬇️ Descargar Reporte Excel (Solo Finalizados)",
                excel_bytes,
                file_name=f"reportes_generales_finalizados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_reportes_generales"
            )
        with col_desc2:
            st.caption(f"📊 Incluye: {len(entregas_finalizadas_pub)} Publicidad, {len(perchas_finalizadas)} Perchas Entregadas, {len(eventos_finalizados)} Eventos Realizados, {len(entregas_comerciales_finalizadas)} Entregas Comerciales")
    
    st.markdown("---")
    
    # Agrupar por cliente
    entregas_por_cliente = {}
    
    # Agregar entregas de publicidad
    for entrega in entregas_filtradas:
        cliente = entrega.get('Cliente', 'Sin Cliente')
        if cliente not in entregas_por_cliente:
            entregas_por_cliente[cliente] = {
                'publicidad': [],
                'perchas': [],
                'eventos': [],
                'entregas_comerciales': []
            }
        entregas_por_cliente[cliente]['publicidad'].append(entrega)
    
    # Agregar perchas
    for percha in perchas_filtradas:
        cliente = percha.get('Cliente', 'Sin Cliente')
        if cliente not in entregas_por_cliente:
            entregas_por_cliente[cliente] = {
                'publicidad': [],
                'perchas': [],
                'eventos': [],
                'entregas_comerciales': []
            }
        entregas_por_cliente[cliente]['perchas'].append(percha)
    
    # Agregar eventos
    for evento in eventos_filtrados:
        cliente = evento.get('Cliente', 'Sin Cliente')
        if cliente not in entregas_por_cliente:
            entregas_por_cliente[cliente] = {
                'publicidad': [],
                'perchas': [],
                'eventos': [],
                'entregas_comerciales': []
            }
        entregas_por_cliente[cliente]['eventos'].append(evento)
    
    # Agregar entregas comerciales a clientes
    for entrega_com in entregas_comerciales_filtradas:
        cliente = entrega_com.get('Cliente Destino', 'Sin Cliente')
        if cliente not in entregas_por_cliente:
            entregas_por_cliente[cliente] = {
                'publicidad': [],
                'perchas': [],
                'eventos': [],
                'entregas_comerciales': []
            }
        entregas_por_cliente[cliente]['entregas_comerciales'].append(entrega_com)
    
    # Mostrar entregas por cliente
    if entregas_por_cliente:
        st.markdown("### 📦 Entregas por Cliente")
        for cliente_nombre in sorted(entregas_por_cliente.keys()):
            cliente_data = entregas_por_cliente[cliente_nombre]
            total_cliente = len(cliente_data['publicidad']) + len(cliente_data['perchas']) + len(cliente_data['eventos']) + len(cliente_data.get('entregas_comerciales', []))
            
            with st.expander(f"📦 **{cliente_nombre}** - {total_cliente} entrega(s)", expanded=False):
                # Resumen del cliente
                col_res1, col_res2, col_res3, col_res4 = st.columns(4)
                col_res1.metric("Publicidad", len(cliente_data['publicidad']))
                col_res2.metric("Perchas", len(cliente_data['perchas']))
                col_res3.metric("Eventos", len(cliente_data['eventos']))
                col_res4.metric("Entregas Comerciales", len(cliente_data.get('entregas_comerciales', [])))
                
                st.markdown("---")
                
                # Mostrar entregas de publicidad
                if cliente_data['publicidad']:
                    st.markdown("#### 📰 Entregas de Publicidad")
                    for idx, entrega in enumerate(cliente_data['publicidad']):
                        st.markdown(f"**📰 Entrega #{entrega.get('ID', 'N/A')} - {entrega.get('Fecha', 'N/A')}**")
                        col1, col2 = st.columns(2)
                        col1.write(f"**ID:** {entrega.get('ID', 'N/A')}")
                        col1.write(f"**Fecha:** {entrega.get('Fecha', 'N/A')}")
                        col1.write(f"**Comercial/Vendedor:** {entrega.get('Comercial', 'N/A')}")
                        col2.write(f"**Productos:**")
                        productos_pub = entrega.get('Productos', '')
                        if productos_pub:
                            for prod_line in productos_pub.split('\n'):
                                if prod_line.strip():
                                    col2.write(f"- {prod_line.strip()}")
                        if entrega.get('Observaciones'):
                            st.write(f"**Observaciones:** {entrega.get('Observaciones')}")
                        if idx < len(cliente_data['publicidad']) - 1:
                            st.markdown("---")
                
                # Mostrar perchas
                if cliente_data['perchas']:
                    st.markdown("#### 📦 Perchas/Exhibidores")
                    for idx, percha in enumerate(cliente_data['perchas']):
                        st.markdown(f"**📦 Percha #{percha.get('ID', 'N/A')} - {percha.get('Fecha', 'N/A')} - {percha.get('Estado', 'N/A')}**")
                        col1, col2 = st.columns(2)
                        col1.write(f"**ID:** {percha.get('ID', 'N/A')}")
                        col1.write(f"**Fecha:** {percha.get('Fecha', 'N/A')}")
                        col1.write(f"**Comercial/Vendedor:** {percha.get('Comercial/Vendedor', 'N/A')}")
                        col1.write(f"**Estado:** {percha.get('Estado', 'N/A')}")
                        col2.write(f"**Cantidad de Compra:** {percha.get('Cantidad de Compra', 'N/A')}")
                        if percha.get('Número de Guía'):
                            col2.write(f"**Número de Guía:** {percha.get('Número de Guía')}")
                        if percha.get('Medio de Envío'):
                            col2.write(f"**Medio de Envío:** {percha.get('Medio de Envío')}")
                        if percha.get('Productos'):
                            st.write(f"**Productos:**")
                            for prod_line in percha.get('Productos', '').split('\n'):
                                if prod_line.strip():
                                    st.write(f"- {prod_line.strip()}")
                        if percha.get('Observaciones'):
                            st.write(f"**Observaciones:** {percha.get('Observaciones')}")
                        if idx < len(cliente_data['perchas']) - 1:
                            st.markdown("---")
                
                # Mostrar eventos
                if cliente_data['eventos']:
                    st.markdown("#### 🎉 Eventos")
                    for idx, evento in enumerate(cliente_data['eventos']):
                        st.markdown(f"**🎉 Evento #{evento.get('ID', 'N/A')} - {evento.get('Tipo', 'N/A')} - {evento.get('Fecha', 'N/A')}**")
                        col1, col2 = st.columns(2)
                        col1.write(f"**ID:** {evento.get('ID', 'N/A')}")
                        col1.write(f"**Tipo:** {evento.get('Tipo', 'N/A')}")
                        col1.write(f"**Fecha:** {evento.get('Fecha', 'N/A')}")
                        col1.write(f"**Estado:** {evento.get('Estado', 'N/A')}")
                        col2.write(f"**Comercial/Agente:** {evento.get('Comercial/Agente', 'N/A')}")
                        if evento.get('Descripcion'):
                            st.write(f"**Descripción:** {evento.get('Descripcion')}")
                        if evento.get('Productos'):
                            st.write(f"**Productos:**")
                            for prod_line in evento.get('Productos', '').split('\n'):
                                if prod_line.strip():
                                    st.write(f"- {prod_line.strip()}")
                        if evento.get('Número de Guía'):
                            st.write(f"**Número de Guía:** {evento.get('Número de Guía')}")
                        if evento.get('Observaciones Envío'):
                            st.write(f"**Observaciones Envío:** {evento.get('Observaciones Envío')}")
                        if idx < len(cliente_data['eventos']) - 1:
                            st.markdown("---")
                
                # Mostrar entregas comerciales a clientes
                if cliente_data.get('entregas_comerciales'):
                    st.markdown("#### 💼 Entregas Comerciales a Cliente")
                    for idx, entrega_com in enumerate(cliente_data['entregas_comerciales']):
                        st.markdown(f"**💼 Entrega #{entrega_com.get('ID', 'N/A')} - {entrega_com.get('Fecha', 'N/A')} - {entrega_com.get('Estado', 'N/A')}**")
                        col1, col2 = st.columns(2)
                        col1.write(f"**ID:** {entrega_com.get('ID', 'N/A')}")
                        col1.write(f"**Fecha:** {entrega_com.get('Fecha', 'N/A')}")
                        col1.write(f"**Vendedor:** {entrega_com.get('Comercial/Vendedor', 'N/A')}")
                        col1.write(f"**Estado:** {entrega_com.get('Estado', 'N/A')}")
                        col2.write(f"**Productos Entregados:**")
                        productos_com = entrega_com.get('Productos Entregados', '')
                        if productos_com:
                            for prod_line in productos_com.split('\n'):
                                if prod_line.strip():
                                    col2.write(f"- {prod_line.strip()}")
                        if entrega_com.get('Observaciones'):
                            st.write(f"**Observaciones:** {entrega_com.get('Observaciones')}")
                        if entrega_com.get('ID_Entrega_Original'):
                            st.write(f"**ID Entrega Original:** {entrega_com.get('ID_Entrega_Original')}")
                        if idx < len(cliente_data['entregas_comerciales']) - 1:
                            st.markdown("---")
    else:
        st.warning("⚠️ No se encontraron entregas con los filtros seleccionados. Intenta ajustar los filtros.")

# ===== FOOTER =====
st.markdown("---")
st.markdown("""
<div style='text-align: center; padding: 20px; color: #666;'>
    <p>🎯 <strong>Marketing Extrememax</strong> | Sistema de Gestión y Seguimiento de Clientes</p>
    <p style='font-size: 0.9em;'>© 2025 - Todos los derechos reservados</p>
    <p style='font-size: 0.85em; color: #FF9A00;'>Creador: Mario Ponce</p>
</div>
""", unsafe_allow_html=True)


